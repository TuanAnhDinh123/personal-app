"""Xuất danh sách ứng viên ra Excel theo bố cục sheet "Candidates".

Sheet được DỰNG THẲNG BẰNG CODE, không đọc file .xlsx mẫu nào. Bố cục của mẫu
(tiêu đề, font, bề rộng cột, dropdown, tô màu theo trạng thái) đều là dữ liệu
tĩnh nên mô tả bằng hằng số ở đây là đủ — mà lại nhanh hơn hẳn: file mẫu cũ mang
theo ~90 liên kết ngoài với 15 MB XML cache, openpyxl mất gần 9 giây chỉ để MỞ
nó; dựng mới xong trong vài chục mili giây và file kết quả chỉ vài chục KB.

Mọi ô đều là CHỮ / SỐ / NGÀY thuần — không công thức, không liên kết ngoài. Phần
"động" duy nhất là dropdown và tô màu điều kiện: cả hai chỉ so khớp chuỗi có
sẵn, Excel không phải tính gì.

Ghi NỐI TIẾP vào file đã xuất trước đó cũng qua đây (`open_existing`): file do
module này tạo ra nhẹ nên mở lại gần như tức thì.
"""
from __future__ import annotations

import datetime as _dt

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core import cv_schema

SHEET_NAME = "Candidates"
HEADER_ROW = 1
DATA_START_ROW = 2

# Sheet ẩn chứa các danh sách nguồn cho dropdown. Dropdown phải trỏ tới một VÙNG
# ô khi danh sách dài (chuỗi nhúng thẳng vào ô kiểm tra dữ liệu giới hạn 255 ký
# tự), nên mọi danh sách đều đặt ở đây cho nhất quán.
_LIST_SHEET = "_Lists"

# Số dòng TRỐNG phía dưới vùng dữ liệu vẫn được gắn dropdown, để HR nhập thêm
# ứng viên bằng tay mà vẫn có sẵn ô chọn.
_DV_SPARE_ROWS = 200

# Cột của sheet: (khóa, tiêu đề, bề rộng). Thứ tự trong list = thứ tự cột A→Y.
COLUMNS = [
    ("batch",           "Batch",                        9.0),
    ("id",              "ID",                          12.0),
    ("name",            "NAME",                        18.875),
    ("apply",           "APPLYING FOR",                21.5),
    ("source",          "SOURCE",                      18.875),
    ("email",           "EMAIL ADDRESS",               30.5),
    ("phone",           "PHONE",                       20.125),
    ("ai_score",        "Score (AI scan)",             12.0),
    # Các cột nhận xét (AI + 3 vòng phỏng vấn) rộng hơn mẫu cũ: nội dung là đoạn
    # văn nhiều dòng chứ không phải một nhãn ngắn.
    ("ai_eval",         "AI Evaluation",               45.0),
    ("status",          "STATUS",                      18.875),
    ("result",          "Results",                     18.875),
    ("ps_date",         "PHONE SCREEN DATE",           18.875),
    ("ps_note",         "PHONE SCREEN NOTE",           25.0),
    ("r1_date",         "1ST INTERVIEW DATE",          20.0),
    ("r1_interviewer",  "1ST ASSIGNED \nINTERVIEWER",  22.875),
    ("r1_eval",         "1ST INTERVIEW \nEVALUATION",  45.0),
    ("r1_result",       "1ST FINAL RESULT",            14.25),
    ("r2_date",         "2ND INTERVIEW DATE",          20.0),
    ("r2_interviewer",  "2ND ASSIGNED \nINTERVIEWER",  22.875),
    ("r2_eval",         "2ND INTERVIEW \nEVALUATION",  45.0),
    ("r2_result",       "2ND FINAL RESULT",            14.25),
    ("r3_date",         "3RD INTERVIEW DATE",          20.0),
    ("r3_interviewer",  "3RD ASSIGNED \nINTERVIEWER",  22.875),
    ("r3_eval",         "3RD INTERVIEW \nEVALUATION",  45.0),
    ("r3_result",       "3RD FINAL RESULT",            14.25),
]

# khóa cột → chỉ số cột (1-based)
COL = {key: i for i, (key, _t, _w) in enumerate(COLUMNS, start=1)}
LAST_COL = len(COLUMNS)

# Các cột dùng để xác định "dòng đã có dữ liệu" khi ghi nối tiếp.
_KEY_COLUMNS = ("id", "name", "apply", "email", "phone")

_DATE_COLUMNS = ("ps_date", "r1_date", "r2_date", "r3_date")
# ID và số điện thoại là CHUỖI: ép định dạng text để Excel không nuốt số 0 đầu.
_TEXT_COLUMNS = ("id", "phone")

_DATE_FORMAT = "mm/dd/yyyy"
_SCORE_FORMAT = "0"

# ── Style ────────────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FONT = Font(name="Verdana", size=10, bold=True)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_HEADER_HEIGHT = 32.1

_CELL_FONT = Font(name="Tahoma", size=10)
_CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Tô màu theo điều kiện ────────────────────────────────────────────────
# (nền, chữ) — cùng bảng màu Excel dùng cho Good / Bad / Neutral.
_GOOD = ("C6EFCE", "006100")
_BAD = ("FFC7CE", "9C0006")
_WARN = ("FFEB9C", "9C6500")

# Trạng thái đơn (cột STATUS): chỉ tô những mốc cần chú ý, không tô cả luồng.
_STATUS_COLORS = {
    "Ready To Hire":         _GOOD,
    "Offer Approval":        _GOOD,
    "Not Proceed":           _BAD,
    "Rejected Offer":        _BAD,
    "Fail Probation Period": _BAD,
}
_RESULT_COLORS = {
    "Pass":              _GOOD,
    "Fail":              _BAD,
    "Withdraw":          _BAD,
    "Not proceed":       _BAD,
    "Could not contact": _BAD,
    "Considering":       _WARN,
}
_ROUND_RESULT_COLORS = {
    "Pass":          _GOOD,
    "Fail":          _BAD,
    "Consideration": _WARN,
}


# ═══════════════════════════ dựng workbook ══════════════════════════════

def new_workbook(dropdowns=None, row_count=0):
    """Tạo workbook mới có sẵn sheet "Candidates" (tiêu đề + style + dropdown).

    `dropdowns` là dict phụ thêm cho các danh sách lấy từ DB — hiện dùng hai
    khóa "positions" (cột APPLYING FOR) và "interviewers" (cột ASSIGNED
    INTERVIEWER). Các danh sách cố định lấy thẳng từ `cv_schema`.

    `row_count` là số dòng sắp ghi: dropdown và tô màu phủ hết chỗ đó cộng thêm
    `_DV_SPARE_ROWS` dòng trống để HR nhập tay tiếp.

    Trả về (workbook, worksheet).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for idx, (_key, title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(HEADER_ROW, idx, title)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = _HEADER_HEIGHT
    ws.freeze_panes = ws.cell(DATA_START_ROW, 1).coordinate

    last_row = DATA_START_ROW + row_count + _DV_SPARE_ROWS
    _add_dropdowns(wb, ws, dropdowns or {}, last_row)
    _add_conditional_formats(ws, last_row)
    return wb, ws


def open_existing(path: str):
    """Mở file đã xuất trước đó để ghi nối tiếp; báo lỗi nếu sai mẫu.

    `keep_links=False`: file người dùng chọn có thể là bản cũ xuất từ template
    mang theo hàng chục liên kết ngoài với vài MB cache vô dụng — bỏ chúng đi
    thì mở/lưu nhanh hơn hàng chục lần, các sheet và style vẫn nguyên.
    """
    wb = openpyxl.load_workbook(path, keep_links=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"This Excel file has no '{SHEET_NAME}' sheet.\n"
            "Choose a file exported by this app, or pick a new file name.")
    return wb, wb[SHEET_NAME]


def next_empty_row(ws) -> int:
    """Dòng trống đầu tiên trong vùng dữ liệu ứng viên."""
    row = DATA_START_ROW
    while any(ws.cell(row, COL[k]).value not in (None, "") for k in _KEY_COLUMNS):
        row += 1
    return row


def write_rows(ws, rows) -> int:
    """Ghi các dict `rows` (khóa = khóa cột trong COLUMNS) nối tiếp vào sheet.

    Ô nào cũng được kẻ khung và gán font, kể cả ô để trống, để dòng thêm vào
    trông liền mạch với phần đã có. Trả về dòng cuối cùng đã ghi.
    """
    start = next_empty_row(ws)
    for i, record in enumerate(rows):
        row_no = start + i
        for key, _title, _w in COLUMNS:
            cell = ws.cell(row_no, COL[key])
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border = _BORDER
            if key in _DATE_COLUMNS:
                cell.number_format = _DATE_FORMAT
            elif key in _TEXT_COLUMNS:
                cell.number_format = "@"
            elif key == "ai_score":
                cell.number_format = _SCORE_FORMAT
            value = record.get(key)
            if value not in (None, ""):
                cell.value = value
    last = max(start + len(rows) - 1, HEADER_ROW)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(LAST_COL)}{last}"
    return last


def export(path: str, rows, dropdowns=None, append=False) -> int:
    """Ghi `rows` ra file Excel — tạo mới, hoặc nối tiếp khi `append`.

    Trả về số dòng đã ghi.
    """
    if append:
        wb, ws = open_existing(path)
    else:
        wb, ws = new_workbook(dropdowns, row_count=len(rows))
    last = write_rows(ws, rows)
    if append:
        _stretch_ranges(ws, last + _DV_SPARE_ROWS)
    wb.save(path)
    wb.close()
    return len(rows)


# ───────────────────────── dropdown & tô màu ─────────────────────────────

def _list_sources(extra: dict) -> dict:
    """khóa cột → danh sách giá trị hợp lệ của cột đó."""
    positions = [str(p) for p in extra.get("positions", []) if str(p).strip()]
    interviewers = [str(p) for p in extra.get("interviewers", []) if str(p).strip()]
    sources = {
        "apply":  positions,
        "source": cv_schema.CANDIDATE_SOURCE_CHOICES,
        "status": cv_schema.CANDIDATE_STATUS_CHOICES,
        "result": cv_schema.FINAL_STATUS_CHOICES,
    }
    for n in (1, 2, 3):
        sources[f"r{n}_interviewer"] = interviewers
        sources[f"r{n}_result"] = cv_schema.INTERVIEW_SCORE_CHOICES
    return {k: v for k, v in sources.items() if v}


def _add_dropdowns(wb, ws, extra: dict, last_row: int) -> None:
    """Gắn ô chọn (data validation) cho các cột có danh sách giá trị cố định.

    Danh sách nằm ở sheet ẩn `_Lists`, mỗi cột một danh sách. Không bật báo lỗi
    nhập sai: HR vẫn phải gõ được giá trị ngoài danh sách (nguồn ứng viên mới,
    người phỏng vấn là khách…).
    """
    sources = _list_sources(extra)
    if not sources:
        return

    lists = wb.create_sheet(_LIST_SHEET)
    lists.sheet_state = "hidden"

    # Danh sách giống nhau (3 vòng dùng chung người phỏng vấn / kết luận) chỉ ghi
    # MỘT cột ở sheet ẩn, các cột dữ liệu cùng trỏ về đó.
    ranges: dict[tuple, str] = {}
    for col_key, values in sources.items():
        fingerprint = tuple(values)
        ref = ranges.get(fingerprint)
        if ref is None:
            idx = len(ranges) + 1
            letter = get_column_letter(idx)
            lists.cell(1, idx, col_key)
            for j, value in enumerate(values, start=2):
                lists.cell(j, idx, value)
            ref = (f"'{_LIST_SHEET}'!${letter}$2:${letter}${len(values) + 1}")
            ranges[fingerprint] = ref

        dv = DataValidation(type="list", formula1=ref, allow_blank=True,
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        letter = get_column_letter(COL[col_key])
        dv.add(f"{letter}{DATA_START_ROW}:{letter}{last_row}")


def _text_rule(value: str, colors, priority: int) -> Rule:
    """Luật tô màu khi ô BẰNG ĐÚNG một chuỗi (so khớp thuần, không tính toán)."""
    bg, fg = colors
    style = DifferentialStyle(font=Font(color=fg),
                              fill=PatternFill(bgColor=bg, patternType="solid"))
    return Rule(type="cellIs", operator="equal", formula=[f'"{value}"'],
                dxf=style, stopIfTrue=True, priority=priority)


def _add_conditional_formats(ws, last_row: int) -> None:
    """Tô màu STATUS / Results / kết luận từng vòng, và bôi đỏ email trùng."""
    priority = 1
    palettes = [("status", _STATUS_COLORS), ("result", _RESULT_COLORS)]
    palettes += [(f"r{n}_result", _ROUND_RESULT_COLORS) for n in (1, 2, 3)]

    for col_key, palette in palettes:
        letter = get_column_letter(COL[col_key])
        span = f"{letter}{DATA_START_ROW}:{letter}{last_row}"
        for value, colors in palette.items():
            ws.conditional_formatting.add(span, _text_rule(value, colors, priority))
            priority += 1

    # Email trùng = ứng viên đã có trong danh sách → bôi đỏ để soát bằng mắt.
    bg, fg = _BAD
    letter = get_column_letter(COL["email"])
    ws.conditional_formatting.add(
        f"{letter}{DATA_START_ROW}:{letter}{last_row}",
        Rule(type="duplicateValues", priority=priority,
             dxf=DifferentialStyle(font=Font(color=fg),
                                   fill=PatternFill(bgColor=bg, patternType="solid"))))


def _stretch(ref: str, last_row: int) -> str:
    """Kéo dài một vùng ô một cột xuống tới `last_row` ('J2:J50' → 'J2:J400')."""
    head, _sep, tail = ref.partition(":")
    if not tail:
        return ref
    letter = tail.rstrip("0123456789")
    try:
        if int(tail[len(letter):]) >= last_row:
            return ref
    except ValueError:
        return ref
    return f"{head}:{letter}{last_row}"


def _stretch_ranges(ws, last_row: int) -> None:
    """Nới dropdown & tô màu xuống hết vùng dữ liệu sau khi ghi nối tiếp.

    File nối tiếp vốn chỉ phủ tới dòng cuối của lần xuất trước; thêm dòng mới mà
    không nới thì các dòng đó mất ô chọn và không được tô màu.
    """
    from openpyxl.formatting.formatting import ConditionalFormattingList
    from openpyxl.worksheet.cell_range import MultiCellRange

    for dv in ws.data_validations.dataValidation:
        dv.sqref = MultiCellRange(
            " ".join(_stretch(str(r), last_row) for r in dv.sqref.ranges))

    # `ConditionalFormatting` băm theo chính vùng ô của nó nên không sửa tại chỗ
    # được — dựng lại cả danh sách.
    saved = [(str(cf.sqref), list(rules))
             for cf, rules in ws.conditional_formatting._cf_rules.items()]
    ws.conditional_formatting = ConditionalFormattingList()
    for ref, rules in saved:
        span = " ".join(_stretch(part, last_row) for part in ref.split())
        for rule in rules:
            ws.conditional_formatting.add(span, rule)


# ═════════════════════ dữ liệu DB → dòng của sheet ══════════════════════

def dropdown_sources() -> dict:
    """Danh sách vị trí & người phỏng vấn hiện có trong DB (nguồn cho dropdown)."""
    from app.core import cv_repository as repo

    return {
        "positions": [p["position_title"] for p in repo.list_positions()
                      if p["position_title"]],
        "interviewers": [e["full_name"] for e in repo.list_interviewers()
                         if e["full_name"]],
    }


def rows_from_candidates(candidates) -> list[dict]:
    """Đổi các dòng ứng viên (`cv_repository.search_candidates`) → dòng của sheet.

    Phỏng vấn và nhận xét của CẢ NHÓM được lấy trong hai truy vấn thay vì hỏi
    lại DB cho từng người, nên xuất vài trăm ứng viên vẫn chỉ tốn vài chục ms.
    """
    from app.core import cv_repository as repo

    ids = [c["candidate_id"] for c in candidates]
    interviews = repo.list_interviews_by_candidates(ids)
    feedbacks = repo.list_feedbacks_by_interviews(
        [iv["interview_id"] for iv in interviews])

    by_interview: dict = {}
    for fb in feedbacks:
        by_interview.setdefault(fb["interview_id"], []).append(fb)

    # candidate_id → {vòng: buổi phỏng vấn}. Truy vấn đã sắp mới nhất trước nên
    # buổi gặp đầu tiên của mỗi vòng là buổi cần lấy.
    by_candidate: dict = {}
    for iv in interviews:
        rounds = by_candidate.setdefault(iv["candidate_id"], {})
        rounds.setdefault(iv["round"] or 1, iv)

    return [_candidate_row(c, by_candidate.get(c["candidate_id"], {}), by_interview)
            for c in candidates]


def _candidate_row(cand, rounds: dict, feedbacks_by_interview: dict) -> dict:
    """Một ứng viên (kèm các vòng phỏng vấn của họ) → dict theo khóa cột."""
    record = {
        "batch":    _int(_get(cand, "batch")),
        "id":       _candidate_code(cand),
        "name":     _text(_get(cand, "full_name")),
        # Cột "APPLYING FOR" hỏi ứng tuyển VỊ TRÍ nào; chưa gắn vị trí thì ghi
        # tạm bộ phận để dòng không trống trơn.
        "apply":    _text(_get(cand, "position_title")) or _text(_get(cand, "department_name")),
        "source":   _cv_source(cand),
        "email":    _text(_get(cand, "email")),
        "phone":    _text(_get(cand, "phone")),
        "ai_score": _int(_get(cand, "ai_score")),
        "ai_eval":  _ai_evaluation(cand),
        "status":   _text(_get(cand, "status")),
        "result":   _text(_get(cand, "final_status")),
        "ps_date":  _date(_get(cand, "phone_screen_date")),
        "ps_note":  _text(_get(cand, "application_note")),
    }
    for n in (1, 2, 3):
        interview = rounds.get(n)
        fbs = feedbacks_by_interview.get(interview["interview_id"], []) if interview else []
        record[f"r{n}_date"] = _date(_get(interview, "interview_date"))
        record[f"r{n}_interviewer"] = ", ".join(
            name for name in (_interviewer_name(fb) for fb in fbs) if name)
        record[f"r{n}_eval"] = _interview_evaluation(fbs)
        record[f"r{n}_result"] = _text(_get(interview, "overall_score"))
    return record


def _cv_source(cand) -> str:
    """Cột SOURCE = NƠI CUNG CẤP CV (Itviec, VietnamWorks, LinkedIn…).

    Nguồn của đơn ứng tuyển được ưu tiên hơn nguồn của ứng viên: cùng một người
    có thể nộp lần này qua sàn khác lần trước. Hồ sơ do tool AI CV Scan nạp vào
    mang dấu `CANDIDATE_SOURCE_AUTO` — đó là đường vào app chứ không phải sàn,
    nên bỏ trống để HR chọn lại từ ô chọn của cột.
    """
    for key in ("application_source", "source"):
        value = _text(_get(cand, key))
        if value and value != cv_schema.CANDIDATE_SOURCE_AUTO:
            return value
    return ""


def _ai_evaluation(cand) -> str:
    """Nhận xét của AI: chỉ lấy phần tóm tắt độ phù hợp của lượt chấm mới nhất."""
    return _text(_get(cand, "fit_summary"))


def _interview_evaluation(feedbacks) -> str:
    """Gộp nhận xét của MỌI người phỏng vấn trong một vòng vào một ô.

    Mỗi người một đoạn mở đầu bằng tên (kèm vai trò & kết luận nếu có), cách
    nhau bằng dòng trống.
    """
    blocks = []
    for fb in feedbacks:
        who = _interviewer_name(fb) or "(interviewer)"
        tags = " · ".join(t for t in (_text(_get(fb, "role")),
                                      _text(_get(fb, "score"))) if t)
        lines = [f"{who} ({tags}):" if tags else f"{who}:"]
        for label, key in (("", "feedback"), ("Strengths", "strengths"),
                           ("Weaknesses", "weaknesses")):
            value = _text(_get(fb, key))
            if value:
                lines.append(f"{label}: {value}" if label else value)
        blocks.append("\n".join(lines) if len(lines) > 1 else lines[0].rstrip(":"))
    return "\n\n".join(blocks)


def _interviewer_name(fb) -> str:
    return _text(_get(fb, "display_name")) or _text(_get(fb, "interviewer_name"))


def _candidate_code(cand) -> str:
    """Mã hiển thị ở cột ID: mã bóc từ tên file CV, không có thì dùng id trong DB."""
    from app.core.cv_scan import _split_id_name

    path = _text(_get(cand, "cv_file_path"))
    if path:
        stem = path.replace("\\", "/").rsplit("/", 1)[-1].rsplit(".", 1)[0]
        code, _name = _split_id_name(stem, [])
        if code:
            return code
    return str(_get(cand, "candidate_id") or "")


# ───────────────────────────── ép kiểu ───────────────────────────────────

def _get(row, key):
    """Đọc row[key] an toàn cho cả sqlite3.Row lẫn dict lẫn None."""
    if row is None:
        return None
    try:
        return row[key]
    except (KeyError, IndexError):
        return None


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _int(value):
    """Số nguyên để Excel hiểu là số; giá trị không phải số thì giữ nguyên chữ."""
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return text


def _date(value):
    """Chuỗi ngày trong DB → `datetime.date` cho Excel; không đọc được thì trả chữ."""
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    text = _text(value)
    if not text:
        return None
    try:
        return _dt.date.fromisoformat(text[:10])
    except ValueError:
        return text
