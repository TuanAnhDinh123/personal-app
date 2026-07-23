"""Đổi tên file CV hàng loạt và trích xuất dữ liệu CV ra Excel.

Tính năng đổi tên:
  • Quét thư mục → tìm file PDF/Word
  • Tự động trích xuất tên ứng viên từ tên file
  • Hiển thị bảng xem trước, cho phép chỉnh tên trước khi đổi
  • Đổi tên theo format: {prefixcode}_{startcode}_{TenUngVien}.ext
    start code tăng dần cho từng file theo thứ tự sắp xếp

Tính năng trích xuất:
  • Đọc nội dung từng file CV (PDF / DOCX)
  • Tách ID (mã CV) và Tên ứng viên từ tên file, dùng regex tìm Email/SĐT
  • Xuất theo template Excel có sẵn (sheet "Candidates"):
      – Chọn THƯ MỤC  → tạo file Excel mới theo template
      – Chọn FILE .xlsx → nối tiếp dữ liệu vào sheet "Candidates"
        (báo lỗi nếu file không có sheet này)
"""
import datetime
import html
import os
import re
import sys
import unicodedata
import zipfile
from pathlib import Path


from app.core import config

try:
    import fitz  # PyMuPDF — đọc text từ PDF
    _FITZ_OK = True
except ImportError:
    _FITZ_OK = False

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_CV_EXTENSIONS = {".pdf", ".doc", ".docx"}

SECTION = "scan_cv"
DEFAULTS = {
    "folder":         "",
    "prefix":         "",
    "start":          "01",
    "noise_keywords": (
        "cv\n"
        "c.v\n"
        "resume\n"
        "hồ sơ\n"
        "ứng tuyển\n"
        "ứng viên\n"
        "phỏng vấn\n"
        "đơn xin việc"
    ),
}

_NOISE_NUMBER = re.compile(r"^(?:20\d{2}|\d+)$")

# ----- Template Excel (sheet "Candidates") -----
_TEMPLATE_NAME = "template_cv.xlsx"
CANDIDATES_SHEET = "Candidates"
DATA_START_ROW = 12          # dòng dữ liệu đầu tiên (tiêu đề ở dòng 11)
# Cột (1-based) trong sheet Candidates
COL_BATCH = 1   # A  — Batch
COL_ID    = 2   # B  — ID
COL_NAME  = 3   # C  — NAME
COL_APPLY = 4   # D  — APPLYING FOR (phòng ban)
COL_EMAIL = 7   # G  — EMAIL ADDRESS
COL_PHONE = 8   # H  — PHONE

# Tên thư mục CV thường đặt là "batch 1", "batch 2"… → lấy số batch.
_BATCH_RE = re.compile(r"batch[\s_\-]*0*(\d+)", re.IGNORECASE)


def _batch_from_folder(folder: str):
    """Lấy số batch từ tên thư mục, vd 'batch 2' → 2. Không khớp thì trả None."""
    name = os.path.basename(os.path.normpath(folder))
    m = _BATCH_RE.search(name)
    return int(m.group(1)) if m else None


def _template_path() -> Path:
    """Đường dẫn tới file template Excel, đúng cả khi chạy dev lẫn bản .exe."""
    base = getattr(sys, "_MEIPASS", None)
    if base:
        p = Path(base) / "app" / _TEMPLATE_NAME
        if p.exists():
            return p
    return Path(__file__).resolve().parent.parent / _TEMPLATE_NAME


def _ascii_fold(s: str) -> str:
    s = s.replace("đ", "d").replace("Đ", "D")
    s = s.replace("ư", "u").replace("Ư", "U")
    s = s.replace("ơ", "o").replace("Ơ", "O")
    nfd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfd if ord(c) < 128)


def _parse_noise(raw: str) -> list[str]:
    result = []
    for line in raw.splitlines():
        for kw in line.split(","):
            kw = kw.strip()
            if kw:
                result.append(_ascii_fold(kw).lower())
    return result


def _title_vn(s: str) -> str:
    """Viết hoa chữ cái đầu mỗi từ, giữ nguyên dấu tiếng Việt."""
    return " ".join(w.capitalize() for w in s.split())


def _extract_name(stem: str, noise_list: list[str]) -> str:
    stem = unicodedata.normalize("NFC", stem)
    stem = re.sub(r"[\[\(][^\]\)]*[\]\)]", " ", stem)
    stem = re.sub(r"[_\-|/\\]+", " ", stem)

    words  = stem.split()
    folded = [_ascii_fold(w).lower() for w in words]

    max_win = max((len(kw.split()) for kw in noise_list), default=1)
    max_win = min(max_win, 4)

    keep = [True] * len(words)
    i = 0
    while i < len(words):
        removed = False
        for size in range(min(max_win, len(words) - i), 0, -1):
            chunk = " ".join(folded[i : i + size])
            if chunk in noise_list or _NOISE_NUMBER.match(chunk):
                for j in range(i, i + size):
                    keep[j] = False
                i += size
                removed = True
                break
        if not removed:
            i += 1

    raw = " ".join(w for w, k in zip(words, keep) if k)
    return _title_vn(raw)


def _build_filename(name: str, prefix: str, code: str, suffix: str) -> str:
    cv_code = (prefix + code).strip()   # prefix và startcode viết liền, không có _
    parts = [p for p in (cv_code, name) if p]
    return "_".join(parts) + suffix


def _seq_code(start_str: str, index: int) -> str:
    """Tính mã thứ tự: start_str='01', index=2 → '03'. Giữ nguyên số chữ số."""
    pad = len(start_str) if start_str.isdigit() else 2
    try:
        n = int(start_str) + index
    except ValueError:
        n = 1 + index
    return str(n).zfill(pad)


# ---------------------------------------------------------------------------
# Trích xuất dữ liệu từ CV (độc lập với UI, dễ test)
# ---------------------------------------------------------------------------

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# Số điện thoại VN: bắt đầu bằng 0, 84, +84 hoặc (+84); cho phép dấu
# cách/chấm/gạch xen giữa. Ví dụ: 0369..., +84 369..., (+84) 369...
_PHONE_RE = re.compile(r"(?:\(?\+?84\)?|0)[\d.\-\s]{8,13}")


def _read_pdf_text(path: Path) -> str:
    """Trích toàn bộ lớp text của file PDF (không OCR ảnh scan)."""
    parts = []
    with fitz.open(path) as doc:
        for page in doc:
            parts.append(page.get_text("text"))
    return "\n".join(parts)


def _read_docx_text(path: Path) -> str:
    """Trích text từ .docx bằng cách đọc word/document.xml (không cần thư viện ngoài)."""
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
    xml = xml.replace("</w:p>", "\n")          # mỗi đoạn xuống một dòng
    text = re.sub(r"<[^>]+>", " ", xml)         # bỏ toàn bộ thẻ XML
    return html.unescape(text)


def _extract_cv_text(path: Path) -> str:
    """Đọc nội dung text của một file CV theo phần mở rộng."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        if not _FITZ_OK:
            raise RuntimeError("Cần cài PyMuPDF để đọc PDF: pip install pymupdf")
        return _read_pdf_text(path)
    if suffix == ".docx":
        return _read_docx_text(path)
    # .doc cũ (định dạng nhị phân) không đọc được nếu không có Word/thư viện
    raise ValueError("Định dạng .doc cũ chưa hỗ trợ — hãy lưu lại thành .docx hoặc .pdf")


def _find_email(text: str) -> str:
    m = _EMAIL_RE.search(text)
    return m.group(0) if m else ""


def _find_phone(text: str) -> str:
    """Tìm số điện thoại VN đầu tiên hợp lệ.

    Giữ nguyên định dạng quốc tế: số viết dạng 84/+84/(+84) → trả về '+84…'
    (bỏ ngoặc và khoảng trắng). Số viết bắt đầu bằng 0 → giữ '0…'.
    """
    for m in _PHONE_RE.finditer(text):
        digits = re.sub(r"\D", "", m.group())
        if digits.startswith("84"):
            local = digits[2:]
            if 9 <= len(local) <= 10:
                return "+84" + local
        elif digits.startswith("0"):
            if 10 <= len(digits) <= 11:
                return digits
    return ""


def _split_id_name(stem: str, noise_list: list[str]) -> tuple[str, str]:
    """Tách (ID, Tên) từ tên file CV.

    File đã đổi tên có dạng '{prefix}{startcode}_{Tên ứng viên}', ví dụ
    '250601_Nguyen Van A' → ID='250601', Tên='Nguyen Van A'. Nếu phần đầu
    trước dấu phân cách không phải dãy số thì ID để trống và tên lấy từ
    toàn bộ tên file (dùng bộ lọc từ nhiễu như cũ).
    """
    norm = unicodedata.normalize("NFC", stem)
    m = re.match(r"\s*(\d{4,})[\s_\-]+(.+)", norm)
    if m:
        return m.group(1), _extract_name(m.group(2), noise_list)
    return "", _extract_name(norm, noise_list)


def _next_empty_row(ws) -> int:
    """Dòng trống đầu tiên (từ DATA_START_ROW) trong vùng dữ liệu ứng viên."""
    row = DATA_START_ROW
    while any(
        ws.cell(row=row, column=c).value not in (None, "")
        for c in (COL_ID, COL_NAME, COL_APPLY, COL_EMAIL, COL_PHONE)
    ):
        row += 1
    return row


def _write_candidates(ws, rows: list[dict]) -> None:
    """Ghi danh sách ứng viên vào sheet Candidates, nối tiếp sau dữ liệu cũ.

    Mọi ô được ghi đều chuẩn hóa về cùng một font: Aptos Display, cỡ 12,
    căn trái.
    """
    from openpyxl.styles import Alignment, Border, Font, Side

    cell_font   = Font(name="Aptos Display", size=12)
    cell_align  = Alignment(horizontal="left")
    _thin       = Side(style="thin", color="000000")
    cell_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    start = _next_empty_row(ws)
    # Kẻ khung cho toàn bộ cột từ A (1) đến W (23) của mỗi dòng được thêm.
    border_cols = range(1, 24)
    field_col = {
        "batch": COL_BATCH, "id": COL_ID, "name": COL_NAME, "apply": COL_APPLY,
        "email": COL_EMAIL, "phone": COL_PHONE,
    }
    for i, r in enumerate(rows):
        row_no = start + i
        # Kẻ border cho toàn bộ ô của dòng được thêm (kể cả ô để trống).
        for col in border_cols:
            ws.cell(row=row_no, column=col).border = cell_border
        for field, col in field_col.items():
            value = r.get(field, "")
            if value:
                cell = ws.cell(row=row_no, column=col, value=value)
                cell.font = cell_font
                cell.alignment = cell_align


def _safe_filename(name: str) -> str:
    """Bỏ các ký tự không hợp lệ trong tên file Windows."""
    return re.sub(r'[<>:"/\\|?*]+', "_", name).strip() or "Candidates"


def _open_template_workbook():
    """Mở file template, trả về (workbook, worksheet Candidates)."""
    tpl = _template_path()
    if not tpl.exists():
        raise FileNotFoundError(
            f"Không tìm thấy file template:\n{tpl}")
    wb = openpyxl.load_workbook(tpl)
    if CANDIDATES_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Template thiếu sheet '{CANDIDATES_SHEET}'.")
    return wb, wb[CANDIDATES_SHEET]


def _open_existing_workbook(path: str):
    """Mở file Excel có sẵn để nối tiếp; báo lỗi nếu thiếu sheet Candidates."""
    wb = openpyxl.load_workbook(path)
    if CANDIDATES_SHEET not in wb.sheetnames:
        raise ValueError(
            f"File Excel không có sheet '{CANDIDATES_SHEET}'.\n"
            "Hãy chọn file đúng mẫu hoặc chọn một thư mục để tạo file mới.")
    return wb, wb[CANDIDATES_SHEET]
