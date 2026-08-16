"""Quét CV bằng AI (Google Gemini) → hồ sơ có cấu trúc + điểm phù hợp với JD.

Một lần gọi API trả về BA phần tách bạch (xem `_RESPONSE_SCHEMA`):

    profile     — mô tả CON NGƯỜI, TRUNG TÍNH với mọi JD: họ tên, liên hệ, chức
                  danh gần nhất, ngành, học vấn, kỹ năng, nguyện vọng… Đây là
                  phần dùng lại được mãi: có yêu cầu tuyển dụng mới thì tìm
                  trong pool bằng chính dữ liệu này, không phải đọc lại PDF.
    experiences — DÒNG THỜI GIAN công việc (ngày vào – ngày ra) để tính lại số
                  năm kinh nghiệm ở bất kỳ thời điểm nào về sau.
    evaluation  — mức độ phù hợp với ĐÚNG JD của lần quét này (điểm 0–100,
                  nhận xét, kỹ năng khớp/thiếu). Phần này gắn chặt với JD nên
                  lưu vào bảng lịch sử `candidate_evaluations`, không lưu vào
                  hồ sơ ứng viên.

Cách hoạt động:
    1. Đọc từng file PDF trong thư mục → mã hóa base64.
    2. Gọi Gemini REST API (generateContent) kèm JD làm ngữ cảnh, yêu cầu trả
       về JSON có cấu trúc (responseSchema).
    3. Ghi kết quả vào DB (xem app_qt/tools/ai_scan_cv.py).

Chỉ dùng thư viện chuẩn để gọi API (urllib) nên KHÔNG cần cài thêm gói.
Cần: một API key Gemini (https://aistudio.google.com/apikey) và kết nối mạng.
"""
import base64
import json
import os
import shutil
import threading
import time
import urllib.error
import urllib.request
from pathlib import Path


from app.core import config, settings

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

SECTION = "ai_scan_cv"
# API key & model nay là thiết lập CHUNG (màn hình Cài đặt) — xem app/core/settings.py.
# Tool này chỉ còn giữ đường dẫn vào/ra và JD của riêng nó.
DEFAULTS = {
    "folder":       "",
    "output":       "",
    "jd_file":      "",   # đường dẫn tới file JD (PDF/DOCX/TXT) — thay cho ô dán JD cũ
    "extra_prompt": "",   # yêu cầu bổ sung người dùng gửi kèm cho AI
}

# Tên folder (ngang cấp) chứa các CV đã quét xong, để lần sau không quét lại.
_DONE_SUFFIX = "_scanned"

_API_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "{model}:generateContent"
)

# Schema JSON mà Gemini phải tuân theo khi trả kết quả cho mỗi CV. Mọi mô tả
# đều viết tiếng Anh vì đây là phần "prompt" gửi thẳng lên API — thống nhất
# ngôn ngữ với phần yêu cầu ở _build_prompt (respond entirely in English).
#
# CHIA BA PHẦN có chủ đích: `profile` + `experiences` mô tả con người (dùng lại
# được cho mọi vị trí về sau), `evaluation` chỉ đúng với JD của lần quét này.
_PROFILE_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "name":  {"type": "STRING", "description": "Candidate's full name"},
        "dob":   {"type": "STRING", "description": "Date of birth (dd/mm/yyyy if available)"},
        "email": {"type": "STRING"},
        "phone": {"type": "STRING", "description": "Phone number"},
        "gender": {"type": "STRING", "description": "Male / Female / Other"},
        "address": {"type": "STRING"},
        "city":  {"type": "STRING", "description": "Province or city they live in"},
        "current_title": {"type": "STRING", "description": "Most recent job title"},
        "industry": {"type": "STRING", "description": "Industry they have worked in"},
        "years_experience": {"type": "NUMBER",
                             "description": "Total years of work experience as stated in the CV"},
        "education": {"type": "STRING", "description": "Highest education level"},
        "major": {"type": "STRING", "description": "Field of study"},
        "languages": {"type": "STRING",
                      "description": "Languages with level, separated by ';'"},
        "skills": {"type": "ARRAY", "items": {"type": "STRING"},
                   "description": "Skills exactly as written in the CV, one per item"},
        "profile_summary": {"type": "STRING",
                            "description": "2-3 neutral sentences describing the candidate. "
                                           "Never mention the job description."},
        "expected_salary": {"type": "STRING",
                            "description": "Expected salary if the CV states one"},
        "available_from": {"type": "STRING",
                           "description": "Earliest start date if the CV states one"},
    },
    "required": ["name", "dob", "email", "phone", "current_title",
                 "years_experience", "skills", "profile_summary"],
}

_EXPERIENCE_SCHEMA = {
    "type": "ARRAY",
    "description": "Every job in the CV, most recent first",
    "items": {
        "type": "OBJECT",
        "properties": {
            "company":   {"type": "STRING"},
            "job_title": {"type": "STRING"},
            "industry":  {"type": "STRING"},
            "start_date": {"type": "STRING",
                           "description": "yyyy-mm-dd; use the 1st when only month/year is given"},
            "end_date":  {"type": "STRING",
                          "description": "yyyy-mm-dd; leave empty if this is the current job"},
            "is_current": {"type": "BOOLEAN",
                           "description": "true if the candidate still worked here when writing the CV"},
            "description": {"type": "STRING", "description": "What they did there, one or two lines"},
        },
        "required": ["company", "job_title", "start_date"],
    },
}

_EVALUATION_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "fit_score":   {"type": "INTEGER", "description": "How well the candidate fits the JD, 0-100"},
        "fit_summary": {"type": "STRING", "description": "Short remark on why they fit / don't fit"},
        "strengths":   {"type": "STRING", "description": "Key strengths (short bullet points)"},
        "weaknesses":  {"type": "STRING", "description": "Weaknesses / gaps compared to the JD"},
        "matched_skills": {"type": "STRING",
                           "description": "Skills the JD asks for AND the candidate has, separated by ';'"},
        "missing_skills": {"type": "STRING",
                           "description": "Skills the JD asks for but the candidate lacks, separated by ';'"},
    },
    "required": ["fit_score", "fit_summary", "strengths", "weaknesses"],
}

_RESPONSE_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "profile":     _PROFILE_SCHEMA,
        "experiences": _EXPERIENCE_SCHEMA,
        "evaluation":  _EVALUATION_SCHEMA,
    },
    "required": ["profile", "experiences", "evaluation"],
}

# Cột xuất ra Excel khi bỏ qua ứng viên trùng: (khóa sau khi làm phẳng bằng
# `flatten_result`, tiêu đề hiển thị, độ rộng cột).
_COLUMNS = [
    ("batch",       "Batch",         14),
    ("file",        "File name",     28),
    ("cv_path",     "CV path",       52),
    ("name",        "Full name",     22),
    ("dob",         "Date of birth", 14),
    ("email",       "Email",         28),
    ("phone",       "Phone",         16),
    ("current_title", "Current title", 24),
    ("years_experience", "Years",    10),
    ("skills",      "Skills",        46),
    ("fit_score",   "Fit score",     13),
    ("fit_summary", "Fit summary",   46),
    ("strengths",   "Strengths",     46),
    ("weaknesses",  "Weaknesses",    46),
]


def flatten_result(data: dict) -> dict:
    """Kết quả 3 tầng của Gemini → dict phẳng cho bảng Excel / hiển thị nhanh."""
    profile = data.get("profile") or {}
    evaluation = data.get("evaluation") or {}
    flat = {k: data.get(k, "") for k in ("batch", "file", "cv_path")}
    flat.update({
        "name":  profile.get("name", ""),
        "dob":   profile.get("dob", ""),
        "email": profile.get("email", ""),
        "phone": profile.get("phone", ""),
        "current_title": profile.get("current_title", ""),
        "years_experience": profile.get("years_experience", ""),
        "skills": "; ".join(profile.get("skills") or []),
        "fit_score":   evaluation.get("fit_score", ""),
        "fit_summary": evaluation.get("fit_summary", ""),
        "strengths":   evaluation.get("strengths", ""),
        "weaknesses":  evaluation.get("weaknesses", ""),
    })
    return flat


def read_jd_file(path: str) -> str:
    """Đọc nội dung JD từ file PDF / DOCX / TXT thành text thuần.

    Dùng lại bộ trích text của tool "Quét CV" cho PDF/DOCX; .txt đọc trực tiếp.
    """
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"JD file not found: {path}")
    suffix = p.suffix.lower()
    if suffix == ".txt":
        return p.read_text(encoding="utf-8", errors="ignore")
    # PDF/DOCX → dùng chung logic với app.core.cv_scan (tránh phụ thuộc vòng).
    from app.core.cv_scan import _extract_cv_text
    return _extract_cv_text(p)


def _build_prompt(jd: str, extra: str = "") -> str:
    """Câu lệnh gửi kèm mỗi CV.

    jd    — nội dung mô tả công việc (đọc từ file JD người dùng chọn).
    extra — yêu cầu bổ sung tùy ý người dùng nhập thêm cho AI.
    """
    jd = jd.strip() or "(No specific job description — give a general assessment.)"
    extra = extra.strip()
    extra_block = ""
    if extra:
        extra_block = (
            "\n=== ADDITIONAL INSTRUCTIONS FROM THE USER ===\n"
            f"{extra}\n"
            "=== END OF ADDITIONAL INSTRUCTIONS ===\n"
        )
    return (
        "You are a recruiter. Read the CV in the attached PDF carefully and "
        "return three separate things.\n\n"
        "1. 'profile' — a description of the PERSON. This is stored in a talent "
        "pool and reused for future openings, so it must be written as if no job "
        "description existed. Never compare the candidate to the JD here.\n"
        "2. 'experiences' — every job in the CV as a timeline with start and end "
        "dates, so their years of experience can be recomputed at any later date.\n"
        "3. 'evaluation' — how well they fit the job description below.\n\n"
        "=== JOB DESCRIPTION (JD) ===\n"
        f"{jd}\n"
        "=== END OF JD ===\n"
        f"{extra_block}\n"
        "Requirements:\n"
        "- Respond entirely in English.\n"
        "- Leave a field as an empty string if it is not found in the CV; never "
        "invent information.\n"
        "- Dates use yyyy-mm-dd. When only a month and year are given, use the "
        "1st of that month. Leave 'end_date' empty for the current job and set "
        "'is_current' to true.\n"
        "- 'skills' lists each skill exactly as written in the CV, one per item — "
        "no grouping, no explanation.\n"
        "- 'years_experience' is the total the CV itself supports, as a number.\n"
        "- 'fit_score' is an integer 0-100 for how well the CV matches the JD.\n"
        "- Write 'strengths' and 'weaknesses' as short one-line bullet points.\n"
        "Return only the JSON object matching the given schema."
    )


# Các mã lỗi HTTP mang tính tạm thời (quá tải / giới hạn nhịp) — nên thử lại.
_RETRY_STATUS = {429, 500, 502, 503, 504}
_MAX_RETRIES = 4          # số lần thử lại tối đa cho mỗi CV
_RETRY_BASE_DELAY = 4     # giây; chờ tăng dần: 4s, 8s, 16s…


def _call_gemini(api_key: str, model: str, jd: str, pdf_bytes: bytes,
                 timeout: int = 180, on_retry=None, extra: str = "",
                 should_cancel=None) -> dict:
    """Gửi 1 file PDF cho Gemini, trả về dict theo _RESPONSE_SCHEMA.

    Tự động thử lại khi gặp lỗi tạm thời (429/5xx, ví dụ HTTP 503 "high
    demand"), chờ tăng dần giữa các lần. `on_retry(attempt, wait, reason)`
    (nếu có) được gọi trước mỗi lần chờ để báo tiến trình. Ném RuntimeError
    với thông báo dễ hiểu nếu vẫn thất bại.

    `should_cancel()` (nếu có) được kiểm tra thường xuyên — kể cả trong lúc
    đang chờ giữa các lần thử lại — để bấm Hủy có hiệu lực NGAY, ném _Cancelled.
    """
    last_error = ""
    for attempt in range(1, _MAX_RETRIES + 1):
        if should_cancel and should_cancel():
            raise _Cancelled()
        try:
            return _call_gemini_once(api_key, model, jd, pdf_bytes, timeout, extra)
        except _TransientError as exc:
            last_error = str(exc)
            if attempt == _MAX_RETRIES:
                break
            wait = _RETRY_BASE_DELAY * (2 ** (attempt - 1))
            if on_retry:
                on_retry(attempt, wait, last_error)
            _interruptible_sleep(wait, should_cancel)
    raise RuntimeError(f"Failed after {_MAX_RETRIES} attempts — {last_error}")


class _TransientError(Exception):
    """Lỗi tạm thời (nên thử lại): quá tải, giới hạn nhịp, mất kết nối."""


class _Cancelled(Exception):
    """Người dùng bấm Hủy giữa chừng (kể cả khi đang chờ thử lại)."""


def _interruptible_sleep(seconds: float, should_cancel=None) -> None:
    """Ngủ `seconds` giây nhưng cứ 0.2s lại kiểm tra Hủy; nếu Hủy → _Cancelled."""
    remaining = seconds
    while remaining > 0:
        if should_cancel and should_cancel():
            raise _Cancelled()
        chunk = 0.2 if remaining > 0.2 else remaining
        time.sleep(chunk)
        remaining -= chunk


def _call_gemini_once(api_key: str, model: str, jd: str, pdf_bytes: bytes,
                      timeout: int, extra: str = "") -> dict:
    """Một lần gọi API. Ném _TransientError nếu lỗi có thể thử lại."""
    body = {
        "contents": [{
            "parts": [
                {"text": _build_prompt(jd, extra)},
                {"inline_data": {
                    "mime_type": "application/pdf",
                    "data": base64.b64encode(pdf_bytes).decode("ascii"),
                }},
            ],
        }],
        "generationConfig": {
            "temperature": 0.2,
            "responseMimeType": "application/json",
            "responseSchema": _RESPONSE_SCHEMA,
        },
    }
    url = _API_URL.format(model=model)
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "x-goog-api-key": api_key,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "ignore")
        msg = detail
        try:
            msg = json.loads(detail).get("error", {}).get("message", detail)
        except ValueError:
            pass
        if exc.code in _RETRY_STATUS:
            raise _TransientError(f"HTTP {exc.code}: {msg}") from exc
        raise RuntimeError(f"HTTP {exc.code}: {msg}") from exc
    except urllib.error.URLError as exc:
        # Timeout / mất mạng thường là tạm thời → cho thử lại.
        raise _TransientError(f"Network error: {exc.reason}") from exc

    try:
        text = payload["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError):
        # Có thể bị chặn do safety hoặc phản hồi rỗng.
        reason = payload.get("promptFeedback", {}).get("blockReason")
        raise RuntimeError(
            f"No content returned by the model"
            + (f" (blocked: {reason})" if reason else ""))
    try:
        data = json.loads(text)
    except ValueError:
        raise RuntimeError("The model returned invalid JSON.")
    if not isinstance(data, dict):
        raise RuntimeError("The model returned an unexpected JSON shape.")
    # Chuẩn hóa ngay tại đây để mọi nơi dùng kết quả đều thấy dữ liệu sạch.
    data.setdefault("profile", {})
    data.setdefault("experiences", [])
    data.setdefault("evaluation", {})
    profile = data["profile"]
    if isinstance(profile, dict) and profile.get("name"):
        profile["name"] = normalize_name(profile["name"])
    return data


def normalize_name(name: str) -> str:
    """Chuẩn hóa tên ứng viên AI trả về: viết hoa chữ cái đầu mỗi từ, GIỮ dấu.

    Ví dụ: "ĐINH QUANG SƠN" hoặc "đinh quang sơn" → "Đinh Quang Sơn".
    Chỉ đổi kiểu chữ hoa/thường (Unicode-aware nên dấu tiếng Việt giữ nguyên),
    không bỏ dấu, không xóa ký tự. Khoảng trắng thừa được gộp lại.
    """
    if not name:
        return ""
    return " ".join(w[:1].upper() + w[1:].lower() for w in name.split())


_SHEET_TITLE = "AI CV Scan"


def _write_header(ws) -> None:
    """Kẻ hàng tiêu đề in đậm + đặt độ rộng cột."""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    for col, (_, title, width) in enumerate(_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[c.column_letter].width = width
    ws.freeze_panes = "A2"


def append_rows_to_excel(rows: list[dict], out_path: str) -> None:
    """Ghi NỐI TIẾP các dòng kết quả vào file .xlsx.

    • Nếu file chưa tồn tại → tạo mới kèm hàng tiêu đề.
    • Nếu đã tồn tại → mở ra và ghi tiếp phía dưới các record cũ, nhờ vậy
      nhiều lần quét (mỗi lần dừng do hết hạn mức key free) vẫn dồn chung một
      file thay vì đè lên nhau.
    """
    from openpyxl.styles import Alignment, Font

    if os.path.exists(out_path):
        wb = openpyxl.load_workbook(out_path)
        ws = wb[_SHEET_TITLE] if _SHEET_TITLE in wb.sheetnames else wb.active
        if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
            _write_header(ws)
        start = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _SHEET_TITLE
        _write_header(ws)
        start = 2

    cell_font = Font(name="Segoe UI", size=11)
    wrap      = Alignment(vertical="top", wrap_text=True)
    for r, row in enumerate(rows, start=start):
        for col, (key, _, _) in enumerate(_COLUMNS, start=1):
            c = ws.cell(row=r, column=col, value=row.get(key, ""))
            c.font = cell_font
            c.alignment = wrap
    wb.save(out_path)


# Giữ tên cũ cho tương thích; mặc định giờ là ghi nối tiếp.
_write_excel = append_rows_to_excel


# --------------------------------------------------------------- folder "đã quét"
def done_folder_for(folder: str) -> Path:
    """Trả về đường dẫn folder (ngang cấp) chứa CV đã quét — chưa tạo trên đĩa."""
    src = Path(folder)
    return src.parent / f"{src.name}{_DONE_SUFFIX}"


def resolve_done_target(path: Path, done_dir: Path) -> Path:
    """Tính đường dẫn đích trong folder 'đã quét' cho 1 CV (CHƯA di chuyển).

    Tạo sẵn folder đích; nếu trùng tên thì thêm hậu tố _1, _2… Trả về đường
    dẫn cuối cùng mà file sẽ nằm — dùng để ghi vào Excel TRƯỚC khi thực sự
    move, nhờ vậy cột 'Đường dẫn CV' khớp với vị trí file sau khi quét.
    """
    done_dir.mkdir(parents=True, exist_ok=True)
    target = done_dir / path.name
    i = 1
    while target.exists():
        target = done_dir / f"{path.stem}_{i}{path.suffix}"
        i += 1
    return target


def move_to_done(path: Path, done_dir: Path, target: Path = None) -> Path:
    """Chuyển 1 file CV đã quét xong sang folder 'đã quét'.

    Nếu truyền sẵn `target` (từ resolve_done_target) thì move tới đúng đó;
    nếu không thì tự tính. Trả về đường dẫn mới của file.
    """
    if target is None:
        target = resolve_done_target(path, done_dir)
    else:
        target.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(path), str(target))
    return target
