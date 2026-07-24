"""Engine xuất Excel theo TEMPLATE có sẵn dùng cú pháp placeholder ``{{field}}``.

Ý tưởng: người dùng (HR) tự thiết kế file Excel mẫu — tiêu đề, style, merge cell,
logo, công thức… — và đặt các "chỗ trống" bằng token ``{{ten_field}}``. Code chỉ
lo BƠM dữ liệu vào, không đụng tới layout. Đổi mẫu → sửa file Excel, không sửa code.

Hai loại placeholder:

* **Scalar** (điền một lần, ở phần tiêu đề): ví dụ ``{{course.title}}``,
  ``{{course.date}}``, ``{{count}}`` → lấy giá trị từ ``context``.
* **Theo dòng** (lặp lại cho từng phần tử của danh sách): ví dụ ``{{full_name}}``,
  ``{{code}}``, ``{{stt}}``. Engine tự dò *dòng mẫu* (dòng có chứa token thuộc
  ``row_fields``), rồi với N phần tử sẽ chèn thêm N-1 dòng, sao chép style + merge
  của dòng mẫu và điền giá trị.

Module thuần logic (không phụ thuộc UI) → dễ test và tái dùng.
"""
from __future__ import annotations

import os
import re
from copy import copy

import openpyxl

# Token dạng {{ khóa }} — khóa gồm chữ, số, gạch dưới, dấu chấm (vd course.title).
_TOKEN_RE = re.compile(r"\{\{\s*([\w.]+)\s*\}\}")


def find_placeholders(template_path: str) -> set[str]:
    """Quét file mẫu, trả về tập tất cả khóa placeholder xuất hiện (để debug/kiểm tra)."""
    wb = _load(template_path)
    keys: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    keys.update(_TOKEN_RE.findall(cell.value))
    wb.close()
    return keys


def render_template(template_path, out_path, context, rows, row_fields=None):
    """Đổ dữ liệu vào file mẫu rồi lưu ra ``out_path``.

    Tham số:
      template_path : đường dẫn file Excel mẫu (.xlsx hoặc .xlsm — giữ nguyên macro).
      out_path      : nơi lưu file kết quả.
      context       : dict giá trị scalar, khóa = tên placeholder (vd "course.title").
      rows          : list[dict] — mỗi phần tử là một dòng dữ liệu.
      row_fields    : (tùy chọn) tập tên field "theo dòng" để nhận diện dòng mẫu.
                      Bỏ trống → suy ra từ khóa của ``rows`` (cộng "stt"). Nên truyền
                      tường minh để vẫn xuất được khi danh sách rỗng.

    Ném FileNotFoundError nếu không có file mẫu; ValueError nếu mẫu thiếu dòng chứa
    token theo dòng (khi có row_fields để dò).
    """
    if not template_path or not os.path.isfile(template_path):
        raise FileNotFoundError(template_path)

    if row_fields is None:
        row_fields = set()
        for r in rows:
            row_fields.update(r.keys())
        row_fields.add("stt")
    row_fields = set(row_fields)

    wb = _load(template_path)

    # Dò sheet chứa dòng mẫu (dòng có token thuộc row_fields). Không có → chỉ có
    # phần scalar, bỏ qua bước lặp dòng.
    roster_ws, tpl_row = _find_row_template(wb, row_fields)
    if roster_ws is not None:
        _expand_rows(roster_ws, tpl_row, rows, row_fields)

    # Thay các placeholder scalar còn lại trên MỌI sheet.
    for ws in wb.worksheets:
        _fill_scalars(ws, context)

    wb.save(out_path)
    wb.close()


# ───────────────────────────── nội bộ ──────────────────────────────────────

def _load(path: str):
    """Mở workbook; file .xlsm giữ macro (keep_vba)."""
    keep_vba = path.lower().endswith(".xlsm")
    return openpyxl.load_workbook(path, keep_vba=keep_vba)


def _cell_has_row_token(value, row_fields) -> bool:
    if not isinstance(value, str):
        return False
    return any(k in row_fields for k in _TOKEN_RE.findall(value))


def _find_row_template(wb, row_fields):
    """Trả về (worksheet, chỉ_số_dòng_mẫu) đầu tiên chứa token theo dòng, hoặc (None, None)."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _cell_has_row_token(cell.value, row_fields):
                    return ws, cell.row
    return None, None


def _sub(text: str, mapping: dict):
    """Thay mọi token trong chuỗi. Nếu cả ô CHỈ là một token → trả về giá trị GỐC
    (giữ nguyên kiểu số/ngày để Excel không hiểu nhầm thành text)."""
    full = _TOKEN_RE.fullmatch(text.strip())
    if full:
        val = mapping.get(full.group(1))
        return "" if val is None else val

    def repl(m):
        val = mapping.get(m.group(1))
        return "" if val is None else str(val)

    return _TOKEN_RE.sub(repl, text)


def _copy_style(src, dst):
    """Sao chép định dạng từ ô src sang ô dst (style là proxy bất biến → gán trực tiếp)."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format


def _row_merges(ws, row_idx):
    """Các vùng merge NẰM GỌN trong một dòng (merge dọc bỏ qua) → list (col_min, col_max)."""
    out = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == row_idx and rng.max_row == row_idx:
            out.append((rng.min_col, rng.max_col))
    return out


def _expand_rows(ws, tpl_row, rows, row_fields):
    """Nhân dòng mẫu ``tpl_row`` thành len(rows) dòng dữ liệu (copy style + merge)."""
    max_col = ws.max_column
    # Chụp lại nội dung + style + merge của dòng mẫu TRƯỚC khi chèn thêm dòng.
    tpl_values = [ws.cell(tpl_row, c).value for c in range(1, max_col + 1)]
    tpl_cells = [ws.cell(tpl_row, c) for c in range(1, max_col + 1)]
    tpl_merges = _row_merges(ws, tpl_row)
    tpl_height = ws.row_dimensions[tpl_row].height

    n = len(rows)
    if n == 0:
        # Không có dữ liệu → xóa token ở dòng mẫu cho sạch, giữ style.
        for c in range(1, max_col + 1):
            v = tpl_values[c - 1]
            if isinstance(v, str):
                ws.cell(tpl_row, c).value = _TOKEN_RE.sub("", v) or None
        return

    if n > 1:
        # Chèn N-1 dòng ngay dưới dòng mẫu (đẩy footer xuống, merge dưới tự dịch).
        ws.insert_rows(tpl_row + 1, n - 1)

    for i, record in enumerate(rows):
        r = tpl_row + i
        mapping = {**record, "stt": i + 1}
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            src = tpl_cells[c - 1]
            if i > 0:                       # dòng chèn thêm cần copy style
                _copy_style(src, cell)
            v = tpl_values[c - 1]
            cell.value = _sub(v, mapping) if isinstance(v, str) else v
        if tpl_height is not None:
            ws.row_dimensions[r].height = tpl_height
        if i > 0:                           # tái lập merge cho dòng chèn thêm
            for col_min, col_max in tpl_merges:
                ws.merge_cells(start_row=r, start_column=col_min,
                               end_row=r, end_column=col_max)


def _fill_scalars(ws, context: dict):
    """Thay các placeholder scalar còn lại trên một sheet (bỏ qua ô không phải chuỗi)."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                cell.value = _sub(cell.value, context)
