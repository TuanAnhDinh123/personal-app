import os
import tkinter as tk
from tkinter import messagebox

from app.core.base_tool import BaseTool
from app.ui import theme, widgets

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


class MergeExcelTool(BaseTool):
    name = "Gộp file Excel"
    description = "Tự động điền dữ liệu từ các file con (theo MSNV) vào file tổng hợp."
    icon = "📊"
    category = "Tệp & Tài liệu"
    order = 10
    action_label = "Gộp dữ liệu"

    def build_body(self, parent):
        widgets.section_label(parent, "Nguồn & Đích")
        self._target_var = widgets.file_row(
            parent, "File đích (.xlsx) — cột A chứa MSNV", mode="file")
        self._folder_var = widgets.file_row(
            parent, "Thư mục chứa file con (tên file bắt đầu bằng MSNV)", mode="folder")

        widgets.section_label(parent, "Cấu hình")
        self._target_sheet_var = widgets.text_row(
            parent, "Sheet trong file đích (để trống = sheet đầu tiên)", placeholder="")
        self._child_sheet_var = widgets.text_row(
            parent, "Sheet trong file con (để trống = sheet đầu tiên)", placeholder="")
        self._child_row_var = widgets.text_row(
            parent, "Hàng chứa dữ liệu trong file con", placeholder="2")
        self._child_col_var = widgets.text_row(
            parent, "Cột bắt đầu đọc trong file con (VD: A)", placeholder="A")
        self._target_col_var = widgets.text_row(
            parent, "Cột bắt đầu ghi trong file đích (VD: B)", placeholder="B")

        widgets.hint(
            parent,
            "💡 Ví dụ: MSNV = NV001  →  file con tên NV001_BaoCao.xlsx hoặc NV001.xlsx.\n"
            "Chỉ điền vào các dòng có MSNV ở cột A nhưng chưa có dữ liệu ở các cột phía sau.\n"
            "⚠ Hãy sao lưu file đích trước khi chạy — dữ liệu sẽ được ghi trực tiếp."
        )

    def run(self):
        if not _OPENPYXL_OK:
            messagebox.showerror(
                "Thiếu thư viện",
                "Cần cài đặt openpyxl:\n  pip install openpyxl"
            )
            return

        target_path = self._target_var.get().strip()
        folder_path = self._folder_var.get().strip()

        if not target_path or not os.path.isfile(target_path):
            messagebox.showerror("Lỗi", "Vui lòng chọn file đích hợp lệ.")
            return
        if not folder_path or not os.path.isdir(folder_path):
            messagebox.showerror("Lỗi", "Vui lòng chọn thư mục chứa file con.")
            return

        target_sheet = self._target_sheet_var.get().strip() or None
        child_sheet = self._child_sheet_var.get().strip() or None

        try:
            child_row = int(self._child_row_var.get().strip() or "2")
            if child_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Lỗi", "Hàng dữ liệu phải là số nguyên dương (VD: 2).")
            return

        try:
            child_col = column_index_from_string(
                self._child_col_var.get().strip().upper() or "A")
            target_col = column_index_from_string(
                self._target_col_var.get().strip().upper() or "B")
        except Exception:
            messagebox.showerror("Lỗi", "Cột không hợp lệ — nhập theo dạng A, B, C, AB, …")
            return

        try:
            filled, skipped, not_found, errors = _do_merge(
                target_path=target_path,
                folder_path=folder_path,
                target_sheet=target_sheet,
                child_sheet=child_sheet,
                child_data_row=child_row,
                child_start_col=child_col,
                target_start_col=target_col,
            )
        except Exception as exc:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{exc}")
            return

        lines = [f"✅ Đã điền dữ liệu: {filled} dòng."]
        if skipped:
            lines.append(f"⏭ Bỏ qua (đã có dữ liệu): {skipped} dòng.")
        if not_found:
            preview = ", ".join(not_found[:5])
            extra = f" (+{len(not_found) - 5} nữa)" if len(not_found) > 5 else ""
            lines.append(f"⚠ Không tìm thấy file con: {preview}{extra}")
        if errors:
            preview = "; ".join(errors[:3])
            extra = f" (+{len(errors) - 3} lỗi nữa)" if len(errors) > 3 else ""
            lines.append(f"❌ Lỗi đọc file con: {preview}{extra}")
        messagebox.showinfo("Kết quả gộp", "\n".join(lines))


# ---------------------------------------------------------------------------
# Core logic (độc lập với UI, dễ test)
# ---------------------------------------------------------------------------

def _msnv_matches(filename, msnv):
    """Tên file bắt đầu đúng với MSNV, ký tự tiếp theo phải là dấu phân cách
    hoặc hết tên (trước phần mở rộng) để tránh khớp nhầm NV01 với NV010."""
    if not filename.startswith(msnv):
        return False
    rest = filename[len(msnv):]
    return not rest or rest[0] in ('_', '-', ' ', '.')


def _do_merge(target_path, folder_path, target_sheet, child_sheet,
              child_data_row, child_start_col, target_start_col):
    wb = openpyxl.load_workbook(target_path)
    ws = (
        wb[target_sheet]
        if (target_sheet and target_sheet in wb.sheetnames)
        else wb.active
    )
    max_col = ws.max_column

    # Lập bảng index file con: {filename_lowercase: full_path}
    child_files = {}
    for fname in os.listdir(folder_path):
        if fname.lower().endswith(('.xlsx', '.xlsm')):
            child_files[fname] = os.path.join(folder_path, fname)

    filled = skipped = 0
    not_found = []
    errors = []

    for row in ws.iter_rows(min_row=2):
        msnv = str(row[0].value or "").strip()
        if not msnv:
            continue

        # Bỏ qua nếu đã có dữ liệu từ cột target_start_col trở đi
        target_cells = row[target_start_col - 1:]
        if any(c.value is not None and str(c.value).strip() for c in target_cells):
            skipped += 1
            continue

        # Tìm file con khớp với MSNV
        child_path = next(
            (path for fname, path in child_files.items()
             if _msnv_matches(fname, msnv)),
            None,
        )
        if child_path is None:
            not_found.append(msnv)
            continue

        # Đọc hàng dữ liệu từ file con
        try:
            wb_child = openpyxl.load_workbook(child_path, data_only=True)
            ws_child = (
                wb_child[child_sheet]
                if (child_sheet and child_sheet in wb_child.sheetnames)
                else wb_child.active
            )
            child_rows = list(ws_child.iter_rows(
                min_row=child_data_row, max_row=child_data_row,
                min_col=child_start_col, values_only=True,
            ))
            wb_child.close()
        except Exception as exc:
            errors.append(f"{os.path.basename(child_path)}: {exc}")
            continue

        if not child_rows or not child_rows[0]:
            errors.append(
                f"{os.path.basename(child_path)}: hàng {child_data_row} rỗng hoặc không tồn tại")
            continue

        values = child_rows[0]

        # Ghi vào file đích
        row_num = row[0].row
        for i, val in enumerate(values):
            col_num = target_start_col + i
            if col_num > max_col:
                break
            ws.cell(row=row_num, column=col_num, value=val)

        filled += 1

    wb.save(target_path)
    wb.close()
    return filled, skipped, not_found, errors
