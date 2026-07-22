"""Quét CV bằng AI (Google Gemini) → xuất bảng đánh giá ứng viên ra Excel.

Khác với tool "Quét CV" (chỉ dùng regex để tách Email/SĐT), tool này gửi
NGUYÊN file PDF cho mô hình Gemini để đọc hiểu và trả về:
    • Họ tên, ngày sinh, email, số điện thoại
    • Mức độ phù hợp với JD (điểm 0–100 + nhận xét)
    • Ưu điểm / nhược điểm của ứng viên

Cách hoạt động:
    1. Đọc từng file PDF trong thư mục → mã hóa base64.
    2. Gọi Gemini REST API (generateContent) kèm JD làm ngữ cảnh, yêu cầu
       trả về JSON có cấu trúc (responseSchema).
    3. Gom kết quả tất cả CV → ghi ra một file Excel (openpyxl).

Chỉ dùng thư viện chuẩn để gọi API (urllib) nên KHÔNG cần cài thêm gói.
Cần: một API key Gemini (https://aistudio.google.com/apikey) và kết nối mạng.
"""
import base64
import json
import os
import threading
import time
import urllib.error
import urllib.request
from pathlib import Path
from tkinter import messagebox

import tkinter as tk
import ttkbootstrap as ttk

from app.core import config, settings
from app.core.base_tool import BaseTool
from app.ui import widgets, theme

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

SECTION = "ai_scan_cv"
# API key & model nay là thiết lập CHUNG (màn hình Cài đặt) — xem app/core/settings.py.
# Tool này chỉ còn giữ đường dẫn vào/ra và JD của riêng nó.
DEFAULTS = {
    "folder":  "",
    "output":  "",
    "jd":      "",
}

_API_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "{model}:generateContent"
)

# Schema JSON mà Gemini phải tuân theo khi trả kết quả cho mỗi CV.
_RESPONSE_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "name":       {"type": "STRING", "description": "Họ và tên ứng viên"},
        "dob":        {"type": "STRING", "description": "Ngày/tháng/năm sinh (dd/mm/yyyy nếu có)"},
        "email":      {"type": "STRING"},
        "phone":      {"type": "STRING", "description": "Số điện thoại"},
        "fit_score":  {"type": "INTEGER", "description": "Mức độ phù hợp với JD, 0-100"},
        "fit_summary": {"type": "STRING", "description": "Nhận xét ngắn vì sao phù hợp / chưa phù hợp"},
        "strengths":  {"type": "STRING", "description": "Ưu điểm nổi bật (gạch đầu dòng ngắn)"},
        "weaknesses": {"type": "STRING", "description": "Nhược điểm / điểm còn thiếu so với JD"},
    },
    "required": [
        "name", "dob", "email", "phone",
        "fit_score", "fit_summary", "strengths", "weaknesses",
    ],
}

# Cột xuất ra Excel: (khóa trong JSON, tiêu đề hiển thị, độ rộng cột).
_COLUMNS = [
    ("file",        "Tên file",         28),
    ("name",        "Họ tên",           22),
    ("dob",         "Ngày sinh",        14),
    ("email",       "Email",            28),
    ("phone",       "Số điện thoại",    16),
    ("fit_score",   "Điểm phù hợp",     13),
    ("fit_summary", "Đánh giá phù hợp", 46),
    ("strengths",   "Ưu điểm",          46),
    ("weaknesses",  "Nhược điểm",       46),
]


def _build_prompt(jd: str) -> str:
    """Câu lệnh gửi kèm mỗi CV. JD là yêu cầu công việc do người dùng nhập."""
    jd = jd.strip() or "(Không có mô tả công việc cụ thể — hãy đánh giá tổng quát.)"
    return (
        "Bạn là chuyên viên tuyển dụng. Hãy đọc kỹ CV trong file PDF đính kèm "
        "và trích xuất thông tin ứng viên, đồng thời đánh giá mức độ phù hợp với "
        "mô tả công việc (JD) dưới đây.\n\n"
        "=== MÔ TẢ CÔNG VIỆC (JD) ===\n"
        f"{jd}\n"
        "=== HẾT JD ===\n\n"
        "Yêu cầu:\n"
        "- Trả lời hoàn toàn bằng tiếng Việt.\n"
        "- 'fit_score' là số nguyên 0-100 thể hiện độ khớp giữa CV và JD.\n"
        "- Nếu thông tin nào không tìm thấy trong CV thì để chuỗi rỗng.\n"
        "- 'strengths' và 'weaknesses' viết mỗi ý một dòng, ngắn gọn.\n"
        "Chỉ trả về đúng đối tượng JSON theo schema đã cho."
    )


# Các mã lỗi HTTP mang tính tạm thời (quá tải / giới hạn nhịp) — nên thử lại.
_RETRY_STATUS = {429, 500, 502, 503, 504}
_MAX_RETRIES = 4          # số lần thử lại tối đa cho mỗi CV
_RETRY_BASE_DELAY = 4     # giây; chờ tăng dần: 4s, 8s, 16s…


def _call_gemini(api_key: str, model: str, jd: str, pdf_bytes: bytes,
                 timeout: int = 180, on_retry=None) -> dict:
    """Gửi 1 file PDF cho Gemini, trả về dict theo _RESPONSE_SCHEMA.

    Tự động thử lại khi gặp lỗi tạm thời (429/5xx, ví dụ HTTP 503 "high
    demand"), chờ tăng dần giữa các lần. `on_retry(attempt, wait, reason)`
    (nếu có) được gọi trước mỗi lần chờ để báo tiến trình. Ném RuntimeError
    với thông báo dễ hiểu nếu vẫn thất bại.
    """
    last_error = ""
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            return _call_gemini_once(api_key, model, jd, pdf_bytes, timeout)
        except _TransientError as exc:
            last_error = str(exc)
            if attempt == _MAX_RETRIES:
                break
            wait = _RETRY_BASE_DELAY * (2 ** (attempt - 1))
            if on_retry:
                on_retry(attempt, wait, last_error)
            time.sleep(wait)
    raise RuntimeError(f"Thất bại sau {_MAX_RETRIES} lần thử — {last_error}")


class _TransientError(Exception):
    """Lỗi tạm thời (nên thử lại): quá tải, giới hạn nhịp, mất kết nối."""


def _call_gemini_once(api_key: str, model: str, jd: str, pdf_bytes: bytes,
                      timeout: int) -> dict:
    """Một lần gọi API. Ném _TransientError nếu lỗi có thể thử lại."""
    body = {
        "contents": [{
            "parts": [
                {"text": _build_prompt(jd)},
                {"inline_data": {
                    "mime_type": "application/pdf",
                    "data": base64.b64encode(pdf_bytes).decode("ascii"),
                }},
            ],
        }],
        "generationConfig": {
            "temperature": 0.2,
            "responseMimeType": "application/json",
            "responseSchema": _RESPONSE_SCHEMA,
        },
    }
    url = _API_URL.format(model=model)
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "x-goog-api-key": api_key,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "ignore")
        msg = detail
        try:
            msg = json.loads(detail).get("error", {}).get("message", detail)
        except ValueError:
            pass
        if exc.code in _RETRY_STATUS:
            raise _TransientError(f"HTTP {exc.code}: {msg}") from exc
        raise RuntimeError(f"HTTP {exc.code}: {msg}") from exc
    except urllib.error.URLError as exc:
        # Timeout / mất mạng thường là tạm thời → cho thử lại.
        raise _TransientError(f"Lỗi kết nối mạng: {exc.reason}") from exc

    try:
        text = payload["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError):
        # Có thể bị chặn do safety hoặc phản hồi rỗng.
        reason = payload.get("promptFeedback", {}).get("blockReason")
        raise RuntimeError(
            f"Không nhận được nội dung từ mô hình"
            + (f" (bị chặn: {reason})" if reason else ""))
    try:
        return json.loads(text)
    except ValueError:
        raise RuntimeError("Mô hình trả về không đúng JSON.")


def _write_excel(rows: list[dict], out_path: str) -> None:
    """Ghi danh sách kết quả ra file .xlsx với tiêu đề in đậm + auto-width."""
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI CV Scan"

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    cell_font   = Font(name="Segoe UI", size=11)
    wrap        = Alignment(vertical="top", wrap_text=True)

    for col, (_, title, width) in enumerate(_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[c.column_letter].width = width

    for r, row in enumerate(rows, start=2):
        for col, (key, _, _) in enumerate(_COLUMNS, start=1):
            value = row.get(key, "")
            c = ws.cell(row=r, column=col, value=value)
            c.font = cell_font
            c.alignment = wrap

    ws.freeze_panes = "A2"
    wb.save(out_path)


class AiScanCvTool(BaseTool):
    name = "Quét CV bằng AI"
    description = "Dùng Gemini đọc PDF, đánh giá độ phù hợp với JD và xuất Excel."
    icon = "🤖"
    category = "Tệp & Tài liệu"
    order = 11
    action_label = "Quét CV bằng AI"
    action_style = "success"
    action_icon = "sparkles"

    def build_body(self, parent):
        self._parent = parent
        cfg = config.load(SECTION, DEFAULTS)

        gen = settings.load()

        # Hàng đầu: tiêu đề mục bên trái + chip model AI nhỏ nép ở góc phải trên.
        header = tk.Frame(parent, bg=theme.CARD_BG)
        header.pack(fill="x", pady=(2, 8))
        tk.Label(
            header, text="Đầu vào / đầu ra", bg=theme.CARD_BG, fg=theme.TEXT,
            font=(theme.FONT_FAMILY, 10, "bold"),
        ).pack(side="left")
        tk.Label(
            header, text=f"🤖  {gen['ai_model']}",
            bg=theme.SIDEBAR_BG_ACTIVE, fg=theme.ACCENT,
            font=(theme.FONT_FAMILY, 8, "bold"), padx=10, pady=4,
        ).pack(side="right")

        self.var_folder = widgets.file_row(parent, "Thư mục chứa CV (PDF)", mode="folder")
        if cfg["folder"]:
            self.var_folder.set(cfg["folder"])

        self.var_output = widgets.file_row(
            parent, "Lưu file Excel kết quả tại", mode="save")
        if cfg["output"]:
            self.var_output.set(cfg["output"])
        widgets.hint(
            parent,
            "Chọn đường dẫn file .xlsx sẽ tạo. Nếu để tên không có .xlsx "
            "hệ thống sẽ tự thêm.",
        )

        widgets.section_label(parent, "Yêu cầu công việc (JD)")
        self.jd_box = widgets.text_area(
            parent,
            "Dán mô tả công việc / tiêu chí tuyển dụng để AI chấm độ phù hợp:",
            value=cfg["jd"],
            height=8,
        )

    # ------------------------------------------------------------------ config

    def _collect(self):
        """Cấu hình riêng của tool (đường dẫn + JD). API key/model lấy từ Cài đặt."""
        return {
            "folder":  self.var_folder.get().strip(),
            "output":  self.var_output.get().strip(),
            "jd":      self.jd_box.get("1.0", "end-1c"),
        }

    # ------------------------------------------------------------------- action

    def run(self):
        if not _OPENPYXL_OK:
            messagebox.showerror(
                "Thiếu thư viện",
                "Cần cài openpyxl để xuất Excel:\n  pip install openpyxl")
            return

        cfg = self._collect()

        # API key & model là thiết lập chung (màn hình Cài đặt) — nạp vào cfg
        # để phần tiến trình/luồng nền bên dưới dùng như trước.
        gen = settings.load()
        cfg["api_key"] = gen.get("api_key", "").strip()
        cfg["model"] = gen.get("ai_model", "").strip() or settings.DEFAULTS["ai_model"]
        if not cfg["api_key"]:
            messagebox.showwarning(
                "Thiếu API key",
                "Chưa cấu hình API key Gemini.\n\n"
                "Vào màn hình ⚙️ Cài đặt (cuối thanh bên) để nhập API key.")
            return
        if not cfg["folder"] or not os.path.isdir(cfg["folder"]):
            messagebox.showwarning("Thiếu thư mục", "Vui lòng chọn thư mục chứa CV.")
            return

        out = cfg["output"].strip()
        if not out:
            messagebox.showwarning("Thiếu đường dẫn", "Vui lòng chọn nơi lưu file Excel.")
            return
        if not out.lower().endswith(".xlsx"):
            out += ".xlsx"
        out_dir = os.path.dirname(out) or "."
        if not os.path.isdir(out_dir):
            messagebox.showwarning(
                "Đường dẫn không hợp lệ", f"Thư mục không tồn tại:\n{out_dir}")
            return

        files = sorted(
            p for p in Path(cfg["folder"]).iterdir()
            if p.is_file() and p.suffix.lower() == ".pdf"
        )
        if not files:
            messagebox.showinfo(
                "Không có file", "Không tìm thấy file PDF nào trong thư mục đã chọn.")
            return

        # Chỉ lưu cấu hình RIÊNG của tool (không ghi api_key/model — đó là
        # thiết lập chung, đã lưu ở section "general").
        config.save(SECTION, {k: cfg[k] for k in DEFAULTS})
        self._open_progress(files, cfg, out)

    # ----------------------------------------------------------------- tiến trình

    def _open_progress(self, files, cfg, out_path):
        """Hộp thoại tiến trình + chạy quét trong luồng nền (không treo UI)."""
        root = self._parent.winfo_toplevel()
        dlg = tk.Toplevel(root)
        dlg.title("Đang quét CV bằng AI…")
        dlg.configure(bg=theme.CONTENT_BG)
        dlg.geometry("560x360")
        dlg.transient(root)

        tk.Label(
            dlg, text=f"Đang quét {len(files)} CV bằng {cfg['model']}",
            bg=theme.CONTENT_BG, fg=theme.TEXT,
            font=(theme.FONT_FAMILY, 13, "bold"),
        ).pack(anchor="w", padx=20, pady=(16, 4))

        bar = ttk.Progressbar(dlg, maximum=len(files), bootstyle="success")
        bar.pack(fill="x", padx=20, pady=(4, 8))

        status = tk.Label(
            dlg, text="Chuẩn bị…", bg=theme.CONTENT_BG, fg=theme.MUTED,
            font=(theme.FONT_FAMILY, 9),
        )
        status.pack(anchor="w", padx=20)

        log_box = widgets.text_area(dlg, "Nhật ký:", height=9, bg=theme.CONTENT_BG)
        state = {"cancel": False}

        def add_log(line):
            log_box.insert("end", line + "\n")
            log_box.see("end")

        btns = tk.Frame(dlg, bg=theme.CONTENT_BG)
        btns.pack(fill="x", padx=20, pady=12)
        close_btn = widgets.button(
            btns, text="Hủy", variant="neutral", icon="x",
            command=lambda: state.update(cancel=True),
        )
        close_btn.pack(side="right")

        # --- Chạy trong luồng nền, đẩy cập nhật về luồng chính qua after() ---
        def worker():
            rows, errors = [], []
            for i, p in enumerate(files, start=1):
                if state["cancel"]:
                    break
                dlg.after(0, lambda n=p.name, k=i: status.configure(
                    text=f"({k}/{len(files)}) {n}"))
                def _on_retry(attempt, wait, reason, name=p.name):
                    dlg.after(0, add_log,
                              f"… {name}: {reason} — thử lại lần {attempt} sau {wait}s")

                try:
                    data = _call_gemini(
                        cfg["api_key"], cfg["model"], cfg["jd"], p.read_bytes(),
                        on_retry=_on_retry)
                    data["file"] = p.name
                    # Chuẩn hóa mọi giá trị về chuỗi để ghi Excel gọn gàng.
                    rows.append({k: (v if v is not None else "") for k, v in data.items()})
                    dlg.after(0, add_log, f"✅ {p.name} — điểm {data.get('fit_score', '?')}")
                except Exception as exc:
                    errors.append(f"{p.name}: {exc}")
                    dlg.after(0, add_log, f"⚠ {p.name}: {exc}")
                dlg.after(0, bar.step, 1)
            dlg.after(0, finish, rows, errors)

        def finish(rows, errors):
            close_btn.configure(text="Đóng", command=dlg.destroy)
            if not rows:
                status.configure(text="Không có CV nào được xử lý.")
                if errors:
                    add_log("— Không ghi Excel vì không có kết quả hợp lệ.")
                return
            try:
                _write_excel(rows, out_path)
            except PermissionError:
                add_log(f"⚠ Không ghi được file (đang mở trong Excel?): {out_path}")
                messagebox.showerror(
                    "Không ghi được file",
                    f"Không có quyền ghi:\n{out_path}\n\nHãy đóng file nếu đang mở.")
                return
            except Exception as exc:
                add_log(f"⚠ Lỗi ghi Excel: {exc}")
                messagebox.showerror("Lỗi", f"Không xuất được Excel:\n{exc}")
                return
            status.configure(text=f"Hoàn thành — {len(rows)} CV đã lưu.")
            add_log(f"\n✅ Đã lưu {len(rows)} CV vào:\n{out_path}")
            msg = f"✅ Đã quét {len(rows)}/{len(files)} CV và lưu vào:\n{out_path}"
            if errors:
                msg += f"\n\n⚠ {len(errors)} file lỗi (xem nhật ký)."
            messagebox.showinfo("Hoàn thành", msg)

        threading.Thread(target=worker, daemon=True).start()
