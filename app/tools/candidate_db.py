"""Quản lý CV ứng viên & danh mục tuyển dụng (SQLite local).

Màn hình chính = ỨNG VIÊN:
    • Thanh tìm kiếm: từ khóa (tên/email/SĐT) + lọc vị trí + lọc trạng thái.
    • Bảng kết quả (double-click để sửa).
    • Nút: Thêm mới · Sửa · Xóa · Danh mục · Nhập từ Excel · Tải lại.

Nút "🏷️ Danh mục" mở hộp quản lý MASTER DATA gồm 3 tab:
    Bộ phận · Vị trí tuyển dụng · Mô tả công việc (JD) — mỗi tab có
    thêm/sửa/xóa riêng.

Thiết kế DB:  app/core/cv_schema.py   ·  Truy cập DB:  app/core/cv_repository.py
"""
import os
from tkinter import filedialog, messagebox

import tkinter as tk
import ttkbootstrap as ttk

from app.core.base_tool import BaseTool
from app.core import cv_repository as repo
from app.core import cv_schema
from app.ui import theme, widgets

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_NONE = "— Chưa chọn —"       # giá trị dropdown = None
_ALL_POS = "— Mọi vị trí —"

# Cột bảng ỨNG VIÊN: (khóa, tiêu đề, rộng, rộng-tối-thiểu, canh lề).
_CAND_COLUMNS = [
    ("candidate_id",   "ID",         50,  50,  "center"),
    ("full_name",      "Họ tên",     180, 130, "w"),
    ("date_of_birth",  "Ngày sinh",  95,  90,  "center"),
    ("email",          "Email",      210, 160, "w"),
    ("phone",          "SĐT",        115, 100, "w"),
    ("position_title", "Vị trí",     160, 120, "w"),
    ("department_name", "Bộ phận",   140, 110, "w"),
    ("fit_score",      "Điểm",       60,  55,  "center"),
    ("status",         "Trạng thái", 100, 90,  "center"),
    ("applied_at",     "Ngày nộp",   105, 95,  "center"),
]

# Ánh xạ tiêu đề cột Excel (tool Quét CV AI) → khóa DB của candidates.
_EXCEL_HEADER_MAP = {
    "họ tên":           "full_name",
    "ngày sinh":        "date_of_birth",
    "email":            "email",
    "số điện thoại":    "phone",
    "điểm phù hợp":     "fit_score",
    "đánh giá phù hợp": "fit_summary",
    "ưu điểm":          "strengths",
    "nhược điểm":       "weaknesses",
    "tên file":         "cv_file_path",   # lưu thẳng vào candidates
}


# ═════════════════════════ Tiện ích dùng chung ══════════════════════════

def _num(text, kind):
    """Chuyển chuỗi → int/float, trả None nếu rỗng/không hợp lệ."""
    s = str(text).strip()
    if not s:
        return None
    try:
        return int(float(s)) if kind == "int" else float(s)
    except ValueError:
        return None


def _open_entity_form(root, title, specs, current, on_save, height=660):
    """Dựng form nhập liệu tổng quát từ danh sách `specs` rồi gọi on_save(data).

    Mỗi spec là dict: {"key","label","kind"[, ...]} với kind ∈
        text | int | decimal | textarea | dropdown | choice | section
    - dropdown: cần "options" = dict {tên → id} (hoặc callable trả dict).
    - choice:   cần "choices" = list[str]; "allow_empty" (mặc định False).
    - section:  chỉ hiện tiêu đề nhóm, không nhập.
    - "required": True → bắt buộc nhập.
    """
    dlg = tk.Toplevel(root)
    dlg.title(title)
    dlg.configure(bg=theme.CONTENT_BG)
    dlg.geometry(f"600x{height}")
    dlg.transient(root)
    dlg.grab_set()

    sf = widgets.ScrollableFrame(dlg, bg=theme.CONTENT_BG)
    sf.pack(fill="both", expand=True)
    form = tk.Frame(sf.inner, bg=theme.CONTENT_BG)
    form.pack(fill="both", expand=True, padx=22, pady=18)

    getters = {}      # key → callable trả về giá trị
    required = {}      # key → nhãn (để báo lỗi)

    def cur_val(key):
        return current[key] if current is not None and current[key] is not None else ""

    for spec in specs:
        kind = spec.get("kind", "text")
        key = spec.get("key")
        label = spec.get("label", "")

        if kind == "section":
            widgets.section_label(form, label, bg=theme.CONTENT_BG)
            continue

        if spec.get("required"):
            required[key] = label

        if kind in ("text", "int", "decimal"):
            val = cur_val(key)
            var = widgets.text_row(form, label, str(val) if val != "" else "",
                                   bg=theme.CONTENT_BG)
            if kind == "text":
                getters[key] = lambda v=var: v.get().strip() or None
            else:
                getters[key] = lambda v=var, k=kind: _num(v.get(), k)

        elif kind == "textarea":
            box = widgets.text_area(form, label, value=str(cur_val(key)),
                                    height=spec.get("height", 3),
                                    bg=theme.CONTENT_BG)
            getters[key] = lambda b=box: b.get("1.0", "end-1c").strip() or None

        elif kind == "file":
            # ô nhập đường dẫn + nút Chọn file
            tk.Label(form, text=label, bg=theme.CONTENT_BG, fg=theme.TEXT,
                     font=(theme.FONT_FAMILY, 9)).pack(anchor="w", pady=(6, 4))
            row = tk.Frame(form, bg=theme.CONTENT_BG)
            row.pack(fill="x")
            var = tk.StringVar(value=str(cur_val(key)))
            ttk.Entry(row, textvariable=var).pack(side="left", fill="x",
                                                  expand=True, ipady=3)

            def _pick(v=var, ft=spec.get("filetypes")):
                p = filedialog.askopenfilename(
                    filetypes=ft or [("Tất cả", "*.*")])
                if p:
                    v.set(p)

            ttk.Button(row, text="Chọn…", bootstyle="secondary-outline",
                       command=_pick).pack(side="left", padx=(6, 0))
            getters[key] = lambda v=var: v.get().strip() or None

        elif kind == "dropdown":
            options = spec["options"]() if callable(spec["options"]) else spec["options"]
            tk.Label(form, text=label, bg=theme.CONTENT_BG, fg=theme.TEXT,
                     font=(theme.FONT_FAMILY, 9)).pack(anchor="w", pady=(6, 4))
            var = tk.StringVar()
            cur_id = current[key] if current is not None else None
            cur_name = next((n for n, i in options.items() if i == cur_id), _NONE)
            var.set(cur_name)
            ttk.Combobox(form, textvariable=var, state="readonly",
                         values=[_NONE] + list(options)).pack(fill="x", ipady=2)
            getters[key] = lambda v=var, o=options: o.get(v.get())

        elif kind == "choice":
            choices = list(spec["choices"])
            allow_empty = spec.get("allow_empty", False)
            values = ([_NONE] + choices) if allow_empty else choices
            tk.Label(form, text=label, bg=theme.CONTENT_BG, fg=theme.TEXT,
                     font=(theme.FONT_FAMILY, 9)).pack(anchor="w", pady=(6, 4))
            var = tk.StringVar()
            cur = cur_val(key)
            var.set(cur if cur else (_NONE if allow_empty else choices[0]))
            ttk.Combobox(form, textvariable=var, state="readonly",
                         values=values).pack(fill="x", ipady=2)
            getters[key] = lambda v=var: (None if v.get() == _NONE else v.get())

    acts = tk.Frame(dlg, bg=theme.CONTENT_BG)
    acts.pack(side="bottom", fill="x", padx=22, pady=14)

    def save():
        data = {k: g() for k, g in getters.items()}
        for k, lbl in required.items():
            if not data.get(k):
                messagebox.showwarning("Thiếu thông tin",
                                       f'Vui lòng nhập "{lbl}".')
                return
        # on_save trả về False → giữ form mở (vd người dùng hủy khi báo trùng).
        if on_save(data) is False:
            return
        dlg.destroy()

    ttk.Button(acts, text="💾 Lưu", bootstyle="success",
               command=save).pack(side="left", ipadx=10, ipady=3)
    ttk.Button(acts, text="Hủy", bootstyle="secondary-outline",
               command=dlg.destroy).pack(side="left", padx=(8, 0), ipady=3)


def _dept_options():
    return {d["department_name"] or f"#{d['department_id']}": d["department_id"]
            for d in repo.list_departments()}


def _position_options():
    return {p["position_title"] or f"#{p['position_id']}": p["position_id"]
            for p in repo.list_positions()}


# ═══════════════════════════ Panel MASTER DATA ══════════════════════════

class _MasterTab(tk.Frame):
    """Một tab CRUD cho bảng master (bộ phận / vị trí / JD)."""

    def __init__(self, parent, spec, on_change=None, bg=None):
        bg = bg or theme.CONTENT_BG
        super().__init__(parent, bg=bg)
        self.spec = spec
        self.on_change = on_change

        bar = tk.Frame(self, bg=bg)
        bar.pack(fill="x", padx=14, pady=(12, 8))
        ttk.Button(bar, text="➕ Thêm", bootstyle="success",
                   command=self._add).pack(side="left", ipady=2)
        ttk.Button(bar, text="✏️ Sửa", bootstyle="info-outline",
                   command=self._edit).pack(side="left", padx=(6, 0), ipady=2)
        ttk.Button(bar, text="🗑️ Xóa", bootstyle="danger-outline",
                   command=self._delete).pack(side="left", padx=(6, 0), ipady=2)
        ttk.Button(bar, text="🔄 Tải lại", bootstyle="secondary-outline",
                   command=self.reload).pack(side="right", ipady=2)

        frm = tk.Frame(self, bg=bg)
        frm.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)
        cols = [c[0] for c in spec["columns"]]
        self.tree = ttk.Treeview(frm, columns=cols, show="headings", height=11)
        last = len(spec["columns"]) - 1
        for i, (key, title, width) in enumerate(spec["columns"]):
            self.tree.heading(key, text=title)
            # min-width = width để không bị ép nhỏ; cột cuối co giãn lấp chỗ trống.
            self.tree.column(key, width=width, minwidth=width, anchor="w",
                             stretch=(i == last))
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<Double-1>", lambda _e: self._edit())

        self.reload()

    def reload(self):
        self.tree.delete(*self.tree.get_children())
        for r in self.spec["list_fn"]():
            values = [r[c[0]] if r[c[0]] is not None else "" for c in self.spec["columns"]]
            self.tree.insert("", "end", iid=str(r[self.spec["pk"]]), values=values)

    def _selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Chưa chọn", "Vui lòng chọn một dòng.")
            return None
        return int(sel[0])

    def _changed(self):
        self.reload()
        if self.on_change:
            self.on_change()

    def _add(self):
        _open_entity_form(
            self.winfo_toplevel(), "Thêm " + self.spec["title"],
            self.spec["form"], None,
            on_save=lambda data: (self.spec["insert"](data), self._changed()),
            height=self.spec.get("form_height", 520))

    def _edit(self):
        rid = self._selected()
        if rid is None:
            return
        current = self.spec["get"](rid)
        _open_entity_form(
            self.winfo_toplevel(), "Sửa " + self.spec["title"],
            self.spec["form"], current,
            on_save=lambda data: (self.spec["update"](rid, data), self._changed()),
            height=self.spec.get("form_height", 520))

    def _delete(self):
        rid = self._selected()
        if rid is None:
            return
        if messagebox.askyesno("Xác nhận xóa", f'Xóa {self.spec["title"]} #{rid}?'):
            self.spec["delete"](rid)
            self._changed()


def _master_specs():
    """Định nghĩa 3 nhóm master. Options dropdown là callable để luôn cập nhật."""
    return {
        "department": {
            "tab": "Bộ phận", "title": "bộ phận", "pk": "department_id",
            "list_fn": repo.list_departments,
            "get": repo.get_department, "insert": repo.insert_department,
            "update": repo.update_department, "delete": repo.delete_department,
            "columns": [
                ("department_id", "ID", 50),
                ("department_code", "Mã", 90),
                ("department_name", "Tên bộ phận", 200),
                ("manager_name", "Quản lý", 150),
                ("description", "Mô tả", 220),
            ],
            "form": [
                {"key": "department_code", "label": "Mã bộ phận", "kind": "text"},
                {"key": "department_name", "label": "Tên bộ phận (*)",
                 "kind": "text", "required": True},
                {"key": "manager_name", "label": "Người quản lý", "kind": "text"},
                {"key": "description", "label": "Mô tả", "kind": "textarea", "height": 3},
            ],
        },
        "position": {
            "tab": "Vị trí tuyển dụng", "title": "vị trí", "pk": "position_id",
            "list_fn": repo.list_positions,
            "get": repo.get_position, "insert": repo.insert_position,
            "update": repo.update_position, "delete": repo.delete_position,
            "columns": [
                ("position_id", "ID", 50),
                ("position_code", "Mã", 90),
                ("position_title", "Vị trí", 200),
                ("department_name", "Bộ phận", 150),
                ("level", "Cấp", 90),
                ("headcount", "SL", 55),
                ("status", "Trạng thái", 110),
            ],
            "form": [
                {"key": "department_id", "label": "Bộ phận", "kind": "dropdown",
                 "options": _dept_options},
                {"key": "position_code", "label": "Mã vị trí", "kind": "text"},
                {"key": "position_title", "label": "Tên vị trí (*)",
                 "kind": "text", "required": True},
                {"key": "level", "label": "Cấp bậc", "kind": "text"},
                {"key": "headcount", "label": "Số lượng cần tuyển", "kind": "int"},
                {"key": "status", "label": "Trạng thái", "kind": "choice",
                 "choices": cv_schema.POSITION_STATUS_CHOICES, "allow_empty": True},
            ],
        },
        "jd": {
            "tab": "Mô tả công việc (JD)", "title": "JD", "pk": "jd_id",
            "list_fn": repo.list_job_descriptions,
            "get": repo.get_job_description, "insert": repo.insert_job_description,
            "update": repo.update_job_description, "delete": repo.delete_job_description,
            "columns": [
                ("jd_id", "ID", 50),
                ("jd_title", "Tiêu đề JD", 230),
                ("position_title", "Vị trí", 180),
                ("salary_range", "Mức lương", 130),
                ("created_at", "Ngày tạo", 140),
            ],
            "form": [
                {"key": "position_id", "label": "Vị trí", "kind": "dropdown",
                 "options": _position_options},
                {"key": "jd_title", "label": "Tiêu đề JD (*)",
                 "kind": "text", "required": True},
                {"key": "summary", "label": "Tóm tắt", "kind": "textarea", "height": 3},
                {"key": "requirements", "label": "Yêu cầu", "kind": "textarea", "height": 4},
                {"key": "salary_range", "label": "Khoảng lương", "kind": "text"},
                {"key": "jd_file_path", "label": "File JD (đường dẫn trên máy)",
                 "kind": "file"},
            ],
            "form_height": 640,
        },
    }


# ═════════════════════════ TRANG MASTER DATA (sidebar) ══════════════════

class _MasterPageTool(BaseTool):
    """Lớp cha cho các trang master ở sidebar (Bộ phận / Vị trí / JD).

    Mỗi trang con chỉ cần khai báo `spec_key` trỏ vào `_master_specs()`.
    """
    category = "Master Data"
    show_on_home = False   # không hiện thẻ ở Trang chủ
    fills_height = True     # chiếm full chiều cao
    spec_key = ""

    def build(self, parent):
        repo.init_db()
        outer = tk.Frame(parent, bg=theme.CONTENT_BG)
        card = tk.Frame(outer, bg=theme.CARD_BG,
                        highlightbackground=theme.BORDER, highlightthickness=1)
        card.pack(fill="both", expand=True)
        inner = tk.Frame(card, bg=theme.CARD_BG)
        inner.pack(fill="both", expand=True, padx=22, pady=18)

        spec = _master_specs()[self.spec_key]
        _MasterTab(inner, spec, bg=theme.CARD_BG).pack(fill="both", expand=True)
        return outer

    def build_body(self, parent):  # không dùng (đã override build)
        pass


class DepartmentTool(_MasterPageTool):
    name = "Bộ phận"
    description = "Danh mục phòng ban / bộ phận."
    icon = "🏢"
    order = 10
    spec_key = "department"


class PositionTool(_MasterPageTool):
    name = "Vị trí tuyển dụng"
    description = "Danh mục vị trí cần tuyển."
    icon = "💼"
    order = 20
    spec_key = "position"


class JobDescriptionTool(_MasterPageTool):
    name = "Mô tả công việc (JD)"
    description = "Danh mục JD gắn với từng vị trí."
    icon = "📋"
    order = 30
    spec_key = "jd"


# ═══════════════════════════════ TOOL CHÍNH ═════════════════════════════

class CandidateDbTool(BaseTool):
    name = "Quản lý CV ứng viên"
    description = "Tìm kiếm ứng viên, quản lý bộ phận/vị trí/JD, nhập hàng loạt (SQLite)."
    icon = "🗂️"
    category = "Tuyển dụng"
    order = 10
    fills_height = True   # chiếm full chiều cao khi phóng to cửa sổ

    def build(self, parent):
        repo.init_db()
        outer = tk.Frame(parent, bg=theme.CONTENT_BG)
        card = tk.Frame(outer, bg=theme.CARD_BG,
                        highlightbackground=theme.BORDER, highlightthickness=1)
        card.pack(fill="both", expand=True)
        inner = tk.Frame(card, bg=theme.CARD_BG)
        inner.pack(fill="both", expand=True, padx=22, pady=20)

        self._build_search_bar(inner)
        self._build_toolbar(inner)

        # count ở đáy trước, để bảng lấp hết khoảng giữa (full height).
        self.count_lbl = tk.Label(inner, text="", bg=theme.CARD_BG, fg=theme.MUTED,
                                  font=(theme.FONT_FAMILY, 9))
        self.count_lbl.pack(side="bottom", anchor="w", pady=(8, 0))

        self._build_table(inner)

        self._reload()
        return outer

    def build_body(self, parent):  # không dùng (đã override build)
        pass

    # ------------------------------------------------------------- tìm kiếm
    def _build_search_bar(self, parent):
        widgets.section_label(parent, "Tìm kiếm ứng viên")
        bar = tk.Frame(parent, bg=theme.CARD_BG)
        bar.pack(fill="x", pady=(0, 10))

        self.var_kw = tk.StringVar()
        ent = ttk.Entry(bar, textvariable=self.var_kw)
        ent.pack(side="left", fill="x", expand=True, ipady=4)
        ent.bind("<Return>", lambda _e: self._reload())

        self.var_pos = tk.StringVar(value=_ALL_POS)
        self.cb_pos = ttk.Combobox(bar, textvariable=self.var_pos,
                                   state="readonly", width=20)
        self.cb_pos.pack(side="left", padx=(8, 0), ipady=2)

        self.var_status = tk.StringVar(value="Tất cả")
        ttk.Combobox(bar, textvariable=self.var_status, state="readonly", width=13,
                     values=["Tất cả"] + cv_schema.STATUS_CHOICES).pack(
            side="left", padx=(8, 0), ipady=2)

        ttk.Button(bar, text="🔍 Tìm", bootstyle="primary",
                   command=self._reload).pack(side="left", padx=(8, 0), ipady=2)
        ttk.Button(bar, text="✖", bootstyle="secondary-outline",
                   command=self._clear_filters, width=3).pack(
            side="left", padx=(6, 0), ipady=2)

        widgets.hint(parent, "Gõ tên / email / SĐT rồi Enter. Lọc thêm theo vị "
                             "trí và trạng thái.")

    def _build_toolbar(self, parent):
        bar = tk.Frame(parent, bg=theme.CARD_BG)
        bar.pack(fill="x", pady=(4, 10))
        ttk.Button(bar, text="➕ Thêm mới", bootstyle="success",
                   command=self._add).pack(side="left", ipady=2)
        ttk.Button(bar, text="✏️ Sửa", bootstyle="info-outline",
                   command=self._edit).pack(side="left", padx=(6, 0), ipady=2)
        ttk.Button(bar, text="🗑️ Xóa", bootstyle="danger-outline",
                   command=self._delete).pack(side="left", padx=(6, 0), ipady=2)
        ttk.Button(bar, text="📂 Mở CV", bootstyle="info-outline",
                   command=self._open_cv).pack(side="left", padx=(6, 0), ipady=2)
        ttk.Button(bar, text="📥 Nhập từ Excel", bootstyle="secondary",
                   command=self._batch_import).pack(side="left", padx=(6, 0), ipady=2)
        ttk.Button(bar, text="🔄 Tải lại", bootstyle="secondary-outline",
                   command=self._reload).pack(side="right", ipady=2)

    def _build_table(self, parent):
        frm = tk.Frame(parent, bg=theme.CARD_BG)
        frm.pack(fill="both", expand=True)
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)
        cols = [c[0] for c in _CAND_COLUMNS]
        self.tree = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        for key, title, width, minwidth, anchor in _CAND_COLUMNS:
            self.tree.heading(key, text=title)
            # minwidth → cột không bị ép nhỏ; email co giãn lấp chỗ trống còn lại.
            self.tree.column(key, width=width, minwidth=minwidth, anchor=anchor,
                             stretch=(key == "email"))
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<Double-1>", lambda _e: self._edit())

    # -------------------------------------------------------------- dữ liệu
    def _reload(self):
        pos_opts = _position_options()
        self.cb_pos.configure(values=[_ALL_POS] + list(pos_opts))
        if self.var_pos.get() not in ([_ALL_POS] + list(pos_opts)):
            self.var_pos.set(_ALL_POS)

        pos_id = pos_opts.get(self.var_pos.get())
        status = self.var_status.get()
        status = "" if status == "Tất cả" else status

        rows = repo.search_candidates(self.var_kw.get(), pos_id, status)
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            values = [r[key] if r[key] is not None else "" for key, *_ in _CAND_COLUMNS]
            self.tree.insert("", "end", iid=str(r["candidate_id"]), values=values)
        self.count_lbl.configure(
            text=f"Hiển thị {len(rows)} ứng viên · Tổng trong DB: "
                 f"{repo.count_candidates()}")

    def _clear_filters(self):
        self.var_kw.set("")
        self.var_pos.set(_ALL_POS)
        self.var_status.set("Tất cả")
        self._reload()

    def _selected_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Chưa chọn", "Vui lòng chọn một ứng viên trong bảng.")
            return None
        return int(sel[0])

    # ------------------------------------------------------------- hành động
    def _candidate_form_specs(self):
        return [
            {"kind": "section", "label": "Thông tin cá nhân"},
            {"key": "full_name", "label": "Họ và tên (*)", "kind": "text", "required": True},
            {"key": "email", "label": "Email", "kind": "text"},
            {"key": "phone", "label": "Số điện thoại", "kind": "text"},
            {"key": "date_of_birth", "label": "Ngày sinh (dd/mm/yyyy)", "kind": "text"},
            {"key": "gender", "label": "Giới tính", "kind": "choice",
             "choices": cv_schema.GENDER_CHOICES, "allow_empty": True},
            {"key": "address", "label": "Địa chỉ", "kind": "text"},
            {"kind": "section", "label": "Ứng tuyển"},
            {"key": "position_id", "label": "Vị trí ứng tuyển", "kind": "dropdown",
             "options": _position_options},
            {"key": "years_experience", "label": "Số năm kinh nghiệm", "kind": "int"},
            {"key": "education", "label": "Học vấn", "kind": "text"},
            {"key": "applied_at", "label": "Ngày nộp (yyyy-mm-dd)", "kind": "text"},
            {"key": "status", "label": "Trạng thái", "kind": "choice",
             "choices": cv_schema.STATUS_CHOICES},
            {"key": "source", "label": "Nguồn CV", "kind": "text"},
            {"key": "cv_file_path", "label": "File CV (đường dẫn trên máy)",
             "kind": "file",
             "filetypes": [("PDF/Word", "*.pdf *.doc *.docx"), ("Tất cả", "*.*")]},
            {"kind": "section", "label": "Đánh giá (từ quét CV)"},
            {"key": "fit_score", "label": "Điểm phù hợp (0-100)", "kind": "decimal"},
            {"key": "fit_summary", "label": "Nhận xét phù hợp", "kind": "textarea", "height": 3},
            {"key": "strengths", "label": "Ưu điểm", "kind": "textarea", "height": 3},
            {"key": "weaknesses", "label": "Nhược điểm", "kind": "textarea", "height": 3},
            {"key": "note", "label": "Ghi chú", "kind": "textarea", "height": 3},
        ]

    def _add(self):
        def _save(data):
            dups = repo.find_duplicates(data.get("email"), data.get("phone"))
            if dups and not self._confirm_duplicate(dups):
                return False   # người dùng hủy → giữ form mở
            repo.insert_candidate(data)
            self._reload()

        _open_entity_form(
            self.tree.winfo_toplevel(), "Thêm ứng viên mới",
            self._candidate_form_specs(), None, on_save=_save)

    def _edit(self):
        cid = self._selected_id()
        if cid is None:
            return
        current = repo.get_candidate(cid)

        def _save(data):
            dups = repo.find_duplicates(data.get("email"), data.get("phone"),
                                        exclude_id=cid)
            if dups and not self._confirm_duplicate(dups):
                return False
            repo.update_candidate(cid, data)
            self._reload()

        _open_entity_form(
            self.tree.winfo_toplevel(), "Sửa ứng viên",
            self._candidate_form_specs(), current, on_save=_save)

    @staticmethod
    def _confirm_duplicate(dups) -> bool:
        """Báo danh sách ứng viên trùng email/SĐT, hỏi có tiếp tục lưu không."""
        lines = "\n".join(
            f"  • #{d['candidate_id']} {d['full_name'] or ''}"
            f"  ({d['email'] or '—'} / {d['phone'] or '—'})"
            for d in dups[:8])
        more = "" if len(dups) <= 8 else f"\n  … và {len(dups) - 8} người khác"
        return messagebox.askyesno(
            "Có thể trùng ứng viên",
            f"Đã có {len(dups)} ứng viên trùng email hoặc số điện thoại:\n\n"
            f"{lines}{more}\n\nVẫn lưu ứng viên này?",
            icon="warning", default="no")

    def _delete(self):
        cid = self._selected_id()
        if cid is None:
            return
        row = repo.get_candidate(cid)
        name = row["full_name"] if row else f"#{cid}"
        if messagebox.askyesno("Xác nhận xóa", f'Xóa ứng viên "{name}" khỏi DB?'):
            repo.delete_candidate(cid)
            self._reload()

    # ------------------------------------------------------------- mở file CV
    def _open_cv(self):
        """Mở file CV của ứng viên đang chọn.

        Xử lý bài toán 'đường dẫn lệch' (file bị di chuyển / đổi tên): nếu
        đường dẫn lưu trong DB không còn trỏ tới file thật, mời người dùng
        định vị lại; đường dẫn mới được LƯU lại vào DB để lần sau khỏi hỏi.
        """
        cid = self._selected_id()
        if cid is None:
            return
        row = repo.get_candidate(cid)
        path = (row["cv_file_path"] or "").strip() if row else ""

        if path and os.path.isfile(path):
            self._launch(path)
            return

        # Không có / không còn đúng đường dẫn → mời chọn (định vị) lại.
        if path:
            msg = (f"Không tìm thấy file CV ở đường dẫn đã lưu:\n{path}\n\n"
                   "Có thể file đã bị di chuyển hoặc đổi tên. Chọn lại vị trí "
                   "file bây giờ?")
        else:
            msg = "Ứng viên này chưa gắn file CV. Chọn file bây giờ?"
        if not messagebox.askyesno("Không tìm thấy file", msg):
            return

        initial = os.path.basename(path) if path else ""
        new_path = filedialog.askopenfilename(
            title="Chọn lại file CV",
            initialfile=initial,
            filetypes=[("PDF/Word", "*.pdf *.doc *.docx"), ("Tất cả", "*.*")])
        if not new_path:
            return
        repo.set_cv_file_path(cid, new_path)   # cập nhật lại DB
        self._reload()
        self._launch(new_path)

    @staticmethod
    def _launch(path):
        try:
            os.startfile(path)   # Windows: mở bằng ứng dụng mặc định
        except AttributeError:
            import subprocess
            subprocess.Popen(["xdg-open", path])   # Linux (lúc dev)
        except Exception as exc:
            messagebox.showerror("Lỗi mở file", f"Không mở được file:\n{exc}")

    # ----------------------------------------------------- nhập hàng loạt Excel
    def _batch_import(self):
        if not _OPENPYXL_OK:
            messagebox.showerror("Thiếu thư viện",
                                 "Cần openpyxl để đọc Excel:\n  pip install openpyxl")
            return
        path = filedialog.askopenfilename(
            title="Chọn file Excel kết quả quét CV",
            filetypes=[("Excel", "*.xlsx"), ("Tất cả", "*.*")])
        if not path:
            return
        try:
            rows = self._read_excel(path)
        except Exception as exc:
            messagebox.showerror("Lỗi đọc file", f"Không đọc được Excel:\n{exc}")
            return
        if not rows:
            messagebox.showinfo("Trống", "Không tìm thấy dòng dữ liệu hợp lệ.")
            return
        if not messagebox.askyesno(
                "Xác nhận nhập",
                f"Đọc được {len(rows)} ứng viên từ file.\n\nNhập vào DB?"):
            return

        # Excel chỉ có TÊN file → hỏi thư mục chứa CV để lưu ĐƯỜNG DẪN ĐẦY ĐỦ,
        # nhờ vậy nút 'Mở CV' sau này bấm được. Bỏ qua nếu người dùng không chọn.
        folder = ""
        if any((r.get("cv_file_path") or "").strip() for r in rows):
            folder = filedialog.askdirectory(
                title="Thư mục chứa các file CV (bỏ qua nếu không có)") or ""

        # Phân loại: bản KHÔNG trùng nhập ngay; bản TRÙNG (email/SĐT — trong file
        # hoặc so với DB) gom lại để hỏi ở cuối.
        added = 0
        dups = []
        seen = set()   # email/SĐT đã gặp trong chính file này
        for rec in rows:
            self._apply_cv_folder(rec, folder)
            email = (rec.get("email") or "").strip().lower()
            phone = (rec.get("phone") or "").strip()
            keys = set()
            if email:
                keys.add(("e", email))
            if phone:
                keys.add(("p", phone))
            is_dup = bool(keys & seen) or bool(
                repo.find_duplicates(rec.get("email"), rec.get("phone")))
            if is_dup:
                dups.append(rec)
            else:
                repo.insert_candidate(rec)
                added += 1
                seen |= keys

        # Cuối cùng mới hỏi về các bản trùng.
        added_dup = 0
        if dups and self._confirm_import_dups(dups):
            for rec in dups:
                repo.insert_candidate(rec)
                added_dup += 1

        self._reload()
        msg = f"Đã nhập {added} ứng viên (không trùng)."
        if dups:
            if added_dup:
                msg += f"\nĐã nhập thêm {added_dup} bản trùng."
            else:
                msg += f"\nBỏ qua {len(dups)} bản trùng."
        messagebox.showinfo("Hoàn tất", msg)

    @staticmethod
    def _apply_cv_folder(rec, folder):
        """Ghép thư mục CV vào tên file để thành đường dẫn đầy đủ (nếu file có thật)."""
        fname = (rec.get("cv_file_path") or "").strip()
        if fname and folder:
            full = os.path.join(folder, fname)
            rec["cv_file_path"] = full if os.path.isfile(full) else fname

    def _confirm_import_dups(self, dups) -> bool:
        """Hiện danh sách bản trùng (họ tên/email/SĐT), hỏi có nhập hay không.

        Trả về True nếu người dùng chọn 'Vẫn nhập', False nếu 'Bỏ qua'.
        """
        root = self.tree.winfo_toplevel()
        dlg = tk.Toplevel(root)
        dlg.title("Ứng viên trùng")
        dlg.configure(bg=theme.CONTENT_BG)
        dlg.geometry("620x420")
        dlg.transient(root)
        dlg.grab_set()

        tk.Label(
            dlg, text=f"Có {len(dups)} ứng viên trùng email hoặc SĐT "
                      "(với người đã có trong DB hoặc trùng nhau trong file):",
            bg=theme.CONTENT_BG, fg=theme.TEXT, wraplength=580, justify="left",
            font=(theme.FONT_FAMILY, 10),
        ).pack(anchor="w", padx=18, pady=(16, 8))

        frm = tk.Frame(dlg, bg=theme.CONTENT_BG)
        frm.pack(fill="both", expand=True, padx=18)
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)
        cols = ("full_name", "email", "phone")
        tree = ttk.Treeview(frm, columns=cols, show="headings", height=10)
        for key, title, w in (("full_name", "Họ tên", 220),
                              ("email", "Email", 240), ("phone", "SĐT", 120)):
            tree.heading(key, text=title)
            tree.column(key, width=w, minwidth=w, anchor="w")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        for rec in dups:
            tree.insert("", "end", values=(rec.get("full_name") or "",
                                           rec.get("email") or "",
                                           rec.get("phone") or ""))

        result = {"ok": False}
        acts = tk.Frame(dlg, bg=theme.CONTENT_BG)
        acts.pack(fill="x", padx=18, pady=14)
        ttk.Button(acts, text="✅ Vẫn nhập các bản trùng", bootstyle="success",
                   command=lambda: (result.update(ok=True), dlg.destroy())).pack(
            side="left", ipadx=6, ipady=3)
        ttk.Button(acts, text="🚫 Bỏ qua", bootstyle="secondary-outline",
                   command=dlg.destroy).pack(side="left", padx=(8, 0), ipady=3)
        dlg.wait_window()
        return result["ok"]

    @staticmethod
    def _read_excel(path) -> list[dict]:
        """Đọc Excel do tool quét CV xuất → list dict theo cột DB.

        Nhận cột theo TIÊU ĐỀ ở hàng đầu (khớp _EXCEL_HEADER_MAP), không phụ
        thuộc thứ tự cột.
        """
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            wb.close()
            return []
        col_key = {}
        for idx, title in enumerate(header):
            if title is None:
                continue
            key = _EXCEL_HEADER_MAP.get(str(title).strip().lower())
            if key:
                col_key[idx] = key

        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for idx, key in col_key.items():
                if idx < len(values):
                    v = values[idx]
                    rec[key] = v if v is not None else ""
            rec["fit_score"] = _num(rec.get("fit_score", ""), "decimal")
            if (rec.get("full_name") or "").strip():
                rows.append(rec)
        wb.close()
        return rows
