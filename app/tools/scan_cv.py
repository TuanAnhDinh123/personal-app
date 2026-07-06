"""Đổi tên file CV hàng loạt và trích xuất dữ liệu CV ra Excel.

Tính năng đổi tên:
  • Quét thư mục → tìm file PDF/Word
  • Tự động trích xuất tên ứng viên từ tên file
  • Hiển thị bảng xem trước, cho phép chỉnh tên trước khi đổi
  • Đổi tên theo format: {prefixcode}_{startcode}_{TenUngVien}.ext
    start code tăng dần cho từng file theo thứ tự sắp xếp

Tính năng trích xuất:
  • Đọc nội dung từng file CV (PDF / DOCX)
  • Tách ID (mã CV) và Tên ứng viên từ tên file, dùng regex tìm Email/SĐT
  • Xuất theo template Excel có sẵn (sheet "Candidates"):
      – Chọn THƯ MỤC  → tạo file Excel mới theo template
      – Chọn FILE .xlsx → nối tiếp dữ liệu vào sheet "Candidates"
        (báo lỗi nếu file không có sheet này)
"""
import datetime
import html
import os
import re
import sys
import unicodedata
import zipfile
from pathlib import Path
from tkinter import messagebox

import tkinter as tk
import ttkbootstrap as ttk

from app.core import config
from app.core.base_tool import BaseTool
from app.ui import widgets, theme

try:
    import fitz  # PyMuPDF — đọc text từ PDF
    _FITZ_OK = True
except ImportError:
    _FITZ_OK = False

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_CV_EXTENSIONS = {".pdf", ".doc", ".docx"}

SECTION = "scan_cv"
DEFAULTS = {
    "folder":         "",
    "prefix":         "",
    "start":          "01",
    "noise_keywords": (
        "cv\n"
        "c.v\n"
        "resume\n"
        "hồ sơ\n"
        "ứng tuyển\n"
        "ứng viên\n"
        "phỏng vấn\n"
        "đơn xin việc"
    ),
}

_NOISE_NUMBER = re.compile(r"^(?:20\d{2}|\d+)$")

# ----- Template Excel (sheet "Candidates") -----
_TEMPLATE_NAME = "template_cv.xlsx"
CANDIDATES_SHEET = "Candidates"
DATA_START_ROW = 12          # dòng dữ liệu đầu tiên (tiêu đề ở dòng 11)
# Cột (1-based) trong sheet Candidates
COL_ID    = 2   # B  — ID
COL_NAME  = 3   # C  — NAME
COL_APPLY = 4   # D  — APPLYING FOR (phòng ban)
COL_EMAIL = 7   # G  — EMAIL ADDRESS
COL_PHONE = 8   # H  — PHONE


def _template_path() -> Path:
    """Đường dẫn tới file template Excel, đúng cả khi chạy dev lẫn bản .exe."""
    base = getattr(sys, "_MEIPASS", None)
    if base:
        p = Path(base) / "app" / _TEMPLATE_NAME
        if p.exists():
            return p
    return Path(__file__).resolve().parent.parent / _TEMPLATE_NAME


def _ascii_fold(s: str) -> str:
    s = s.replace("đ", "d").replace("Đ", "D")
    s = s.replace("ư", "u").replace("Ư", "U")
    s = s.replace("ơ", "o").replace("Ơ", "O")
    nfd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfd if ord(c) < 128)


def _parse_noise(raw: str) -> list[str]:
    result = []
    for line in raw.splitlines():
        for kw in line.split(","):
            kw = kw.strip()
            if kw:
                result.append(_ascii_fold(kw).lower())
    return result


def _title_vn(s: str) -> str:
    """Viết hoa chữ cái đầu mỗi từ, giữ nguyên dấu tiếng Việt."""
    return " ".join(w.capitalize() for w in s.split())


def _extract_name(stem: str, noise_list: list[str]) -> str:
    stem = unicodedata.normalize("NFC", stem)
    stem = re.sub(r"[\[\(][^\]\)]*[\]\)]", " ", stem)
    stem = re.sub(r"[_\-|/\\]+", " ", stem)

    words  = stem.split()
    folded = [_ascii_fold(w).lower() for w in words]

    max_win = max((len(kw.split()) for kw in noise_list), default=1)
    max_win = min(max_win, 4)

    keep = [True] * len(words)
    i = 0
    while i < len(words):
        removed = False
        for size in range(min(max_win, len(words) - i), 0, -1):
            chunk = " ".join(folded[i : i + size])
            if chunk in noise_list or _NOISE_NUMBER.match(chunk):
                for j in range(i, i + size):
                    keep[j] = False
                i += size
                removed = True
                break
        if not removed:
            i += 1

    raw = " ".join(w for w, k in zip(words, keep) if k)
    return _title_vn(raw)


def _build_filename(name: str, prefix: str, code: str, suffix: str) -> str:
    cv_code = (prefix + code).strip()   # prefix và startcode viết liền, không có _
    parts = [p for p in (cv_code, name) if p]
    return "_".join(parts) + suffix


def _seq_code(start_str: str, index: int) -> str:
    """Tính mã thứ tự: start_str='01', index=2 → '03'. Giữ nguyên số chữ số."""
    pad = len(start_str) if start_str.isdigit() else 2
    try:
        n = int(start_str) + index
    except ValueError:
        n = 1 + index
    return str(n).zfill(pad)


# ---------------------------------------------------------------------------
# Trích xuất dữ liệu từ CV (độc lập với UI, dễ test)
# ---------------------------------------------------------------------------

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# Số điện thoại VN: bắt đầu bằng 0, 84, +84 hoặc (+84); cho phép dấu
# cách/chấm/gạch xen giữa. Ví dụ: 0369..., +84 369..., (+84) 369...
_PHONE_RE = re.compile(r"(?:\(?\+?84\)?|0)[\d.\-\s]{8,13}")


def _read_pdf_text(path: Path) -> str:
    """Trích toàn bộ lớp text của file PDF (không OCR ảnh scan)."""
    parts = []
    with fitz.open(path) as doc:
        for page in doc:
            parts.append(page.get_text("text"))
    return "\n".join(parts)


def _read_docx_text(path: Path) -> str:
    """Trích text từ .docx bằng cách đọc word/document.xml (không cần thư viện ngoài)."""
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
    xml = xml.replace("</w:p>", "\n")          # mỗi đoạn xuống một dòng
    text = re.sub(r"<[^>]+>", " ", xml)         # bỏ toàn bộ thẻ XML
    return html.unescape(text)


def _extract_cv_text(path: Path) -> str:
    """Đọc nội dung text của một file CV theo phần mở rộng."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        if not _FITZ_OK:
            raise RuntimeError("Cần cài PyMuPDF để đọc PDF: pip install pymupdf")
        return _read_pdf_text(path)
    if suffix == ".docx":
        return _read_docx_text(path)
    # .doc cũ (định dạng nhị phân) không đọc được nếu không có Word/thư viện
    raise ValueError("Định dạng .doc cũ chưa hỗ trợ — hãy lưu lại thành .docx hoặc .pdf")


def _find_email(text: str) -> str:
    m = _EMAIL_RE.search(text)
    return m.group(0) if m else ""


def _find_phone(text: str) -> str:
    """Tìm số điện thoại VN đầu tiên hợp lệ.

    Giữ nguyên định dạng quốc tế: số viết dạng 84/+84/(+84) → trả về '+84…'
    (bỏ ngoặc và khoảng trắng). Số viết bắt đầu bằng 0 → giữ '0…'.
    """
    for m in _PHONE_RE.finditer(text):
        digits = re.sub(r"\D", "", m.group())
        if digits.startswith("84"):
            local = digits[2:]
            if 9 <= len(local) <= 10:
                return "+84" + local
        elif digits.startswith("0"):
            if 10 <= len(digits) <= 11:
                return digits
    return ""


def _split_id_name(stem: str, noise_list: list[str]) -> tuple[str, str]:
    """Tách (ID, Tên) từ tên file CV.

    File đã đổi tên có dạng '{prefix}{startcode}_{Tên ứng viên}', ví dụ
    '250601_Nguyen Van A' → ID='250601', Tên='Nguyen Van A'. Nếu phần đầu
    trước dấu phân cách không phải dãy số thì ID để trống và tên lấy từ
    toàn bộ tên file (dùng bộ lọc từ nhiễu như cũ).
    """
    norm = unicodedata.normalize("NFC", stem)
    m = re.match(r"\s*(\d{4,})[\s_\-]+(.+)", norm)
    if m:
        return m.group(1), _extract_name(m.group(2), noise_list)
    return "", _extract_name(norm, noise_list)


def _next_empty_row(ws) -> int:
    """Dòng trống đầu tiên (từ DATA_START_ROW) trong vùng dữ liệu ứng viên."""
    row = DATA_START_ROW
    while any(
        ws.cell(row=row, column=c).value not in (None, "")
        for c in (COL_ID, COL_NAME, COL_APPLY, COL_EMAIL, COL_PHONE)
    ):
        row += 1
    return row


def _write_candidates(ws, rows: list[dict]) -> None:
    """Ghi danh sách ứng viên vào sheet Candidates, nối tiếp sau dữ liệu cũ.

    Mọi ô được ghi đều chuẩn hóa về cùng một font: Aptos Display, cỡ 12,
    căn trái.
    """
    from openpyxl.styles import Alignment, Border, Font, Side

    cell_font   = Font(name="Aptos Display", size=12)
    cell_align  = Alignment(horizontal="left")
    _thin       = Side(style="thin", color="000000")
    cell_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    start = _next_empty_row(ws)
    # Các cột nằm trong khung border cho mỗi dòng (từ ID đến PHONE).
    border_cols = range(COL_ID, COL_PHONE + 1)
    field_col = {
        "id": COL_ID, "name": COL_NAME, "apply": COL_APPLY,
        "email": COL_EMAIL, "phone": COL_PHONE,
    }
    for i, r in enumerate(rows):
        row_no = start + i
        # Kẻ border cho toàn bộ ô của dòng được thêm (kể cả ô để trống).
        for col in border_cols:
            ws.cell(row=row_no, column=col).border = cell_border
        for field, col in field_col.items():
            value = r.get(field, "")
            if value:
                cell = ws.cell(row=row_no, column=col, value=value)
                cell.font = cell_font
                cell.alignment = cell_align


def _safe_filename(name: str) -> str:
    """Bỏ các ký tự không hợp lệ trong tên file Windows."""
    return re.sub(r'[<>:"/\\|?*]+', "_", name).strip() or "Candidates"


def _open_template_workbook():
    """Mở file template, trả về (workbook, worksheet Candidates)."""
    tpl = _template_path()
    if not tpl.exists():
        raise FileNotFoundError(
            f"Không tìm thấy file template:\n{tpl}")
    wb = openpyxl.load_workbook(tpl)
    if CANDIDATES_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Template thiếu sheet '{CANDIDATES_SHEET}'.")
    return wb, wb[CANDIDATES_SHEET]


def _open_existing_workbook(path: str):
    """Mở file Excel có sẵn để nối tiếp; báo lỗi nếu thiếu sheet Candidates."""
    wb = openpyxl.load_workbook(path)
    if CANDIDATES_SHEET not in wb.sheetnames:
        raise ValueError(
            f"File Excel không có sheet '{CANDIDATES_SHEET}'.\n"
            "Hãy chọn file đúng mẫu hoặc chọn một thư mục để tạo file mới.")
    return wb, wb[CANDIDATES_SHEET]


class ScanCvTool(BaseTool):
    name = "Quét CV"
    description = "Đổi tên hàng loạt file CV và trích xuất Email, Số điện thoại ra Excel."
    icon = "📇"
    category = "Tệp & Tài liệu"
    order = 10
    # Hai tính năng tách thành 2 tab riêng nên ghi đè build() thay vì dùng
    # nút hành động mặc định của BaseTool.
    def build(self, parent):
        self._parent = parent
        cfg = config.load(SECTION, DEFAULTS)

        outer = tk.Frame(parent, bg=theme.CONTENT_BG)

        # ---- Card dùng chung cho cả hai tab: thư mục + từ nhiễu ----
        shared = self._card(outer, pady=(0, 14))
        widgets.section_label(shared, "Thư mục chứa CV")
        self.var_folder = widgets.file_row(shared, "Thư mục", mode="folder")
        if cfg["folder"]:
            self.var_folder.set(cfg["folder"])

        widgets.section_label(shared, "Từ cần xóa khi trích tên ứng viên")
        self.noise_box = widgets.text_area(
            shared,
            "Mỗi từ / cụm từ một dòng (hoặc cách nhau bởi dấu phẩy):",
            value=cfg["noise_keywords"],
            height=5,
        )
        save_row = tk.Frame(shared, bg=theme.CARD_BG)
        save_row.pack(fill="x", pady=(4, 0))
        ttk.Button(
            save_row, text="💾 Lưu cấu hình", bootstyle="secondary-outline",
            command=self._save_config,
        ).pack(side="left")

        # ---- Notebook: mỗi tính năng một tab ----
        nb = ttk.Notebook(outer)
        nb.pack(fill="both", expand=True)

        page_rename,  body_rename  = self._tab_page(nb)
        page_extract, body_extract = self._tab_page(nb)
        nb.add(page_rename,  text="Đổi tên file")
        nb.add(page_extract, text="Trích xuất Excel")

        self._build_rename_tab(body_rename, cfg)
        self._build_extract_tab(body_extract)

        return outer

    # ----- Khung tiện ích -----

    def _card(self, parent, **pack):
        card = tk.Frame(
            parent, bg=theme.CARD_BG,
            highlightbackground=theme.BORDER, highlightthickness=1,
        )
        card.pack(fill="x", **pack)
        inner = tk.Frame(card, bg=theme.CARD_BG)
        inner.pack(fill="both", expand=True, padx=24, pady=18)
        return inner

    def _tab_page(self, notebook):
        """Tạo một trang tab có nền thẻ và lề trong. Trả về (page, inner)."""
        page = tk.Frame(notebook, bg=theme.CARD_BG)
        inner = tk.Frame(page, bg=theme.CARD_BG)
        inner.pack(fill="both", expand=True, padx=24, pady=20)
        return page, inner

    # ----- Tab 1: Đổi tên file -----

    def _build_rename_tab(self, parent, cfg):
        widgets.section_label(parent, "Mã CV")

        code_row = tk.Frame(parent, bg=theme.CARD_BG)
        code_row.pack(fill="x", pady=(0, 6))

        col_l = tk.Frame(code_row, bg=theme.CARD_BG)
        col_l.pack(side="left", fill="x", expand=True, padx=(0, 12))
        tk.Label(
            col_l, text="Prefix code (4 số)",
            bg=theme.CARD_BG, fg=theme.TEXT,
            font=(theme.FONT_FAMILY, 9),
        ).pack(anchor="w", pady=(0, 4))
        self.var_prefix = tk.StringVar(value=cfg["prefix"])
        widgets.digit_entry(col_l, self.var_prefix).pack(fill="x", ipady=4)

        col_r = tk.Frame(code_row, bg=theme.CARD_BG)
        col_r.pack(side="left", fill="x", expand=True)
        tk.Label(
            col_r, text="Start code (2 số, tăng dần)",
            bg=theme.CARD_BG, fg=theme.TEXT,
            font=(theme.FONT_FAMILY, 9),
        ).pack(anchor="w", pady=(0, 4))
        self.var_start = tk.StringVar(value=cfg["start"])
        widgets.digit_entry(col_r, self.var_start).pack(fill="x", ipady=4)

        widgets.hint(
            parent,
            "Ví dụ: prefix=2506, start=01  →  250601_Nguyen Van A.pdf, "
            "250602_Tran Thi B.pdf, …",
        )

        act = tk.Frame(parent, bg=theme.CARD_BG)
        act.pack(fill="x", pady=(16, 0))
        ttk.Button(
            act, text="📝 Đổi tên file CV", bootstyle="success",
            command=self.run,
        ).pack(side="left", ipadx=12, ipady=5)

    # ----- Tab 2: Trích xuất Excel -----

    def _build_extract_tab(self, parent):
        widgets.section_label(parent, "Trích xuất dữ liệu ra Excel")

        self.var_dept = widgets.text_row(
            parent, "Phòng ban (điền vào cột APPLYING FOR)")

        self.var_output = widgets.export_target_row(
            parent, "Xuất bảng tổng hợp ra")
        widgets.hint(
            parent,
            "• Chọn 📁 Thư mục → tạo file Excel mới theo template có sẵn.\n"
            "• Chọn 📄 File Excel → nối tiếp dữ liệu vào sheet 'Candidates' "
            "(báo lỗi nếu file không có sheet này).",
        )

        widgets.section_label(parent, "Trường cần lấy")
        self.chk_name  = widgets.checkbox(parent, "Tên ứng viên (không kèm ID)")
        self.chk_id    = widgets.checkbox(parent, "ID (prefix + start code từ tên file)")
        self.chk_email = widgets.checkbox(parent, "Email")
        self.chk_phone = widgets.checkbox(parent, "Số điện thoại")
        widgets.hint(
            parent,
            "Tên & ID tách từ tên file (vd '250601_Nguyen Van A' → ID=250601, "
            "Tên=Nguyen Van A) — nên đổi tên file ở tab 'Đổi tên file' trước. "
            "Email & SĐT đọc từ nội dung CV (PDF/DOCX); SĐT hỗ trợ cả dạng (+84). "
            "File .doc cũ chưa hỗ trợ — hãy lưu lại thành .docx hoặc .pdf.",
        )

        act = tk.Frame(parent, bg=theme.CARD_BG)
        act.pack(fill="x", pady=(16, 0))
        ttk.Button(
            act, text="📊 Trích xuất ra Excel", bootstyle="primary",
            command=self._run_extract,
        ).pack(side="left", ipadx=12, ipady=5)

    def build_body(self, parent):
        # Không dùng — đã ghi đè build(). Bắt buộc cài đặt vì là abstract.
        pass

    # ----------------------------------------------------------------- config

    def _collect(self):
        return {
            "folder":         self.var_folder.get().strip(),
            "prefix":         self.var_prefix.get().strip(),
            "start":          self.var_start.get().strip(),
            "noise_keywords": self.noise_box.get("1.0", "end-1c"),
        }

    def _save_config(self):
        config.save(SECTION, self._collect())
        messagebox.showinfo("Đã lưu", "Đã lưu cấu hình ✅")

    # ------------------------------------------------------- trích xuất Excel

    def _run_extract(self):
        if not _OPENPYXL_OK:
            messagebox.showerror(
                "Thiếu thư viện",
                "Cần cài openpyxl để xuất Excel:\n  pip install openpyxl")
            return

        folder = self.var_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("Thiếu thư mục", "Vui lòng chọn thư mục chứa CV.")
            return

        target = self.var_output.get().strip()
        if not target:
            messagebox.showwarning(
                "Thiếu đích xuất",
                "Vui lòng chọn thư mục (tạo file mới) hoặc file Excel (nối tiếp).")
            return

        dept       = self.var_dept.get().strip()
        want_name  = self.chk_name.get()
        want_id    = self.chk_id.get()
        want_email = self.chk_email.get()
        want_phone = self.chk_phone.get()
        if not (want_name or want_id or want_email or want_phone):
            messagebox.showwarning(
                "Chưa chọn trường", "Hãy chọn ít nhất một trường để trích xuất.")
            return

        noise = _parse_noise(self.noise_box.get("1.0", "end-1c"))
        files = sorted(
            p for p in Path(folder).iterdir()
            if p.is_file() and p.suffix.lower() in _CV_EXTENSIONS
        )
        if not files:
            messagebox.showinfo(
                "Không có file",
                "Không tìm thấy file PDF/DOC/DOCX trong thư mục đã chọn.")
            return

        config.save(SECTION, self._collect())

        rows      = []
        errors    = []
        need_text = want_email or want_phone
        for p in files:
            text = ""
            if need_text:
                try:
                    text = _extract_cv_text(p)
                except Exception as exc:
                    errors.append(f"{p.name}: {exc}")

            cv_id, cv_name = _split_id_name(p.stem, noise)
            row = {}
            if want_name:
                row["name"] = cv_name
            if want_id:
                row["id"] = cv_id
            if dept:
                row["apply"] = dept
            if want_email:
                row["email"] = _find_email(text)
            if want_phone:
                row["phone"] = _find_phone(text)
            rows.append(row)

        # ---- Xác định chế độ: tạo mới theo template hay nối tiếp file có sẵn ----
        try:
            if os.path.isdir(target):
                wb, ws = _open_template_workbook()
                stamp  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                base   = _safe_filename(f"Candidates_{dept}" if dept else "Candidates")
                out    = os.path.join(target, f"{base}_{stamp}.xlsx")
                mode   = "mới"
            elif os.path.isfile(target):
                wb, ws = _open_existing_workbook(target)
                out    = target
                mode   = "nối tiếp"
            elif target.lower().endswith(".xlsx") and os.path.isdir(os.path.dirname(target)):
                # Đường dẫn file chưa tồn tại nhưng thư mục cha hợp lệ → tạo mới.
                wb, ws = _open_template_workbook()
                out    = target
                mode   = "mới"
            else:
                messagebox.showerror(
                    "Đích không hợp lệ",
                    "Hãy chọn một THƯ MỤC (tạo file mới) hoặc một FILE .xlsx có sẵn "
                    "(nối tiếp) bằng nút '📁 Thư mục' / '📄 File Excel'.")
                return

            _write_candidates(ws, rows)
            wb.save(out)
        except Exception as exc:
            messagebox.showerror("Lỗi", f"Không xuất được Excel:\n{exc}")
            return

        msg = f"✅ Đã trích xuất {len(rows)} CV ({mode}) và lưu vào:\n{out}"
        if errors:
            preview = "\n".join(errors[:5])
            extra = f"\n(+{len(errors) - 5} file nữa)" if len(errors) > 5 else ""
            msg += f"\n\n⚠ Một số file không đọc được nội dung:\n{preview}{extra}"
        messagebox.showinfo("Hoàn thành", msg)

    # ----------------------------------------------------------------- hành động

    def run(self):
        folder = self.var_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("Thiếu thư mục", "Vui lòng chọn thư mục chứa CV.")
            return

        prefix    = self.var_prefix.get().strip()
        start_str = self.var_start.get().strip() or "01"
        noise_raw = self.noise_box.get("1.0", "end-1c")
        noise     = _parse_noise(noise_raw)

        config.save(SECTION, self._collect())

        files = sorted(
            p for p in Path(folder).iterdir()
            if p.is_file() and p.suffix.lower() in _CV_EXTENSIONS
        )
        if not files:
            messagebox.showinfo(
                "Không có file",
                "Không tìm thấy file PDF/DOC/DOCX trong thư mục đã chọn.",
            )
            return

        self._open_preview(files, prefix, start_str, noise)

    # --------------------------------------------------------------- hộp thoại

    def _open_preview(
        self, files: list, prefix: str, start_str: str, noise: list[str]
    ):
        dlg = tk.Toplevel(self._parent.winfo_toplevel())
        dlg.title("Xem trước & xác nhận đổi tên")
        dlg.configure(bg=theme.CONTENT_BG)
        dlg.geometry("1020x660")
        dlg.transient(self._parent.winfo_toplevel())
        dlg.grab_set()

        tk.Label(
            dlg,
            text=f"Tìm thấy {len(files)} file  —  double-click dòng để chỉnh tên ứng viên",
            bg=theme.CONTENT_BG, fg=theme.TEXT,
            font=(theme.FONT_FAMILY, 13, "bold"),
        ).pack(anchor="w", padx=22, pady=(16, 2))
        tk.Label(
            dlg,
            text="Tên ứng viên tự động trích từ tên file gốc. "
                 "Double-click để chỉnh, cột 'Tên file mới' cập nhật ngay.",
            bg=theme.CONTENT_BG, fg=theme.MUTED,
            font=(theme.FONT_FAMILY, 9),
        ).pack(anchor="w", padx=22, pady=(0, 10))

        # ---- Nút hành động — pack trước để luôn hiển thị ở đáy ----
        acts = tk.Frame(dlg, bg=theme.CONTENT_BG)
        acts.pack(side="bottom", fill="x", padx=22, pady=14)

        # ---- Treeview ----
        frm = tk.Frame(dlg, bg=theme.CONTENT_BG)
        frm.pack(fill="both", expand=True, padx=22)

        cols = ("original", "cname", "newname")
        tree = ttk.Treeview(frm, columns=cols, show="headings")
        tree.heading("original", text="Tên file gốc")
        tree.heading("cname",    text="Tên ứng viên  (double-click để sửa)")
        tree.heading("newname",  text="Tên file mới (xem trước)")
        tree.column("original", width=290, anchor="w", stretch=False)
        tree.column("cname",    width=230, anchor="w", stretch=False)
        tree.column("newname",  width=420, anchor="w")

        vsb = ttk.Scrollbar(frm, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        _path_map: dict[str, Path] = {}
        _code_map: dict[str, str]  = {}   # iid → mã thứ tự cố định của file đó

        for idx, p in enumerate(files):
            code = _seq_code(start_str, idx)
            name = _extract_name(p.stem, noise)
            new  = _build_filename(name, prefix, code, p.suffix)
            iid  = tree.insert("", "end", values=(p.name, name, new))
            _path_map[iid] = p
            _code_map[iid] = code

        # ---- Popup edit khi double-click ----
        def _on_double_click(event):
            region = tree.identify_region(event.x, event.y)
            iid    = tree.identify_row(event.y)
            if region != "cell" or not iid:
                return

            current = tree.set(iid, "cname")
            popup = tk.Toplevel(dlg)
            popup.title("Chỉnh tên ứng viên")
            popup.configure(bg=theme.CONTENT_BG)
            popup.resizable(False, False)
            popup.transient(dlg)
            popup.grab_set()
            popup.geometry(f"440x210+{event.x_root}+{event.y_root}")

            tk.Label(
                popup, text="Tên ứng viên:",
                bg=theme.CONTENT_BG, fg=theme.TEXT,
                font=(theme.FONT_FAMILY, 9),
            ).pack(anchor="w", padx=14, pady=(14, 4))

            var = tk.StringVar(value=current)
            entry = ttk.Entry(popup, textvariable=var, font=(theme.FONT_FAMILY, 10))
            entry.pack(fill="x", padx=14, ipady=3)
            entry.select_range(0, "end")
            entry.focus_set()

            def _save(_=None):
                new_name = var.get().strip()
                if new_name:
                    new_file = _build_filename(
                        new_name, prefix, _code_map[iid], _path_map[iid].suffix,
                    )
                    tree.set(iid, "cname",   new_name)
                    tree.set(iid, "newname", new_file)
                popup.destroy()

            entry.bind("<Return>", _save)
            entry.bind("<Escape>", lambda _: popup.destroy())

            btn_row = tk.Frame(popup, bg=theme.CONTENT_BG)
            btn_row.pack(fill="x", padx=14, pady=(18, 14))
            ttk.Button(
                btn_row, text="OK", bootstyle="primary", command=_save,
            ).pack(side="left", ipadx=10, ipady=3)
            ttk.Button(
                btn_row, text="Hủy", bootstyle="secondary-outline",
                command=popup.destroy,
            ).pack(side="left", padx=(8, 0), ipady=3)

        tree.bind("<Double-1>", _on_double_click)

        def do_rename():
            errors  = []
            renamed = 0
            skipped = 0
            for iid in tree.get_children():
                path     = _path_map[iid]
                cname    = tree.set(iid, "cname").strip()
                if not cname:
                    skipped += 1
                    continue
                new_name = _build_filename(cname, prefix, _code_map[iid], path.suffix)
                new_path = path.parent / new_name
                if new_path == path:
                    skipped += 1
                    continue
                try:
                    path.rename(new_path)
                    renamed += 1
                except Exception as exc:
                    errors.append(f"{path.name}: {exc}")

            dlg.destroy()
            msg = f"Đã đổi tên {renamed} file."
            if skipped:
                msg += f"\n{skipped} file bỏ qua (tên không thay đổi)."
            if errors:
                msg += "\n\nLỗi:\n" + "\n".join(errors)
                messagebox.showwarning("Hoàn thành (có lỗi)", msg)
            else:
                messagebox.showinfo("Hoàn thành ✅", msg)

        ttk.Button(
            acts, text="✅ Đổi tên tất cả",
            bootstyle="success", command=do_rename,
        ).pack(side="left", ipadx=14, ipady=5)
        ttk.Button(
            acts, text="Hủy",
            bootstyle="secondary-outline", command=dlg.destroy,
        ).pack(side="left", padx=(10, 0), ipady=5)
