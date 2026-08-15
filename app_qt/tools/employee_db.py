"""Quản lý nhân viên (SQLite) — bản PySide6.

Giao diện dựng theo mẫu tool "Quản lý CV ứng viên": tìm kiếm (từ khóa + lọc bộ
phận / giới tính / level) + bảng liệt kê đầy đủ cột (có checkbox) + CRUD.
Tầng dữ liệu dùng lại app.core.cv_repository (bảng `employees`).
"""
import datetime
import os
import re
import unicodedata

from PySide6.QtCore import QPoint, Qt
from PySide6.QtWidgets import (
    QCheckBox, QFileDialog, QHBoxLayout, QLabel, QLineEdit, QMenu, QVBoxLayout,
)

from app.core import application_form, config, settings
from app.core import cv_repository as repo
from app.core import cv_schema
from app_qt import dialogs, theme, widgets
from app_qt.base_tool import BaseTool
from app_qt.components.column_picker import ColumnPicker
from app_qt.components.form_dialog import FormDialog
from app_qt.components.modal import ModalDialog
from app_qt.components.progress_dialog import ProgressDialog
from app_qt.components.table import DataTable

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# Sentinel: cột trong Excel là TEXT, cần tra ra id ở bảng master trước khi ghi.
_DEPT_TEXT = "__department_short_name__"   # tra theo departments.short_name
_LEVEL_TEXT = "__level_name__"             # tra theo levels.level_name
_CC_TEXT = "__cost_center_code__"          # tra theo cost_centers.code
_ETYPE_TEXT = "__employee_type_code__"     # tra theo employee_types.code

# Map tiêu đề cột trong "Master HC file.xlsx" → field trong bảng `employees`.
# Import KHỚP THEO TÊN CỘT, KHÔNG theo thứ tự cột: mỗi ô ở dòng header được
# chuẩn hóa (`_norm`: chữ thường + gộp khoảng trắng) rồi tra ở map này, nên đổi
# vị trí cột / thêm cột lạ trong file đều không ảnh hưởng.
#
# GIÁ TRỊ là chuỗi (1 field) hoặc TUPLE khi file có NHIỀU cột TRÙNG TÊN — phần
# tử thứ n dùng cho lần xuất hiện thứ n. Ví dụ "Issued date" xuất hiện 2 lần
# (sau "ID no." rồi sau "Passport No.") → id_issued_date, passport_issued_date.
#
# Bốn cột text được tra sang id của bảng danh mục (xem `_MASTER_LOOKUPS`):
# "Function (Common)" · "New Cost center" · "IBC/DBC/WC" · "Job level".
_EXCEL_HEADER_MAP = {
    # ── định danh ──
    "ec":                             "code",
    "emp code":                       "code",
    "employee code":                  "code",
    "code":                           "code",
    "globalempcode":                  "global_code",
    "globalemp code":                 "global_code",
    "global emp code":                "global_code",
    "global code":                    "global_code",
    "global_code":                    "global_code",
    # ── họ tên ──
    "full name":                      "full_name",
    "fullname":                       "full_name",
    "surname":                        "surname",
    "name":                           "name",
    "middle name (only for vietnam)": "middle_name",
    "middle name":                    "middle_name",
    # ── thông tin cá nhân ──
    "date of birth":                  "date_of_birth",
    "dob":                            "date_of_birth",
    "gender":                         "gender",
    "place of birth":                 "place_of_birth",
    "native country":                 "native_place",
    "nationality":                    "nationality",
    "religion":                       "religion",
    "marriage status (yes)":          "marriage_status",
    "marriage status":                "marriage_status",
    "marital status":                 "marital_status",
    "spouse name":                    "spouse_name",
    "spouse date":                    "spouse_dob",
    "number of children":             "children_count",
    "children's name":                "children_names",
    "children's birthday":            "children_birthdays",
    # ── liên hệ ──
    "phone number":                   "phone",
    "phone":                          "phone",
    "personal email address":         "email",
    "email":                          "email",
    "company email":                  "company_email",
    "street (address)":               "address",
    "address":                        "address",
    "city (address)":                 "city",
    "country (address)":              "country",
    "địa chỉ thường trú":             "permanent_address",
    "địa chỉ tạm trú":                "temporary_address",
    # Ô này trong file gộp cả tên lẫn SĐT ("tên ⏎ số ĐT") — `_read_excel` tách
    # ra hai cột. File nào có sẵn cột SĐT riêng thì lấy thẳng theo tên cột.
    "emergency contact name":         "emergency_contact_name",
    "emergency contact phone":        "emergency_contact_phone",
    "emergency contact number":       "emergency_contact_phone",
    "relationship":                   "emergency_contact_relationship",
    # ── học vấn ──
    "education level":                "education",
    "education":                      "education",
    "trình độ theo lĩnh vực":         "education_field",
    "major":                          "major",
    "year of graduated":              "graduation_year",
    "school name":                    "school_name",
    "qualification":                  "qualification",
    "qualification code":             "qualification_code",
    # ── giấy tờ · ngân hàng · thuế · bảo hiểm ──
    "id no.":                         "id_no",
    "id no":                          "id_no",
    "issued date":                    ("id_issued_date", "passport_issued_date"),
    "issued place":                   "id_issued_place",
    "passport no.":                   "passport_no",
    "passport no":                    "passport_no",
    "bank account no.":               "bank_account_no",
    "bank address":                   "bank_address",
    "personal tax code":              "tax_code",
    "dependance":                     "dependants",
    "insurance book no.":             "insurance_book_no",
    # ── tổ chức & công việc ──
    "function (common)":              _DEPT_TEXT,
    "function":                       _DEPT_TEXT,
    "new cost center":                _CC_TEXT,
    "cost center":                    _CC_TEXT,
    "ibc/dbc/wc":                     _ETYPE_TEXT,
    "job level":                      _LEVEL_TEXT,
    "full name of manager":           "manager_name",
    "job title (description)":        "job_title",
    "current position":               "current_position",
    "time in position":               "time_in_position",
    "country of facility":            "facility_country",
    "town of facility":               "facility_town",
    "function (for local only)":      "local_function",
    "by group":                       "by_group",
    "type of labor":                  "labor_type",
    "production line (internal)":     "production_line",
    "operator skill":                 "operator_skill",
    "driving forklift":               "driving_forklift",
    "working hour/week":              "working_hours_per_week",
    "eligible -smart working policy eligible": "smart_working_eligible",
    "#er/ jrf":                       "er_jrf",
    "#er/jrf":                        "er_jrf",
    # ── hợp đồng & thời gian làm việc ──
    "date of employment":             "date_of_employment",
    "cột tính thâm niên":             "seniority_date",
    "permanent/temporary contract":   "contract_permanency",
    "full time/ part time":           "work_time_type",
    "full time/part time":            "work_time_type",
    "% working time":                 "working_time_pct",
    "direct/indirect":                "direct_indirect",
    "type of contract":               "contract_type",
    "starting date of contract":      "contract_start_date",
    "ending date of contract":        "contract_end_date",
    "termination date":               "termination_date",
    "reason for leaving":             "leaving_reason",
    # ── số liệu file Excel tự tính ──
    "year of service":                "years_of_service",
    "length of service":              "length_of_service",
    "year of birthday (year)":        "birth_year",
    "age":                            "age",
    "age range":                      "age_range",
    # ── ghi chú ── ("Changing date" có 2 lần: 1 ngày đơn + 1 ô lịch sử nhiều dòng)
    "changing notes":                 "changing_notes",
    "changing date":                  ("changing_date", "changing_dates"),
    "updated changing date":          "updated_changing_date",
    "note":                           "note",
}

# Tiêu đề cột CỐ TÌNH bỏ qua khi import (không báo "unrecognized column" vì đã
# biết rõ lý do bỏ): cột phụ trợ/công thức trong file, hoặc thông tin đã có
# nguồn khác nên chỉ hiển thị qua mapping bảng master thay vì lưu thẳng text.
_EXCEL_IGNORED_HEADERS = {
    "legal entity (company)",         # cố định 1 pháp nhân, không cần lưu
    "stt",                            # số thứ tự (=ROW())
    "birthday",                       # =MONTH(ngày sinh)
    "level",                          # cột số phụ trợ, trùng nghĩa "Job level"
    "job title with level (no use)",  # file ghi rõ "no use"
    "(old) phone number",             # số điện thoại cũ
    "business unit (department)",      # hiển thị qua departments (department_id)
    "business unit",
    "department",
    "bc/wc",                          # hiển thị qua employee_types.collar
    "position status",                # suy ra từ Termination Date (xem README)
    "status",
}

# Ký tự thuộc vùng Private Use Area (U+E000–U+F8FF) — font ký hiệu (Wingdings…)
# lẫn vào tiêu đề cột trong file gốc, vd 'Marital Status '. Bỏ đi để tiêu
# đề chuẩn hóa vẫn khớp map.
_PUA_RE = re.compile(r"[-]")


def _norm(text):
    """Chuẩn hóa tiêu đề cột / tên bộ phận: về chữ thường, gộp khoảng trắng.

    Chuẩn hóa Unicode về NFC trước khi so: tiêu đề tiếng Việt trong file Excel
    có thể ở dạng TỔ HỢP (NFD — "ê" = "e" + dấu mũ rời), so chuỗi thô sẽ không
    khớp với khóa trong `_EXCEL_HEADER_MAP`.
    """
    s = unicodedata.normalize("NFC", str(text))
    return " ".join(_PUA_RE.sub("", s).strip().lower().split())


def _cell_str(value):
    """Đổi 1 ô Excel → chuỗi. Ngày/giờ → 'dd/mm/yyyy'; còn lại strip chuỗi."""
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


_PHONE_SPLIT_RE = re.compile(r"[,;/\n]+")


def _normalize_phones(text):
    """Chuẩn hóa 1 ô số điện thoại → chuỗi nhiều số ngăn cách bởi "; ".

    Cột `phone` có thể chứa nhiều số (1 nhân viên nhiều SĐT) — file Excel
    thường gộp chung 1 ô, ngăn cách bởi dấu phẩy/chấm phẩy/gạch chéo/xuống
    dòng. Ghép lại đồng nhất về "; " để hiển thị & tìm kiếm (LIKE) nhất quán.
    """
    parts = [p.strip() for p in _PHONE_SPLIT_RE.split(text) if p.strip()]
    return "; ".join(parts)


# Dò dòng header THẬT trong N dòng đầu file — file gốc thường có vài dòng
# tiêu đề/logo/ghi chú phía trên (đôi khi bị ẨN) trước khi tới dòng tên cột.
_HEADER_SCAN_ROWS = 20
_HEADER_MIN_MATCHES = 3    # số cột khớp tối thiểu để coi 1 dòng là header thật
_KNOWN_HEADERS = set(_EXCEL_HEADER_MAP) | _EXCEL_IGNORED_HEADERS


def _find_header_row(ws):
    """Trả về SỐ THỨ TỰ dòng chứa tên cột thật (1-based).

    Quét `_HEADER_SCAN_ROWS` dòng đầu, chọn dòng có nhiều ô khớp tên cột đã
    biết (`_EXCEL_HEADER_MAP`/`_EXCEL_IGNORED_HEADERS`) nhất. Không dòng nào đạt
    tối thiểu `_HEADER_MIN_MATCHES` → coi dòng 1 là header (hành vi cũ, để
    không vỡ với file đơn giản không có phần tiêu đề thừa phía trên).
    """
    best_row, best_score = 1, -1
    for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS, values_only=True), start=1):
        score = sum(1 for v in row if v and _norm(v) in _KNOWN_HEADERS)
        if score > best_score:
            best_row, best_score = row_idx, score
    return best_row if best_score >= _HEADER_MIN_MATCHES else 1


# ─────────────────────────────────────────────────────────────────────────
#  BỀ RỘNG (px) CÁC CỘT BẢNG NHÂN VIÊN — chỉnh tùy ý ở đây.
#  Cột không khai báo ở đây dùng EMP_COL_WIDTH_DEFAULT.
# ─────────────────────────────────────────────────────────────────────────
EMP_COL_WIDTH_DEFAULT = 130

EMP_COL_WIDTHS = {
    "employee_id":       56,   # vừa đủ 4 ký tự (kể cả padding 8px 2 bên)
    "code":              90,
    "global_code":       100,
    "full_name":         170,
    "surname":           90,
    "middle_name":       100,
    "name":              90,
    "date_of_birth":     95,
    "gender":            70,
    "phone":             160,
    "email":             210,
    "company_email":     210,
    "level_name":        90,
    "department_name":   140,
    "address":           200,
    "permanent_address": 200,
    "temporary_address": 200,
    "city":              110,
    "country":           90,
    "work_status":       90,
    "cost_center_code":  95,
    "employee_type_code": 95,
    "collar":            100,
    "emergency_contact_name": 160,
    "emergency_contact_phone": 130,
    "children_count":    80,
    "dependants":        80,
    "age":               60,
    "working_time_pct":  95,
    "job_title":         180,
    "manager_name":      160,
    "changing_notes":    220,
    "note":              220,
}

_W = EMP_COL_WIDTHS

# Cột bảng NHÂN VIÊN: (khóa, tiêu đề, canh lề) — xếp THEO ĐÚNG THỨ TỰ cột của
# "Master HC file.xlsx" để dễ đối chiếu với file gốc. Cột lấy từ bảng danh mục
# (department_name/level_name/cost_center_code/employee_type_code/collar) và cột
# suy ra (work_status) đứng ở đúng chỗ của cột Excel tương ứng.
_EMP_COLUMN_SPECS = [
    ("employee_id",         "ID",                 "center"),
    ("code",                "Emp code",           "w"),
    ("global_code",         "Global code",        "w"),
    ("full_name",           "Full name",          "w"),
    ("surname",             "Surname",            "w"),
    ("name",                "Name",               "w"),
    ("middle_name",         "Middle name",        "w"),
    ("date_of_birth",       "Date of birth",      "center"),
    ("gender",              "Gender",             "center"),
    ("education",           "Education level",    "w"),
    ("address",             "Street (address)",   "w"),
    ("city",                "City",               "w"),
    ("country",             "Country",            "w"),
    ("phone",               "Phone",              "w"),
    ("manager_name",        "Manager",            "w"),
    ("department_name",     "Function (dept.)",   "w"),
    ("cost_center_code",    "Cost center",        "center"),
    ("cost_center_group",   "CC group",           "w"),
    ("date_of_employment",  "Date of employment", "center"),
    ("job_title",           "Job title",          "w"),
    ("facility_country",    "Country of facility", "w"),
    ("facility_town",       "Town of facility",   "w"),
    ("contract_permanency", "Perm./Temp.",        "center"),
    ("work_time_type",      "FT/PT",              "center"),
    ("working_time_pct",    "% working time",     "center"),
    ("direct_indirect",     "Direct/Indirect",    "center"),
    ("termination_date",    "Termination date",   "center"),
    ("work_status",         "Status",             "center"),
    ("leaving_reason",      "Reason for leaving", "w"),
    ("major",               "Major",              "w"),
    ("graduation_year",     "Year of graduated",  "center"),
    ("school_name",         "School name",        "w"),
    ("place_of_birth",      "Place of birth",     "w"),
    ("id_no",               "ID no.",             "w"),
    ("id_issued_date",      "ID issued date",     "center"),
    ("id_issued_place",     "ID issued place",    "w"),
    ("native_place",        "Native place",       "w"),
    ("bank_account_no",     "Bank account no.",   "w"),
    ("bank_address",        "Bank address",       "w"),
    ("tax_code",            "Personal tax code",  "w"),
    ("dependants",          "Dependants",         "center"),
    ("insurance_book_no",   "Insurance book no.", "w"),
    ("passport_no",         "Passport no.",       "w"),
    ("passport_issued_date", "Passport issued",   "center"),
    ("emergency_contact_name", "Emergency contact", "w"),
    ("emergency_contact_phone", "Emergency phone", "w"),
    ("emergency_contact_relationship", "Relationship", "w"),
    ("email",               "Personal email",     "w"),
    ("company_email",       "Company email",      "w"),
    ("permanent_address",   "Permanent address",  "w"),
    ("temporary_address",   "Temporary address",  "w"),
    ("marriage_status",     "Marriage status",    "center"),
    ("children_count",      "Children",           "center"),
    ("children_names",      "Children's names",   "w"),
    ("children_birthdays",  "Children's birthdays", "w"),
    ("religion",            "Religion",           "w"),
    ("qualification",       "Qualification",      "w"),
    ("qualification_code",  "Qualification code", "center"),
    ("level_name",          "Job level",          "center"),
    ("operator_skill",      "Operator skill",     "w"),
    ("driving_forklift",    "Driving forklift",   "center"),
    ("working_hours_per_week", "Working hour/week", "center"),
    ("production_line",     "Production line",    "w"),
    ("er_jrf",              "#ER/JRF",            "w"),
    ("contract_type",       "Type of contract",   "w"),
    ("contract_start_date", "Contract start",     "center"),
    ("contract_end_date",   "Contract end",       "center"),
    ("changing_date",       "Changing date",      "center"),
    ("marital_status",      "Marital status",     "w"),
    ("spouse_name",         "Spouse name",        "w"),
    ("spouse_dob",          "Spouse date",        "center"),
    ("nationality",         "Nationality",        "w"),
    ("years_of_service",    "Year of service",    "center"),
    ("education_field",     "Education field",    "w"),
    ("birth_year",          "Year of birthday",   "center"),
    ("collar",              "BC/WC",              "w"),
    ("employee_type_code",  "IBC/DBC/WC",         "center"),
    ("age",                 "Age",                "center"),
    ("age_range",           "Age range",          "center"),
    ("length_of_service",   "Length of service",  "w"),
    ("local_function",      "Function (local)",   "w"),
    ("by_group",            "By group",           "w"),
    ("labor_type",          "Type of labor",      "w"),
    ("smart_working_eligible", "Smart working",   "center"),
    ("changing_notes",      "Changing notes",     "w"),
    ("changing_dates",      "Changing dates",     "w"),
    ("updated_changing_date", "Updated changing", "center"),
    ("note",                "Note",               "w"),
    ("seniority_date",      "Seniority date",     "center"),
    ("time_in_position",    "Time in position",   "w"),
    ("current_position",    "Current position",   "w"),
]

# Cột bảng dạng DataTable cần: (khóa, tiêu đề, rộng, canh lề).
_EMP_COLUMNS = [(key, title, _W.get(key, EMP_COL_WIDTH_DEFAULT), align)
                for key, title, align in _EMP_COLUMN_SPECS]

# Bảng có ~90 cột → UI KHÔNG hiện hết. Đây là các cột hiện MẶC ĐỊNH; người dùng
# bật/tắt thêm ở modal "Columns" (lựa chọn được lưu lại).
_EMP_DEFAULT_COLUMNS = [
    "code", "global_code", "full_name", "date_of_birth", "gender",
    "phone", "company_email", "manager_name", "department_name",
    "cost_center_code", "job_title",
]

# Nhóm cột cho modal chọn cột: (tên nhóm, [khóa cột…]) — thứ tự nhóm & tên nhóm
# soi theo các mục của bảng `employees` trong cv_schema.py (nhưng bằng tiếng Anh
# vì đây là text người dùng NHÌN THẤY). Cột không nằm trong nhóm nào tự dồn vào
# nhóm "Other" ở cuối (xem ColumnPicker._resolve_groups) nên thêm cột mới vào
# _EMP_COLUMN_SPECS mà quên khai ở đây thì vẫn không bị mất cột.
_EMP_COLUMN_GROUPS = [
    ("Identity", [
        "employee_id", "code", "global_code",
    ]),
    ("Personal info", [
        "full_name", "surname", "middle_name", "name", "date_of_birth",
        "gender", "place_of_birth", "native_place", "nationality", "religion",
    ]),
    ("Family", [
        "marriage_status", "marital_status", "spouse_name", "spouse_dob",
        "children_count", "children_names", "children_birthdays",
    ]),
    ("Contact", [
        "phone", "email", "company_email", "address", "city", "country",
        "permanent_address", "temporary_address", "emergency_contact_name",
        "emergency_contact_phone", "emergency_contact_relationship",
    ]),
    ("Education", [
        "education", "education_field", "major", "graduation_year",
        "school_name", "qualification", "qualification_code",
    ]),
    ("ID · bank · tax · insurance", [
        "id_no", "id_issued_date", "id_issued_place", "passport_no",
        "passport_issued_date", "bank_account_no", "bank_address", "tax_code",
        "dependants", "insurance_book_no",
    ]),
    ("Organization", [
        "department_name", "cost_center_code", "cost_center_group",
        "employee_type_code", "collar", "level_name", "manager_name",
        "job_title", "current_position", "time_in_position",
        "facility_country", "facility_town", "local_function", "by_group",
        "labor_type", "production_line", "operator_skill", "driving_forklift",
        "working_hours_per_week", "smart_working_eligible", "er_jrf",
    ]),
    ("Contract & working time", [
        "date_of_employment", "seniority_date", "contract_permanency",
        "work_time_type", "working_time_pct", "direct_indirect",
        "contract_type", "contract_start_date", "contract_end_date",
        "changing_date",
    ]),
    ("Termination", [
        "work_status", "termination_date", "leaving_reason",
    ]),
    ("Figures (computed in the Excel file)", [
        "years_of_service", "length_of_service", "birth_year", "age",
        "age_range",
    ]),
    ("Notes", [
        "changing_notes", "changing_dates", "updated_changing_date", "note",
    ]),
]

# Section cấu hình để nhớ tập cột người dùng đã chọn (%APPDATA%/…/config.json).
_CFG_SECTION = "employee_db"
_CFG_COLUMNS = "visible_columns"
# Đánh số phiên bản của _EMP_DEFAULT_COLUMNS: TĂNG số này khi đổi danh sách cột
# mặc định → cấu hình đã lưu của người dùng bị bỏ qua MỘT LẦN để họ thấy ngay
# tập cột mặc định mới (không phải tự bấm "Reset to default").
_CFG_COLUMNS_VERSION = "columns_default_version"
_EMP_COLUMNS_VERSION = 2


def _dept_options():
    return {d["department_name"] or f"#{d['department_id']}": d["department_id"]
            for d in repo.list_departments()}


def _level_options():
    """Danh sách cho ô lọc/form Level: lấy theo bảng danh mục `levels` (Master
    Data → Levels), giữ đúng thứ tự sort_order. {tên hiển thị: level_id}."""
    return {r["level_name"] or f"#{r['level_id']}": r["level_id"]
            for r in repo.list_levels()}


def _cost_center_options():
    """{"VN1012 · VNPlant": cost_center_id} — ô chọn cost center trong form."""
    out = {}
    for r in repo.list_cost_centers():
        label = r["code"] or f"#{r['cost_center_id']}"
        if r["group_function"]:
            label += f" · {r['group_function']}"
        out[label] = r["cost_center_id"]
    return out


def _employee_type_options():
    """{"WC · White Collar": employee_type_id} — ô chọn loại nhân viên."""
    out = {}
    for r in repo.list_employee_types():
        label = r["code"] or f"#{r['employee_type_id']}"
        if r["collar"]:
            label += f" · {r['collar']}"
        out[label] = r["employee_type_id"]
    return out


# Bốn cột TEXT trong file Excel được tra sang id của bảng danh mục khi import.
# (sentinel trong rec, field DB, hàm nạp danh mục, cột khớp, cột id, nhãn báo lỗi)
_MASTER_LOOKUPS = (
    (_DEPT_TEXT, "department_id", repo.list_departments,
     "short_name", "department_id", "Departments (short name)"),
    (_CC_TEXT, "cost_center_id", repo.list_cost_centers,
     "code", "cost_center_id", "Cost centers"),
    (_ETYPE_TEXT, "employee_type_id", repo.list_employee_types,
     "code", "employee_type_id", "Employee types"),
    (_LEVEL_TEXT, "level_id", repo.list_levels,
     "level_name", "level_id", "Levels"),
)


class _DuplicateCodesDialog(ModalDialog):
    """Modal cảnh báo mã NV bị trùng khi import Excel (đã có sẵn trong DB).

    `dups` = list sqlite3.Row bảng `employees` (employee_id, code, full_name)
    bị trùng. Hiện trong DataTable để có thể copy (chọn ô → Ctrl+C, hoặc chuột
    phải → Copy — DataTable đã hỗ trợ sẵn). Trả về True (tiếp tục import, bỏ
    qua các dòng trùng) / False (hủy import) qua .run().
    """

    def __init__(self, parent, dups):
        super().__init__(parent, "sm")
        self._result = False
        card, lay = self.build_shell(f"Duplicate employee codes · {len(dups)}")

        desc = QLabel("These employee codes already exist in the database. "
                     "They will be skipped if you continue:")
        desc.setObjectName("DialogMsg")
        desc.setWordWrap(True)
        lay.addWidget(desc)

        rows = [{"code": d["code"] or "", "full_name": d["full_name"] or ""}
                for d in dups]
        table = DataTable([
            ("code", "Emp code", 140),
            ("full_name", "Existing employee", 240),
        ])
        table.set_rows(rows)
        table.setMinimumHeight(min(260, self.modal_h))
        lay.addWidget(table, 1)
        self.set_grow_region(table)

        hint = QLabel("Select a cell and press Ctrl+C (or right-click → Copy) "
                     "to copy the code.")
        hint.setObjectName("Hint")
        hint.setWordWrap(True)
        lay.addWidget(hint)

        foot = QHBoxLayout()
        foot.addWidget(widgets.button(card, "Import the rest, skip duplicates",
                                      variant="success", icon="check",
                                      command=lambda: self._choose(True)))
        foot.addWidget(widgets.button(card, "Cancel import", variant="neutral",
                                      icon="x", command=lambda: self._choose(False)))
        foot.addStretch(1)
        lay.addLayout(foot)

    def _choose(self, value):
        self._result = value
        self.accept()

    def run(self):
        self.exec()
        return self._result


class EmployeeDbTool(BaseTool):
    name = "Employees"
    description = "Search, manage work status, export reports."
    icon = "👥"
    category = "Human Resources"
    order = 10
    fills_height = True

    # Dựng thẳng thẻ full-height (giống CandidateDbTool) thay cho khung mặc định.
    def build(self, parent=None):
        repo.init_db()
        card = widgets.Card(parent)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(22, 20, 22, 18)
        lay.setSpacing(10)
        self._root = card

        widgets.section_label(card, "Search employees")
        self._build_search_bar(lay)

        # Dựng bảng TRƯỚC thanh nút (nút "Columns" cần tham chiếu tới bảng),
        # nhưng thêm vào layout SAU để thứ tự hiển thị vẫn là: nút → bảng.
        self.table = DataTable(_EMP_COLUMNS, pk="employee_id",
                               stretch_key="email", on_double=self._edit,
                               checkable=True)
        self._build_toolbar(lay)
        lay.addWidget(self.table, 1)

        self.count_lbl = QLabel("")
        self.count_lbl.setObjectName("Hint")
        lay.addWidget(self.count_lbl)

        self._reload()
        return card

    def build_body(self, parent):
        pass

    # -------------------------------------------------------------- tìm kiếm
    def _build_search_bar(self, lay):
        # Hàng 1: tìm theo MÃ NV — dán nguyên cột mã từ Excel (nhiều mã cách nhau
        # bởi dấu cách), khớp CHÍNH XÁC từng mã.
        self.ent_codes = QLineEdit()
        self.ent_codes.setPlaceholderText(
            "Search by employee code — paste multiple codes separated by spaces…")
        self.ent_codes.setClearButtonEnabled(True)
        self.ent_codes.addAction(widgets.svg_icon("idcard", theme.TEXT_MUTED, 16),
                                 QLineEdit.LeadingPosition)
        self.ent_codes.editingFinished.connect(self._reload)
        lay.addWidget(self.ent_codes)

        # Hàng 2: ô tìm free-text (rộng hơn) + các ô lọc select + nút đặt lại.
        filters = QHBoxLayout()
        filters.setSpacing(10)
        self.ent_kw = QLineEdit()
        self.ent_kw.setPlaceholderText("Search…")
        self.ent_kw.setClearButtonEnabled(True)
        self.ent_kw.addAction(widgets.svg_icon("search", theme.TEXT_MUTED, 16),
                              QLineEdit.LeadingPosition)
        self.ent_kw.editingFinished.connect(self._reload)
        # Ô free-text: cùng CHIỀU CAO với ô select (QLineEdit & QComboBox chung
        # QSS → cao bằng nhau), chỉ chiếm phần rộng còn lại nên NHÌN dài hơn các
        # ô select một chút. Canh giữa theo chiều dọc để thẳng hàng với ô select
        # (ô select cao 54px, ô text ~36px → tự canh giữa cho khớp).
        filters.addWidget(self.ent_kw, 1, Qt.AlignVCenter)

        # Các ô select để bề rộng CỐ ĐỊNH & bằng nhau (nếu không, combo tự giãn
        # theo nội dung dài nhất — vd tên bộ phận — nuốt hết chỗ của ô text).
        self.sel_dept = widgets.FilterSelect("Department")
        self.sel_gender = widgets.FilterSelect("Gender")
        self.sel_level = widgets.FilterSelect("Level")
        self.sel_gender.set_options(cv_schema.GENDER_CHOICES)
        # self.sel_dept / self.sel_level nạp options ở _reload() (đọc từ DB).
        for w in (self.sel_dept, self.sel_gender, self.sel_level):
            w.setFixedWidth(180)
            w.changed.connect(self._reload)
            filters.addWidget(w, 0)

        filters.addWidget(widgets.button(None, "Reset", variant="neutral",
                                         icon="eraser", command=self._clear_filters), 0)
        lay.addLayout(filters)

    def _build_toolbar(self, lay):
        """Thanh nút: chỉ để LỘ hai việc làm hằng ngày (ghi danh khóa học · nhập
        đơn dự tuyển); ba việc thi thoảng mới dùng (Add · Bulk Import · Reload)
        gom vào nút ⋮ bên phải cho thanh nút đỡ rối."""
        bar = QHBoxLayout()
        bar.setSpacing(6)
        B = widgets.button
        bar.addWidget(B(None, "Enroll", variant="info", icon="award",
                        command=self._enroll_to_course))
        bar.addWidget(B(None, "Import application form", variant="primary",
                        icon="file-text", command=self._import_forms))
        bar.addStretch(1)

        # GLOBAL SCOPE (xem cv_repository._EXCLUDE_RESIGNED_SQL): mặc định ẨN
        # người đã nghỉ việc. Tick vào đây mới gỡ scope, xem luôn cả họ.
        self.chk_include_resigned = QCheckBox("Include resigned employees")
        self.chk_include_resigned.toggled.connect(self._reload)
        bar.addWidget(self.chk_include_resigned, 0, Qt.AlignVCenter)

        bar.addWidget(self._build_column_picker())
        bar.addWidget(self._build_more_button())
        lay.addLayout(bar)

    def _build_more_button(self):
        """Nút ⋮ (vuông, không chữ) mở menu các thao tác ít dùng."""
        self.btn_more = widgets.button(None, "", variant="neutral", icon="more",
                                       command=self._show_more_menu)
        self.btn_more.setFixedWidth(36)
        self.btn_more.setToolTip("More actions")
        return self.btn_more

    def _show_more_menu(self):
        """Menu ⋮ — bung ra ngay dưới nút, canh mép PHẢI (nút nằm sát mép phải)."""
        menu = QMenu(self._root)
        menu.addAction("Add", self._add)
        menu.addAction("Bulk Import", self._batch_import)
        menu.addSeparator()
        menu.addAction("Reload", self._reload)
        corner = self.btn_more.mapToGlobal(self.btn_more.rect().bottomRight())
        menu.exec(corner - QPoint(menu.sizeHint().width(), -4))

    def _build_column_picker(self):
        """Modal tích chọn cột hiển thị; nhớ lựa chọn qua config.json."""
        cfg = config.load(_CFG_SECTION)
        # Cấu hình lưu từ phiên bản cột mặc định CŨ → bỏ qua, dùng mặc định mới.
        saved = (cfg.get(_CFG_COLUMNS)
                 if cfg.get(_CFG_COLUMNS_VERSION) == _EMP_COLUMNS_VERSION else None)

        def _save(keys):
            cfg = config.load(_CFG_SECTION)
            cfg[_CFG_COLUMNS] = list(keys)
            cfg[_CFG_COLUMNS_VERSION] = _EMP_COLUMNS_VERSION
            config.save(_CFG_SECTION, cfg)

        self.col_picker = ColumnPicker(self.table, _EMP_DEFAULT_COLUMNS,
                                       groups=_EMP_COLUMN_GROUPS, on_change=_save)
        if saved:
            self.col_picker.set_keys(saved, notify=False)
        else:
            _save(self.col_picker.keys())   # ghi luôn phiên bản mới xuống config
        return self.col_picker

    # -------------------------------------------------------------- dữ liệu
    def _reload(self):
        dept_opts = _dept_options()
        self.sel_dept.set_options(dept_opts.keys())
        dept_id = dept_opts.get(self.sel_dept.value())
        # Nạp lại danh mục Level mỗi lần tìm → thêm/sửa cấp bậc ở trang Master
        # Data → Levels là thấy ngay, không cần mở lại tool.
        level_opts = _level_options()
        self.sel_level.set_options(level_opts.keys())
        level_id = level_opts.get(self.sel_level.value())

        include_resigned = self.chk_include_resigned.isChecked()
        rows = repo.search_employees(
            self.ent_kw.text(), department_id=dept_id,
            gender=self.sel_gender.value(), level_id=level_id,
            codes=self.ent_codes.text().split(),
            include_resigned=include_resigned)
        self.table.set_rows(rows)
        self.count_lbl.setText(
            f"Showing {len(rows)} employees · Total in DB: "
            f"{repo.count_employees(include_resigned=include_resigned)}")

    def _clear_filters(self):
        self.ent_kw.clear()
        self.ent_codes.clear()
        for w in (self.sel_dept, self.sel_gender, self.sel_level):
            w.clear()
        self.chk_include_resigned.setChecked(False)
        self._reload()

    def _selected_id(self):
        eid = self.table.selected_id()
        if eid is None:
            dialogs.info(self._root, "Nothing selected", "Please select an employee in the table.")
        return eid

    # -------------------------------------------- ghi danh vào khóa học (Enroll)
    # Tick chọn nhiều nhân viên → chọn 1 khóa học trong modal → OK: thêm TẤT CẢ
    # người đã tick vào bảng course_employees (bỏ qua người đã ghi danh trước đó).
    def _enroll_to_course(self):
        rows = self.table.checked_rows()
        if not rows:
            dialogs.info(self._root, "Nothing selected",
                         "Tick at least one employee to enroll in a course.")
            return
        courses = repo.list_courses()
        if not courses:
            dialogs.info(self._root, "No courses",
                         "There are no courses to enroll employees into yet.")
            return
        course_id = self._pick_course(courses, len(rows))
        if course_id is None:
            return

        added = skipped = 0
        for row in rows:
            eid = row["employee_id"]
            if eid is None:
                continue
            rid = repo.enroll_employee(
                course_id, eid, {"status": cv_schema.COURSE_STATUS_CHOICES[0]})
            if rid:
                added += 1
            else:
                skipped += 1   # đã ghi danh trước đó (unique course_id+employee_id)

        title = next((c["title"] for c in courses
                      if c["course_id"] == course_id), "") or f"#{course_id}"
        msg = f'Enrolled {added} employees in "{title}".'
        if skipped:
            msg += f"\nSkipped {skipped} already enrolled."
        dialogs.success(self._root, "Done", msg)

    def _pick_course(self, courses, n_selected):
        """Modal chọn 1 khóa học. Trả về course_id đã chọn, hoặc None nếu hủy."""
        dlg = ModalDialog(self._root, "sm")
        card, lay = dlg.build_shell("Enroll to course")

        info = QLabel(f"Enroll {n_selected} selected employees in a course:")
        info.setObjectName("DialogMsg")
        info.setWordWrap(True)
        lay.addWidget(info)

        lbl = QLabel("Course")
        lbl.setObjectName("FieldLabel")
        lay.addWidget(lbl)
        combo = widgets.ComboBox(card)
        combo.addItems([self._course_label(c) for c in courses])
        lay.addWidget(combo)
        lay.addStretch(1)

        result = {"id": None}

        def _ok():
            i = combo.currentIndex()
            if 0 <= i < len(courses):
                result["id"] = courses[i]["course_id"]
            dlg.accept()

        foot = QHBoxLayout()
        foot.addWidget(widgets.button(card, "OK", variant="success", icon="check",
                                      command=_ok))
        foot.addWidget(widgets.button(card, "Cancel", variant="neutral", icon="x",
                                      command=dlg.reject))
        foot.addStretch(1)
        lay.addLayout(foot)
        dlg.exec()
        return result["id"]

    @staticmethod
    def _course_label(c):
        title = (c["title"] or "").strip() or f"#{c['course_id']}"
        date = (str(c["date"]).strip() if c["date"] else "")
        return f"{title} · {date}" if date else title

    # ----------------------------------------------------- nhập hàng loạt Excel
    def _batch_import(self):
        if not _OPENPYXL_OK:
            dialogs.error(self._root, "Missing library",
                          "openpyxl is required to read Excel:\n  pip install openpyxl")
            return
        path, _ = QFileDialog.getOpenFileName(
            self._root, "Choose the employee list Excel file", "",
            "Excel (*.xlsx *.xlsm);;All files (*.*)")
        if not path:
            return
        try:
            rows, unknown = self._read_excel(path)
        except Exception as exc:
            dialogs.error(self._root, "Read error", f"Couldn't read Excel:\n{exc}")
            return
        if not rows:
            dialogs.info(self._root, "Empty", "No valid data rows found.")
            return

        note = ""
        if unknown:
            note = ("\n\nUnrecognized columns (skipped): "
                    + ", ".join(unknown[:10])
                    + (" …" if len(unknown) > 10 else ""))
        if not dialogs.confirm(
                self._root, "Confirm import",
                f"Found {len(rows)} employees in the file.\n\nImport into the DB?{note}",
                ok_label="Import"):
            return

        # Mã NV (`code`) đã có sẵn trong DB → hỏi lại trước khi ghi trùng.
        dup_rows = repo.find_employees_by_codes([r.get("code") for r in rows])
        skip_codes = set()
        if dup_rows:
            if not _DuplicateCodesDialog(self._root, dup_rows).run():
                return
            skip_codes = {_norm(d["code"]) for d in dup_rows if d["code"]}

        # Tra 4 cột text sang id của bảng danh mục (bộ phận theo short_name,
        # cost center & loại NV theo code, cấp bậc theo tên) — khớp không phân
        # biệt hoa/thường, bỏ khoảng trắng thừa.
        lookups = []
        for sentinel, field, loader, match_col, id_col, label in _MASTER_LOOKUPS:
            table = {_norm(r[match_col]): r[id_col] for r in loader()
                     if r[match_col]}
            lookups.append((sentinel, field, table, label, set()))

        added = 0
        skipped = 0
        for rec in rows:
            if skip_codes and _norm(rec.get("code", "")) in skip_codes:
                skipped += 1
                continue
            for sentinel, field, table, _label, missing in lookups:
                text = rec.pop(sentinel, "")
                if not text:
                    continue
                row_id = table.get(_norm(text))
                if row_id is not None:
                    rec[field] = row_id
                else:
                    missing.add(text)   # không có trong danh mục → để trống link
            repo.insert_employee(rec)
            added += 1

        self._reload()
        msg = f"Imported {added} employees."
        if skipped:
            msg += f"\n\nSkipped {skipped} duplicate employee code(s)."
        for _sentinel, _field, _table, label, missing in lookups:
            if not missing:
                continue
            names = ", ".join(sorted(missing)[:10])
            more = " …" if len(missing) > 10 else ""
            msg += (f"\n\nNot found in master data · {label} "
                    f"(link left empty): {names}{more}\nAdd them on the matching "
                    "Master Data page and re-import if you need the link.")
        dialogs.success(self._root, "Done", msg)

    @staticmethod
    def _read_excel(path):
        """Đọc file Excel → (list rec, list tiêu đề cột không nhận diện được).

        KHỚP THEO TÊN CỘT (không theo thứ tự cột): mỗi ô header được chuẩn hóa
        rồi tra `_EXCEL_HEADER_MAP`. Mỗi rec là dict {field DB → giá trị}, riêng
        4 cột danh mục (bộ phận · cost center · loại NV · cấp bậc) giữ TEXT dưới
        khóa sentinel để tra ra id ở bước sau (xem `_batch_import`). Nếu thiếu
        full_name thì ghép từ surname + middle_name + name (thứ tự tên tiếng Việt).

        File có vài cột TRÙNG TÊN ("Issued date", "Changing date") → giá trị map
        là tuple, lần xuất hiện thứ n lấy phần tử thứ n (phần tử cuối dùng lại
        nếu file có nhiều lần hơn).

        Dòng header KHÔNG chắc luôn ở dòng 1 — file gốc thường có vài dòng tiêu
        đề/logo/ghi chú phía trên (có thể bị ẨN) trước khi tới dòng tên cột thật
        → dò tìm dòng đó bằng `_find_header_row` thay vì đinh cứng dòng 1.
        """
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = _find_header_row(ws)
        header = next(
            ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
        if not header:
            wb.close()
            return [], []

        col_key = {}       # chỉ số cột → field DB
        unknown = []       # tiêu đề không map được (để báo lại)
        seen = {}          # tiêu đề đã gặp → số lần (xử lý cột trùng tên)
        for idx, title in enumerate(header):
            if title is None or not str(title).strip():
                continue
            norm_title = _norm(title)
            key = _EXCEL_HEADER_MAP.get(norm_title)
            if key:
                nth = seen.get(norm_title, 0)
                seen[norm_title] = nth + 1
                if isinstance(key, tuple):
                    key = key[min(nth, len(key) - 1)]
                col_key[idx] = key
            elif norm_title not in _EXCEL_IGNORED_HEADERS:
                unknown.append(str(title).strip())

        rows = []
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            rec = {}
            for idx, key in col_key.items():
                if idx < len(values):
                    v = _cell_str(values[idx])
                    if v == "":
                        continue
                    rec[key] = v
            if rec.get("phone"):
                rec["phone"] = _normalize_phones(rec["phone"])
            # Ô "Emergency Contact Name" gộp "tên ⏎ số ĐT" → tách sang 2 cột
            # (dùng chung logic với lượt di trú dữ liệu cũ trong cv_schema).
            if rec.get("emergency_contact_name") and not rec.get("emergency_contact_phone"):
                name, phone = repo.split_contact_name_phone(rec["emergency_contact_name"])
                rec["emergency_contact_name"] = name
                if phone:
                    rec["emergency_contact_phone"] = phone
                if not name:
                    rec.pop("emergency_contact_name")
            if not rec.get("full_name"):
                parts = [rec.get("surname"), rec.get("middle_name"), rec.get("name")]
                composed = " ".join(p for p in parts if p)
                if composed:
                    rec["full_name"] = composed
            # Bỏ dòng rỗng hoàn toàn (không có định danh nào).
            if any(rec.get(k) for k in ("full_name", "code", "global_code", "email")):
                rows.append(rec)
        wb.close()
        return rows, unknown

    # ------------------------------------- nhập từ ĐƠN DỰ TUYỂN (AI đọc form)
    # Nhân viên mới tự điền "DLVN Application Form": hoặc gõ thẳng vào file Excel
    # mẫu (phần họ điền là chữ màu), hoặc in ra viết tay rồi HR scan thành PDF.
    # Gemini đọc file → app.core.application_form trả về dict đúng cột bảng
    # `employees`; HR xem lại/sửa trong form nhập liệu rồi mới ghi xuống DB (AI
    # đọc chữ viết tay không phải lúc nào cũng đúng — luôn cần người duyệt).
    def _import_forms(self):
        gen = settings.load()
        api_key = gen.get("api_key", "").strip()
        model = gen.get("ai_model", "").strip() or settings.DEFAULTS["ai_model"]
        if not api_key:
            dialogs.error(self._root, "Missing API key",
                          "No Gemini API key configured.\n\nOpen ⚙️ Settings "
                          "(bottom of the sidebar) to add one.")
            return

        paths, _ = QFileDialog.getOpenFileNames(
            self._root, "Choose the filled application form(s)", "",
            "Application form (*.xlsx *.xlsm *.pdf);;All files (*.*)")
        if not paths:
            return
        bad = [os.path.basename(p) for p in paths
               if not p.lower().endswith(application_form.SUPPORTED_EXTS)]
        if bad:
            dialogs.error(self._root, "Unsupported file",
                          "An application form must be the filled Excel file "
                          "(.xlsx/.xlsm) or a scan of the printed form (.pdf).\n\n"
                          + ", ".join(bad))
            return

        total = len(paths)

        def job(ctx):
            # Đọc TUẦN TỰ từng đơn; một đơn lỗi (hết hạn mức key, file hỏng) chỉ
            # bỏ qua đơn đó rồi đọc tiếp, cuối lượt báo lại danh sách lỗi.
            results, errors = [], []
            for i, path in enumerate(paths, start=1):
                if ctx.cancelled:
                    break
                name = os.path.basename(path)
                ctx.status(f"({i}/{total}) {name}")

                def on_retry(attempt, wait, reason, n=name):
                    ctx.log(f"… {n}: {reason} — retry {attempt} in {wait}s")

                try:
                    rec = application_form.extract(
                        api_key, model, path, on_retry=on_retry,
                        should_cancel=lambda: ctx.cancelled)
                except application_form.Cancelled:
                    ctx.log(f"✋ Cancelled while reading {name}.")
                    break
                except Exception as exc:                       # noqa: BLE001
                    errors.append(f"{name}: {exc}")
                    ctx.log(f"⛔ {name}: {exc}")
                    ctx.step()
                    continue
                results.append((name, rec))
                ctx.log(f"✅ {name} — {rec.get('full_name') or 'no name found'} "
                        f"({len(rec)} fields)")
                ctx.step()
            return results, errors

        def on_finish(dlg, result):
            results, errors = result
            dlg.set_final_status(f"Read {len(results)}/{total} application form(s).")
            if errors:
                dlg.log("\n⚠ Couldn't read:\n" + "\n".join(errors))
            if results:
                dlg.log("\n👉 Review each employee in the form that opens next, "
                        "then press Save to add them to the database.")
                self._review_forms(results)

        dlg = ProgressDialog(self._root, "Reading application forms with AI…",
                             total=total, subtitle=f"Reading {total} form(s) with {model}")
        dlg.start(job, on_finish)

    def _review_forms(self, results):
        """Mở form nhập liệu điền sẵn cho TỪNG đơn đã đọc để HR duyệt rồi lưu."""
        added = 0
        for i, (name, rec) in enumerate(results, start=1):
            saved = FormDialog(
                self._root, f"Review employee {i}/{len(results)} · {name}",
                self._employee_form_specs(), rec, on_save=self._save_from_form,
                size="lg").run()
            if saved:
                added += 1
        self._reload()
        dialogs.success(self._root, "Done",
                        f"Added {added} of {len(results)} employees "
                        "from the application forms.")

    def _save_from_form(self, data):
        """Ghi 1 nhân viên đọc từ đơn dự tuyển; cảnh báo nếu đã có người trùng.

        Trả về False để FormDialog giữ form mở khi người dùng hủy ở cảnh báo trùng.
        """
        dups = repo.find_employees_by_identity(
            data.get("code"), data.get("id_no"), data.get("full_name"))
        if dups:
            listing = "\n".join(
                f"• #{d['employee_id']} {d['full_name'] or ''}"
                f"{' · ' + d['code'] if d['code'] else ''}"
                f"{' · ID ' + d['id_no'] if d['id_no'] else ''}" for d in dups[:10])
            if not dialogs.confirm(
                    self._root, "Employee may already exist",
                    "The database already has someone with the same name, "
                    f"employee code or ID number:\n\n{listing}\n\nAdd this "
                    "employee anyway?", ok_label="Add anyway"):
                return False
        repo.insert_employee(data)
        self._reload()

    # ------------------------------------------------------------- form specs
    def _employee_form_specs(self):
        """Toàn bộ field của bảng `employees`, nhóm theo mục (form cuộn được).

        Thứ tự nhóm khớp với cv_schema.py để dễ đối chiếu với "Master HC file".
        Trạng thái làm việc không có field riêng — điền/xóa "Termination date"
        là đủ (có ngày = đã nghỉ việc).
        """
        return [
            {"kind": "section", "label": "Identity"},
            {"key": "code", "label": "Employee code (EC)", "kind": "text"},
            {"key": "global_code", "label": "Global code", "kind": "text"},

            {"kind": "section", "label": "Personal info"},
            {"key": "full_name", "label": "Full name (*)", "kind": "text", "required": True},
            {"key": "surname", "label": "Surname", "kind": "text"},
            {"key": "middle_name", "label": "Middle name", "kind": "text"},
            {"key": "name", "label": "Name", "kind": "text"},
            {"key": "date_of_birth", "label": "Date of birth (dd/mm/yyyy)", "kind": "text"},
            {"key": "gender", "label": "Gender", "kind": "choice",
             "choices": cv_schema.GENDER_CHOICES, "allow_empty": True},
            {"key": "place_of_birth", "label": "Place of birth", "kind": "text"},
            {"key": "native_place", "label": "Native place", "kind": "text"},
            {"key": "nationality", "label": "Nationality", "kind": "text"},
            {"key": "religion", "label": "Religion", "kind": "text"},

            {"kind": "section", "label": "Family"},
            {"key": "marriage_status", "label": "Marriage status", "kind": "choice",
             "choices": cv_schema.YES_NO_CHOICES, "allow_empty": True},
            {"key": "marital_status", "label": "Marital status", "kind": "choice",
             "choices": cv_schema.MARITAL_STATUS_CHOICES, "allow_empty": True},
            {"key": "spouse_name", "label": "Spouse name", "kind": "text"},
            {"key": "spouse_dob", "label": "Spouse date (dd/mm/yyyy)", "kind": "text"},
            {"key": "children_count", "label": "Number of children", "kind": "int"},
            {"key": "children_names", "label": "Children's names (one per line)",
             "kind": "textarea", "height": 3},
            {"key": "children_birthdays", "label": "Children's birthdays (one per line)",
             "kind": "textarea", "height": 3},

            {"kind": "section", "label": "Contact"},
            {"key": "phone", "label": 'Phone (separate multiple with "; ")', "kind": "text"},
            {"key": "email", "label": "Personal email", "kind": "text"},
            {"key": "company_email", "label": "Company email", "kind": "text"},
            {"key": "address", "label": "Street (address)", "kind": "text"},
            {"key": "city", "label": "City (address)", "kind": "text"},
            {"key": "country", "label": "Country (address)", "kind": "text"},
            {"key": "permanent_address", "label": "Permanent address", "kind": "text"},
            {"key": "temporary_address", "label": "Temporary address", "kind": "text"},
            {"key": "emergency_contact_name", "label": "Emergency contact name",
             "kind": "text"},
            {"key": "emergency_contact_phone", "label": "Emergency contact phone",
             "kind": "text"},
            {"key": "emergency_contact_relationship", "label": "Relationship",
             "kind": "text"},

            {"kind": "section", "label": "Education"},
            {"key": "education", "label": "Education level", "kind": "text"},
            {"key": "education_field", "label": "Education field", "kind": "text"},
            {"key": "major", "label": "Major (one per line)", "kind": "textarea",
             "height": 3},
            {"key": "graduation_year", "label": "Year of graduated", "kind": "text"},
            {"key": "school_name", "label": "School name", "kind": "text"},
            {"key": "qualification", "label": "Qualification", "kind": "text"},
            {"key": "qualification_code", "label": "Qualification code", "kind": "text"},

            {"kind": "section", "label": "ID · bank · tax · insurance"},
            {"key": "id_no", "label": "ID no.", "kind": "text"},
            {"key": "id_issued_date", "label": "ID issued date (dd/mm/yyyy)",
             "kind": "text"},
            {"key": "id_issued_place", "label": "ID issued place", "kind": "text"},
            {"key": "passport_no", "label": "Passport no.", "kind": "text"},
            {"key": "passport_issued_date",
             "label": "Passport issued date (dd/mm/yyyy)", "kind": "text"},
            {"key": "bank_account_no", "label": "Bank account no.", "kind": "text"},
            {"key": "bank_address", "label": "Bank address", "kind": "text"},
            {"key": "tax_code", "label": "Personal tax code", "kind": "text"},
            {"key": "dependants", "label": "Dependants", "kind": "int"},
            {"key": "insurance_book_no", "label": "Insurance book no.", "kind": "text"},

            {"kind": "section", "label": "Organization"},
            {"key": "department_id", "label": "Function (department)",
             "kind": "dropdown", "options": _dept_options},
            {"key": "cost_center_id", "label": "New cost center",
             "kind": "dropdown", "options": _cost_center_options},
            {"key": "employee_type_id", "label": "Employee type (IBC/DBC/WC)",
             "kind": "dropdown", "options": _employee_type_options},
            {"key": "level_id", "label": "Job level", "kind": "dropdown",
             "options": _level_options},
            {"key": "manager_name", "label": "Full name of manager", "kind": "text"},
            {"key": "job_title", "label": "Job title (description)", "kind": "text"},
            {"key": "current_position", "label": "Current position", "kind": "text"},
            {"key": "time_in_position", "label": "Time in position", "kind": "text"},
            {"key": "facility_country", "label": "Country of facility", "kind": "text"},
            {"key": "facility_town", "label": "Town of facility", "kind": "text"},
            {"key": "local_function", "label": "Function (for local only)",
             "kind": "text"},
            {"key": "by_group", "label": "By group", "kind": "text"},
            {"key": "labor_type", "label": "Type of labor", "kind": "text"},
            {"key": "production_line", "label": "Production line (internal)",
             "kind": "text"},
            {"key": "operator_skill", "label": "Operator skill", "kind": "text"},
            {"key": "driving_forklift", "label": "Driving forklift", "kind": "text"},
            {"key": "working_hours_per_week", "label": "Working hour/week",
             "kind": "text"},
            {"key": "smart_working_eligible",
             "label": "Smart working policy eligible", "kind": "text"},
            {"key": "er_jrf", "label": "#ER/JRF", "kind": "text"},

            {"kind": "section", "label": "Contract & working time"},
            {"key": "date_of_employment", "label": "Date of employment (dd/mm/yyyy)",
             "kind": "text"},
            {"key": "seniority_date", "label": "Seniority date (dd/mm/yyyy)",
             "kind": "text"},
            {"key": "contract_permanency", "label": "Permanent/Temporary contract",
             "kind": "choice", "choices": cv_schema.CONTRACT_PERMANENCY_CHOICES,
             "allow_empty": True},
            {"key": "work_time_type", "label": "Full time / Part time",
             "kind": "choice", "choices": cv_schema.WORK_TIME_TYPE_CHOICES,
             "allow_empty": True},
            {"key": "working_time_pct", "label": "% working time", "kind": "text"},
            {"key": "direct_indirect", "label": "Direct/Indirect", "kind": "choice",
             "choices": cv_schema.DIRECT_INDIRECT_CHOICES, "allow_empty": True},
            {"key": "contract_type", "label": "Type of contract", "kind": "text"},
            {"key": "contract_start_date",
             "label": "Starting date of contract (dd/mm/yyyy)", "kind": "text"},
            {"key": "contract_end_date",
             "label": "Ending date of contract (dd/mm/yyyy)", "kind": "text"},
            {"key": "changing_date", "label": "Changing date (dd/mm/yyyy)",
             "kind": "text"},

            {"kind": "section", "label": "Termination (filled = resigned)"},
            {"key": "termination_date", "label": "Termination date (dd/mm/yyyy)",
             "kind": "text"},
            {"key": "leaving_reason", "label": "Reason for leaving", "kind": "text"},

            {"kind": "section", "label": "Figures (computed in the Excel file)"},
            {"key": "years_of_service", "label": "Year of service", "kind": "text"},
            {"key": "length_of_service", "label": "Length of service", "kind": "text"},
            {"key": "birth_year", "label": "Year of birthday", "kind": "text"},
            {"key": "age", "label": "Age", "kind": "text"},
            {"key": "age_range", "label": "Age range", "kind": "text"},

            {"kind": "section", "label": "Notes"},
            {"key": "changing_notes", "label": "Changing notes", "kind": "textarea",
             "height": 3},
            {"key": "changing_dates", "label": "Changing dates (one per line)",
             "kind": "textarea", "height": 3},
            {"key": "updated_changing_date",
             "label": "Updated changing date (dd/mm/yyyy)", "kind": "text"},
            {"key": "note", "label": "Note", "kind": "textarea", "height": 3},
        ]

    def _add(self):
        def _save(data):
            repo.insert_employee(data)
            self._reload()

        FormDialog(self._root, "Add employee", self._employee_form_specs(), None,
                   on_save=_save, size="lg").run()

    def _edit(self, eid=None):
        if eid is None:
            eid = self._selected_id()
        if eid is None:
            return
        current = repo.get_employee(eid)

        def _save(data):
            repo.update_employee(eid, data)
            self._reload()

        FormDialog(self._root, "Edit employee",
                   self._employee_form_specs(), current, on_save=_save,
                   on_delete=lambda: self._delete(eid), size="lg").run()

    def _delete(self, eid):
        """Xóa nhân viên; trả về False nếu người dùng hủy xác nhận (giữ form mở)."""
        row = repo.get_employee(eid)
        name = (row["full_name"] if row and row["full_name"] else f"#{eid}")
        if not dialogs.confirm(self._root, "Confirm delete",
                               f'Delete employee "{name}" from the DB?', ok_label="Delete"):
            return False
        repo.delete_employee(eid)
        self._reload()
        return True
