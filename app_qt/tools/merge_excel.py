"""Tool 'Gộp file Excel' — bản PySide6.

Logic gộp (openpyxl) tách hẳn khỏi UI ở cuối file, y như bản Tk. build_body()
chỉ dựng ô nhập; run() đọc giá trị rồi gọi logic.
"""
import os

from app_qt import widgets
from app_qt.base_tool import BaseTool

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


class MergeExcelTool(BaseTool):
    name = "Merge Excel"
    description = "Fill data from child files (by employee ID) into a master file."
    icon = "📊"
    category = "Files & Documents"
    order = 10
    action_label = "Merge data"

    def build_body(self, parent):
        widgets.section_label(parent, "Source & target")
        self._target = widgets.file_row(
            parent, "Target file (.xlsx) — column A holds employee IDs", mode="file")
        self._folder = widgets.file_row(
            parent, "Folder of child files (filenames start with the employee ID)", mode="folder")

        widgets.section_label(parent, "Settings")
        self._target_sheet = widgets.text_row(
            parent, "Sheet in target file (blank = first sheet)")
        self._child_sheet = widgets.text_row(
            parent, "Sheet in child file (blank = first sheet)")
        self._child_row = widgets.text_row(
            parent, "Data row in child file", placeholder="2")
        self._child_col = widgets.text_row(
            parent, "Start column to read in child file (e.g. A)", placeholder="A")
        self._target_col = widgets.text_row(
            parent, "Start column to write in target file (e.g. B)", placeholder="B")

        widgets.hint(
            parent,
            "💡 Example: ID NV001 → child file NV001_Report.xlsx or NV001.xlsx.\n"
            "Only fills rows that have an ID in column A but no data in the following columns.\n"
            "⚠ Back up the target file first — data is written directly.")

    def run(self):
        if not _OPENPYXL_OK:
            self.error("Missing library", "openpyxl is required:\n  pip install openpyxl")
            return

        target_path = self._target.get().strip()
        folder_path = self._folder.get().strip()
        if not target_path or not os.path.isfile(target_path):
            self.error("Error", "Please choose a valid target file.")
            return
        if not folder_path or not os.path.isdir(folder_path):
            self.error("Error", "Please choose the child-files folder.")
            return

        target_sheet = self._target_sheet.get().strip() or None
        child_sheet = self._child_sheet.get().strip() or None
        try:
            child_row = int(self._child_row.get().strip() or "2")
            if child_row < 1:
                raise ValueError
        except ValueError:
            self.error("Error", "The data row must be a positive integer (e.g. 2).")
            return
        try:
            child_col = column_index_from_string(self._child_col.get().strip().upper() or "A")
            target_col = column_index_from_string(self._target_col.get().strip().upper() or "B")
        except Exception:
            self.error("Error", "Invalid column — use A, B, C, AB, …")
            return

        try:
            filled, skipped, not_found, errors = _do_merge(
                target_path, folder_path, target_sheet, child_sheet,
                child_row, child_col, target_col)
        except Exception as exc:
            self.error("Error", f"Something went wrong:\n{exc}")
            return

        lines = [f"✅ Filled: {filled} rows."]
        if skipped:
            lines.append(f"⏭ Skipped (already has data): {skipped} rows.")
        if not_found:
            preview = ", ".join(not_found[:5])
            extra = f" (+{len(not_found) - 5} more)" if len(not_found) > 5 else ""
            lines.append(f"⚠ Child files not found: {preview}{extra}")
        if errors:
            preview = "; ".join(errors[:3])
            extra = f" (+{len(errors) - 3} more errors)" if len(errors) > 3 else ""
            lines.append(f"❌ Child-file read errors: {preview}{extra}")
        self.info("Merge result", "\n".join(lines))


# --------------------------------------------------------------- logic (thuần)
def _msnv_matches(filename, msnv):
    if not filename.startswith(msnv):
        return False
    rest = filename[len(msnv):]
    return not rest or rest[0] in ('_', '-', ' ', '.')


def _do_merge(target_path, folder_path, target_sheet, child_sheet,
              child_data_row, child_start_col, target_start_col):
    wb = openpyxl.load_workbook(target_path)
    ws = wb[target_sheet] if (target_sheet and target_sheet in wb.sheetnames) else wb.active
    max_col = ws.max_column

    child_files = {}
    for fname in os.listdir(folder_path):
        if fname.lower().endswith(('.xlsx', '.xlsm')):
            child_files[fname] = os.path.join(folder_path, fname)

    filled = skipped = 0
    not_found, errors = [], []

    for row in ws.iter_rows(min_row=2):
        msnv = str(row[0].value or "").strip()
        if not msnv:
            continue
        target_cells = row[target_start_col - 1:]
        if any(c.value is not None and str(c.value).strip() for c in target_cells):
            skipped += 1
            continue
        child_path = next(
            (path for fname, path in child_files.items() if _msnv_matches(fname, msnv)),
            None)
        if child_path is None:
            not_found.append(msnv)
            continue
        try:
            wb_child = openpyxl.load_workbook(child_path, data_only=True)
            ws_child = (wb_child[child_sheet]
                        if (child_sheet and child_sheet in wb_child.sheetnames)
                        else wb_child.active)
            child_rows = list(ws_child.iter_rows(
                min_row=child_data_row, max_row=child_data_row,
                min_col=child_start_col, values_only=True))
            wb_child.close()
        except Exception as exc:
            errors.append(f"{os.path.basename(child_path)}: {exc}")
            continue
        if not child_rows or not child_rows[0]:
            errors.append(f"{os.path.basename(child_path)}: row {child_data_row} is empty")
            continue
        values = child_rows[0]
        row_num = row[0].row
        for i, val in enumerate(values):
            col_num = target_start_col + i
            if col_num > max_col:
                break
            ws.cell(row=row_num, column=col_num, value=val)
        filled += 1

    wb.save(target_path)
    wb.close()
    return filled, skipped, not_found, errors
